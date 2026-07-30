"""Safe workbook I/O primitives shared by Step 3 services."""

from __future__ import annotations

import hashlib
import os
import shutil
import tempfile
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Any, Iterator
from uuid import uuid4

from openpyxl import load_workbook

from .models import WorkbookFingerprint


SUPPORTED_EXTENSIONS = frozenset({".xlsx", ".xlsm"})


class WorkbookError(RuntimeError):
    pass


class WorkbookChangedError(WorkbookError):
    pass


class WorkbookLockedError(WorkbookError):
    pass


def ensure_supported_workbook(path: str | Path) -> Path:
    candidate = Path(path)
    suffix = candidate.suffix.casefold()
    if suffix == ".xls":
        raise WorkbookError("Định dạng .xls không được hỗ trợ; hãy dùng .xlsx/.xlsm.")
    if suffix not in SUPPORTED_EXTENSIONS:
        raise WorkbookError("File Excel phải có định dạng .xlsx hoặc .xlsm.")
    return candidate


def workbook_fingerprint(path: str | Path) -> WorkbookFingerprint:
    candidate = Path(path)
    stat = candidate.stat()
    digest = hashlib.sha256()
    with candidate.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return WorkbookFingerprint(stat.st_size, stat.st_mtime_ns, digest.hexdigest())


class WorkbookGateway:
    def load(
        self,
        path: str | Path,
        *,
        read_only: bool = False,
        data_only: bool = False,
    ) -> Any:
        candidate = ensure_supported_workbook(path)
        return load_workbook(
            candidate,
            read_only=read_only,
            data_only=data_only,
            keep_vba=candidate.suffix.casefold() == ".xlsm",
            keep_links=True,
        )

    def fingerprint(self, path: str | Path) -> WorkbookFingerprint:
        return workbook_fingerprint(path)

    def assert_unchanged(
        self,
        path: str | Path,
        expected: WorkbookFingerprint,
        *,
        label: str = "Workbook",
    ) -> None:
        current = self.fingerprint(path)
        if current != expected:
            raise WorkbookChangedError(
                f"{label} đã thay đổi sau khi phân tích; vui lòng đọc lại."
            )

    def copy(
        self, source: str | Path, destination: str | Path
    ) -> Path:
        source_path = ensure_supported_workbook(source)
        destination_path = Path(destination)
        destination_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_path, destination_path)
        return destination_path

    def source_snapshot(
        self, source: str | Path, temp_dir: str | Path
    ) -> Path:
        source_path = ensure_supported_workbook(source)
        directory = Path(temp_dir)
        directory.mkdir(parents=True, exist_ok=True)
        destination = directory / (
            f"{source_path.stem}.{uuid4().hex}.snapshot{source_path.suffix}"
        )
        return self.copy(source_path, destination)

    @staticmethod
    def cleanup_source_snapshots(
        source: str | Path, temp_dir: str | Path
    ) -> int:
        source_path = ensure_supported_workbook(source)
        directory = Path(temp_dir)
        if not directory.is_dir():
            return 0
        removed = 0
        pattern = f"{source_path.stem}.*.snapshot{source_path.suffix}"
        for candidate in directory.glob(pattern):
            if candidate.is_file():
                candidate.unlink(missing_ok=True)
                removed += 1
        return removed

    def save(self, workbook: Any, path: str | Path) -> None:
        workbook.save(Path(path))

    def verify_openable(self, path: str | Path) -> None:
        workbook = self.load(path, read_only=True, data_only=False)
        workbook.close()

    def atomic_replace(
        self,
        working_path: str | Path,
        target_path: str | Path,
        *,
        expected: WorkbookFingerprint | None = None,
    ) -> WorkbookFingerprint:
        working = Path(working_path)
        target = Path(target_path)
        if expected is not None:
            self.assert_unchanged(target, expected, label="File BK")
        try:
            os.replace(working, target)
        except PermissionError as exc:
            raise WorkbookLockedError(
                "Không thể thay file BK; file có thể đang được mở trong Excel."
            ) from exc
        except OSError as exc:
            raise WorkbookError(f"Không thể thay nguyên tử file BK: {exc}") from exc
        return self.fingerprint(target)


class ExcelBackupService:
    def __init__(
        self,
        backup_dir: str | Path,
        *,
        working_dir: str | Path | None = None,
    ) -> None:
        self.backup_dir = Path(backup_dir)
        self.working_dir = Path(working_dir) if working_dir else None

    @staticmethod
    def _token(run_id: int | str | None) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        return f"{timestamp}_{run_id or uuid4().hex[:8]}"

    def create_backup(
        self, target: str | Path, *, run_id: int | str | None = None
    ) -> Path:
        target_path = ensure_supported_workbook(target)
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        backup = self.backup_dir / (
            f"{target_path.stem}_latest{target_path.suffix}"
        )
        temporary = self.backup_dir / (
            f".{target_path.stem}_{self._token(run_id)}.backup"
            f"{target_path.suffix}"
        )
        try:
            shutil.copy2(target_path, temporary)
            workbook = load_workbook(
                temporary,
                read_only=True,
                data_only=False,
                keep_vba=target_path.suffix.casefold() == ".xlsm",
                keep_links=True,
            )
            workbook.close()
            os.replace(temporary, backup)
            self._cleanup_legacy_backups(target_path, keep=backup)
        finally:
            temporary.unlink(missing_ok=True)
        return backup

    def create_working_copy(
        self, target: str | Path, *, run_id: int | str | None = None
    ) -> Path:
        target_path = ensure_supported_workbook(target)
        directory = self.working_dir or target_path.parent
        directory.mkdir(parents=True, exist_ok=True)
        working = directory / (
            f".{target_path.stem}_{self._token(run_id)}.working{target_path.suffix}"
        )
        shutil.copy2(target_path, working)
        return working

    def cleanup_working_copies(self, target: str | Path) -> int:
        target_path = ensure_supported_workbook(target)
        directory = self.working_dir or target_path.parent
        if not directory.is_dir():
            return 0
        removed = 0
        pattern = (
            f".{target_path.stem}_*.working{target_path.suffix}"
        )
        for candidate in directory.glob(pattern):
            if candidate.is_file():
                candidate.unlink(missing_ok=True)
                removed += 1
        return removed

    def _cleanup_legacy_backups(self, target: Path, *, keep: Path) -> None:
        pattern = f"{target.stem}_*{target.suffix}"
        for candidate in self.backup_dir.glob(pattern):
            if candidate == keep or not candidate.is_file():
                continue
            candidate.unlink(missing_ok=True)


class ExcelLockService:
    """Best-effort Windows lock probe plus parent-directory write probe."""

    def ensure_readable(self, path: str | Path) -> None:
        try:
            with Path(path).open("rb") as handle:
                handle.read(1)
        except (PermissionError, OSError) as exc:
            raise WorkbookLockedError(f"Không thể đọc file Excel: {exc}") from exc

    def ensure_writable(self, path: str | Path) -> None:
        candidate = Path(path)
        self.ensure_readable(candidate)
        try:
            descriptor, probe = tempfile.mkstemp(
                prefix=".excel-write-probe-", dir=candidate.parent
            )
            os.close(descriptor)
            Path(probe).unlink()
        except (PermissionError, OSError) as exc:
            raise WorkbookLockedError(
                "Không có quyền tạo file tạm cạnh file BK."
            ) from exc

    @contextmanager
    def acquire(self, path: str | Path) -> Iterator[None]:
        candidate = Path(path)
        self.ensure_writable(candidate)
        handle = None
        locked = False
        try:
            handle = candidate.open("r+b")
            if os.name == "nt":
                import msvcrt

                try:
                    handle.seek(0)
                    msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
                    locked = True
                except OSError as exc:
                    raise WorkbookLockedError(
                        "File BK đang bị khóa bởi chương trình khác."
                    ) from exc
            yield
        finally:
            if handle is not None:
                if locked:
                    import msvcrt

                    try:
                        handle.seek(0)
                        msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
                    except OSError:
                        pass
                handle.close()
