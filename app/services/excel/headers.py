"""Unicode-tolerant, position-independent worksheet header discovery."""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from typing import Any, Iterable, Mapping

from openpyxl.utils import get_column_letter


_SPACE_RE = re.compile(r"\s+")
_PUNCT_RE = re.compile(r"[^0-9a-z]+")


def normalize_header(value: Any) -> str:
    """Return a stable comparison key for Vietnamese/English headers."""

    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value)).strip().casefold()
    text = text.replace("đ", "d")
    text = "".join(
        character
        for character in unicodedata.normalize("NFD", text)
        if unicodedata.category(character) != "Mn"
    )
    text = unicodedata.normalize("NFC", text)
    text = _PUNCT_RE.sub(" ", text)
    return _SPACE_RE.sub(" ", text).strip()


class HeaderResolutionError(ValueError):
    def __init__(
        self,
        message: str,
        *,
        missing: Iterable[str] = (),
        ambiguous: Iterable[str] = (),
    ) -> None:
        super().__init__(message)
        self.missing = tuple(missing)
        self.ambiguous = tuple(ambiguous)


@dataclass(frozen=True, slots=True)
class HeaderResolution:
    columns: dict[str, int]
    row_start: int
    row_end: int
    headers: dict[int, str]

    @property
    def header_row(self) -> int:
        return self.row_end

    def column_letter(self, field: str) -> str:
        return get_column_letter(self.columns[field])


class HeaderResolver:
    """Find a one-to-three row header band in the first 30 worksheet rows."""

    def __init__(self, *, max_rows: int = 30, max_band_height: int = 3) -> None:
        self.max_rows = max_rows
        self.max_band_height = max_band_height

    def resolve(
        self,
        worksheet: Any,
        fields: Mapping[str, str | Iterable[str]],
        *,
        required: Iterable[str] | None = None,
        max_rows: int | None = None,
    ) -> HeaderResolution:
        aliases = {
            field: self._alias_keys(field, values)
            for field, values in fields.items()
        }
        required_set = set(required if required is not None else fields)
        unknown = required_set.difference(fields)
        if unknown:
            raise KeyError(f"Trường header chưa khai báo: {', '.join(sorted(unknown))}")

        limit = min(max_rows or self.max_rows, int(worksheet.max_row or 0))
        max_column = int(worksheet.max_column or 0)
        if limit <= 0 or max_column <= 0:
            raise HeaderResolutionError(
                "Worksheet không có vùng header.",
                missing=sorted(required_set),
            )

        merged_values = self._merged_values(worksheet)
        candidates: list[tuple[int, int, int, dict[str, int], dict[int, str]]] = []
        for row_start in range(1, limit + 1):
            for height in range(1, self.max_band_height + 1):
                row_end = row_start + height - 1
                if row_end > limit:
                    break
                headers = {
                    column: self._combined_header(
                        worksheet, row_start, row_end, column, merged_values
                    )
                    for column in range(1, max_column + 1)
                }
                mapping, ambiguous = self._match_headers(headers, aliases)
                if ambiguous:
                    continue
                score = sum(field in mapping for field in required_set)
                if score:
                    candidates.append((score, row_start, row_end, mapping, headers))

        if not candidates:
            raise HeaderResolutionError(
                "Không nhận diện được header bắt buộc.",
                missing=sorted(required_set),
            )
        best_score = max(candidate[0] for candidate in candidates)
        best = [candidate for candidate in candidates if candidate[0] == best_score]
        complete = [candidate for candidate in best if best_score == len(required_set)]
        if complete:
            best = complete

        # Equivalent mappings from nested bands are harmless. Prefer the shortest,
        # then the earliest band. Different best mappings are unsafe.
        mapping_keys = {
            tuple(sorted(candidate[3].items()))
            for candidate in best
        }
        if len(mapping_keys) > 1:
            raise HeaderResolutionError(
                "Có nhiều vùng header phù hợp ngang nhau; không thể chọn an toàn.",
                ambiguous=sorted(required_set),
            )
        chosen = min(best, key=lambda item: (item[2] - item[1], item[2], item[1]))
        missing = sorted(required_set.difference(chosen[3]))
        if missing:
            raise HeaderResolutionError(
                "Thiếu header bắt buộc: " + ", ".join(missing),
                missing=missing,
            )
        return HeaderResolution(
            columns=chosen[3],
            row_start=chosen[1],
            row_end=chosen[2],
            headers=chosen[4],
        )

    @staticmethod
    def _alias_keys(
        field: str, values: str | Iterable[str]
    ) -> frozenset[str]:
        supplied = (values,) if isinstance(values, str) else tuple(values)
        return frozenset(
            key
            for key in (normalize_header(value) for value in (*supplied, field))
            if key
        )

    @staticmethod
    def _merged_values(worksheet: Any) -> dict[tuple[int, int], Any]:
        result: dict[tuple[int, int], Any] = {}
        ranges = getattr(getattr(worksheet, "merged_cells", None), "ranges", ())
        for merged_range in ranges:
            value = worksheet.cell(
                merged_range.min_row, merged_range.min_col
            ).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for column in range(
                    merged_range.min_col, merged_range.max_col + 1
                ):
                    result[(row, column)] = value
        return result

    @staticmethod
    def _combined_header(
        worksheet: Any,
        row_start: int,
        row_end: int,
        column: int,
        merged_values: Mapping[tuple[int, int], Any],
    ) -> str:
        parts: list[str] = []
        for row in range(row_start, row_end + 1):
            value = merged_values.get(
                (row, column), worksheet.cell(row, column).value
            )
            key = normalize_header(value)
            if key and (not parts or parts[-1] != key):
                parts.append(key)
        return " ".join(parts)

    @staticmethod
    def _match_headers(
        headers: Mapping[int, str],
        aliases: Mapping[str, frozenset[str]],
    ) -> tuple[dict[str, int], set[str]]:
        mapping: dict[str, int] = {}
        ambiguous: set[str] = set()
        for field, field_aliases in aliases.items():
            exact = [
                column
                for column, header in headers.items()
                if header and header in field_aliases
            ]
            matches = exact or [
                column
                for column, header in headers.items()
                if header
                and any(
                    header.endswith(f" {alias}")
                    or header.startswith(f"{alias} ")
                    for alias in field_aliases
                )
            ]
            if len(matches) == 1:
                mapping[field] = matches[0]
            elif len(matches) > 1:
                ambiguous.add(field)
        return mapping, ambiguous

