"""Đồng bộ toàn bộ dữ liệu sheet Hàng ngày vào các cột nghiệp vụ của BK."""

from __future__ import annotations

import copy
import hashlib
import json
from collections import defaultdict
from collections.abc import Callable, Mapping, Sequence
from pathlib import Path
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from app.services.file_stability import FileStabilityChecker

from .headers import HeaderResolution, HeaderResolutionError, HeaderResolver
from .models import (
    ConflictType,
    ExcelOperation,
    ExcelRunStatus,
    MonthCandidate,
    ResolutionAction,
    SourceSheetCandidate,
    SyncAction,
    SyncActionType,
    SyncConflict,
    SyncPlan,
    SyncResolution,
    SyncResult,
    SyncRow,
    resolution_map,
)
from .resolvers import MonthSheetService, YearResolver
from .workbook import (
    ExcelBackupService,
    ExcelLockService,
    WorkbookGateway,
    ensure_supported_workbook,
)


ProgressCallback = Callable[[str], None] | None

SYNC_FIELDS: tuple[str, ...] = (
    "sqt",
    "closing_date",
    "container",
    "weight",
    "cargo_type",
    "closing_place",
    "vessel",
    "departure_date",
    "estimated_delivery",
    "recipient",
    "sea_transport",
    "transport",
)

SOURCE_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "sqt": ("SQT PM", "SQT", "Số thứ tự PM"),
    "closing_date": ("Ngày Đóng", "Ngày đóng hàng"),
    "container": ("Số Container", "Container", "Số cont"),
    "weight": ("Số tấn", "Trọng lượng", "Số lượng tấn"),
    "cargo_type": ("Loại hàng", "Tên hàng"),
    "closing_place": ("Nơi đóng", "Địa điểm đóng"),
    "vessel": ("Tên tàu", "Tàu"),
    "departure_date": ("Ngày chạy", "Ngày tàu chạy"),
    "estimated_delivery": ("Dự kiến giao", "Ngày dự kiến giao"),
    "recipient": ("Người nhận", "Khách hàng"),
    "sea_transport": ("VT biển", "Vận tải biển"),
    "transport": ("Vận chuyển", "Đơn vị vận chuyển", "VT bộ"),
}

TARGET_EXPECTED_COLUMNS: dict[str, int] = {
    **{field: index for index, field in enumerate(SYNC_FIELDS[:11], 1)},
    "transport": 16,
}
SYNC_TARGET_COLUMNS = frozenset(TARGET_EXPECTED_COLUMNS.values())


class DailySyncError(RuntimeError):
    pass


def parse_sqt(value: Any) -> int | None:
    if isinstance(value, bool) or value is None:
        return None
    if type(value) is int:
        return value if value > 0 else None
    if isinstance(value, float):
        return int(value) if value.is_integer() and value > 0 else None
    if isinstance(value, str):
        text = "".join(value.split())
        if text.isdigit():
            parsed = int(text)
            return parsed if parsed > 0 else None
    return None


def _stable_id(*parts: Any) -> str:
    payload = json.dumps(parts, ensure_ascii=False, default=str, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:24]


def _progress(callback: ProgressCallback, message: str) -> None:
    if callback is not None:
        callback(message)


def _settings_value(settings: Any, name: str, default: Any = None) -> Any:
    return getattr(settings, name, default) if settings is not None else default


class DailySyncService:
    def __init__(
        self,
        settings: Any | None = None,
        *,
        daily_path: str | Path | None = None,
        bk_path: str | Path | None = None,
        temp_dir: str | Path | None = None,
        backup_dir: str | Path | None = None,
        gateway: WorkbookGateway | None = None,
        lock_service: ExcelLockService | None = None,
        stability_checker: FileStabilityChecker | None = None,
        header_resolver: HeaderResolver | None = None,
        month_service: MonthSheetService | None = None,
        year_resolver: YearResolver | None = None,
        run_repository: Any | None = None,
    ) -> None:
        paths = getattr(settings, "paths", None)
        self.daily_path = Path(
            daily_path or _settings_value(settings, "daily_workbook_path", "")
        )
        self.bk_path = Path(
            bk_path or _settings_value(settings, "bk_workbook_path", "")
        )
        system_dir = getattr(paths, "system_dir", Path("Output") / "_system")
        self.temp_dir = Path(
            temp_dir
            or getattr(paths, "excel_temp_dir", system_dir / "Excel" / "Temp")
        )
        self.backup_dir = Path(
            backup_dir
            or getattr(paths, "excel_backup_dir", system_dir / "Excel" / "Backup")
        )
        self.gateway = gateway or WorkbookGateway()
        self.lock_service = lock_service or ExcelLockService()
        self.stability_checker = stability_checker or FileStabilityChecker()
        self.headers = header_resolver or HeaderResolver()
        self.months = month_service or MonthSheetService()
        self.years = year_resolver or YearResolver()
        self.run_repository = run_repository
        self.backups = ExcelBackupService(self.backup_dir)

    def cleanup_stale_files(self) -> int:
        """Dọn đúng các artifact do luồng Excel của ứng dụng tạo ra."""

        removed = 0
        if self.bk_path.suffix.casefold() in {".xlsx", ".xlsm"}:
            removed += self.backups.cleanup_working_copies(self.bk_path)
        if self.daily_path.suffix.casefold() in {".xlsx", ".xlsm"}:
            removed += self.gateway.cleanup_source_snapshots(
                self.daily_path, self.temp_dir
            )
        return removed

    def source_sheet_candidates(self) -> list[SourceSheetCandidate]:
        source = ensure_supported_workbook(self.daily_path)
        if not source.is_file():
            raise DailySyncError(f"Không tìm thấy file Hàng ngày: {source}")
        self.lock_service.ensure_readable(source)
        workbook = self.gateway.load(source, read_only=True)
        try:
            source_sheets = self.months.daily_sheets(workbook.sheetnames)
        finally:
            workbook.close()
        if not source_sheets:
            raise DailySyncError(
                "File Hàng ngày không có sheet Tháng 1…Tháng 12."
            )
        return [
            SourceSheetCandidate(month=month, source_sheet=sheet_name)
            for month, sheet_name in sorted(source_sheets.items())
        ]

    def analyze(
        self,
        progress_callback: ProgressCallback = None,
        *,
        source_sheet_name: str | None = None,
    ) -> SyncPlan:
        source = ensure_supported_workbook(self.daily_path)
        target = ensure_supported_workbook(self.bk_path)
        if not source.is_file():
            raise DailySyncError(f"Không tìm thấy file Hàng ngày: {source}")
        if not target.is_file():
            raise DailySyncError(f"Không tìm thấy file BK: {target}")

        run_id = self._create_run(source, target)
        source_book = None
        target_book = None
        try:
            _progress(progress_callback, "Đang chờ file Hàng ngày ổn định…")
            self.stability_checker.wait(source)
            self.lock_service.ensure_readable(source)
            self.lock_service.ensure_readable(target)
            source_fingerprint = self.gateway.fingerprint(source)
            target_fingerprint = self.gateway.fingerprint(target)

            # Workbook nguồn nhỏ và cần truy cập header/cell nhiều lần. Chế độ
            # normal nhanh hơn rất nhiều so với random access trên ReadOnlyWorksheet.
            source_book = self.gateway.load(source, read_only=False)
            target_book = self.gateway.load(target, read_only=False)
            source_sheets = self.months.daily_sheets(source_book.sheetnames)
            if not source_sheets:
                raise DailySyncError(
                    "File Hàng ngày không có sheet Tháng 1…Tháng 12."
                )
            if source_sheet_name is not None:
                source_sheets = {
                    month: sheet_name
                    for month, sheet_name in source_sheets.items()
                    if sheet_name == source_sheet_name
                }
                if not source_sheets:
                    raise DailySyncError(
                        f"Không tìm thấy sheet nguồn {source_sheet_name!r} "
                        "trong file Hàng ngày."
                    )

            target_sheets = self._target_sheets(target_book.sheetnames)
            target_years = sorted(
                {
                    year
                    for sheets in target_sheets.values()
                    for year, _name in sheets
                }
            )
            if not target_years:
                raise DailySyncError("File BK không có sheet tháng TMM YY.")

            rows_by_month: dict[int, list[SyncRow]] = {}
            rows_by_target: dict[str, list[SyncRow]] = {}
            actions_by_target: dict[str, list[SyncAction]] = {}
            candidates: list[MonthCandidate] = []
            conflicts: list[SyncConflict] = []
            invalid_rows: list[int] = []
            source_row_count = 0

            for month, current_source_name in sorted(source_sheets.items()):
                _progress(
                    progress_callback,
                    f"Đang đối chiếu toàn bộ {current_source_name}…",
                )
                source_sheet = source_book[current_source_name]
                source_header = self.headers.resolve(
                    source_sheet, SOURCE_HEADER_ALIASES, required=SYNC_FIELDS
                )
                source_rows, current_invalid = self._source_rows(
                    source_sheet,
                    source_header,
                    month=month,
                )
                source_row_count += len(source_rows)
                invalid_rows.extend(current_invalid)

                existing_targets = list(target_sheets.get(month, ()))
                target_options = (
                    existing_targets
                    if existing_targets
                    else [
                        (year, self.months.target_name(month, year))
                        for year in target_years
                    ]
                )
                existing_names = {name for _year, name in existing_targets}
                for target_year, target_name in target_options:
                    target_rows: list[SyncRow] = []
                    if target_name in existing_names:
                        target_sheet = target_book[target_name]
                        target_header = self._resolve_target_headers(target_sheet)
                        target_rows = self._target_rows(
                            target_sheet, target_header, month=month
                        )
                    actions, action_conflicts = self._build_actions(
                        source_rows,
                        target_rows,
                        source_fingerprint=source_fingerprint.sha256,
                        target_sheet=target_name,
                        target_worksheet=(
                            target_book[target_name]
                            if target_name in existing_names
                            else None
                        ),
                    )
                    actions_by_target[target_name] = actions
                    inserts = [
                        action.source
                        for action in actions
                        if action.action is SyncActionType.INSERT
                        and action.source is not None
                    ]
                    rows_by_target[target_name] = inserts
                    rows_by_month.setdefault(month, inserts)
                    conflicts.extend(action_conflicts)
                    counts = self._action_counts(actions)
                    candidates.append(
                        MonthCandidate(
                            month=month,
                            year=target_year,
                            source_sheet=current_source_name,
                            target_sheet=target_name,
                            new_row_count=counts[SyncActionType.INSERT],
                            update_count=counts[SyncActionType.UPDATE],
                            unchanged_count=counts[SyncActionType.UNCHANGED],
                            target_only_count=counts[SyncActionType.TARGET_ONLY],
                            invalid_count=len(current_invalid),
                            last_sqt=max(
                                (row.sqt for row in target_rows), default=0
                            ),
                            match_count=(
                                counts[SyncActionType.UPDATE]
                                + counts[SyncActionType.UNCHANGED]
                            ),
                        )
                    )

            source_book.close()
            source_book = None
            self.gateway.assert_unchanged(
                source, source_fingerprint, label="File Hàng ngày"
            )
            target_book.close()
            target_book = None

            selected_month = candidates[0].month if len(candidates) == 1 else None
            selected_target = (
                candidates[0].target_sheet if len(candidates) == 1 else None
            )
            if len(candidates) > 1:
                conflicts.append(
                    SyncConflict(
                        conflict_id=_stable_id(
                            "sync-month",
                            source_fingerprint.sha256,
                            [candidate.target_sheet for candidate in candidates],
                        ),
                        conflict_type=ConflictType.TARGET_MONTH_AMBIGUOUS,
                        message=(
                            "Có nhiều sheet BK phù hợp; hãy chọn một sheet đích."
                        ),
                        allowed_actions=(
                            ResolutionAction.SELECT_MONTH,
                            ResolutionAction.CANCEL_ALL,
                        ),
                        details={
                            "months": [candidate.month for candidate in candidates],
                            "sheet_candidates": candidates,
                        },
                    )
                )
            selected_year = (
                candidates[0].year
                if len(candidates) == 1
                else target_years[0]
                if len(target_years) == 1
                else None
            )
            plan = SyncPlan(
                source_path=source,
                target_path=target,
                source_fingerprint=source_fingerprint,
                target_fingerprint=target_fingerprint,
                source_year=None,
                target_year=selected_year,
                month_candidates=candidates,
                rows_by_month=rows_by_month,
                rows_by_target=rows_by_target,
                actions_by_target=actions_by_target,
                invalid_rows=invalid_rows,
                conflicts=conflicts,
                selected_month=selected_month,
                selected_target_sheet=selected_target,
                run_id=run_id,
            )
            status = (
                ExcelRunStatus.WAITING_USER
                if plan.requires_user_input
                else ExcelRunStatus.NO_CHANGES
            )
            self._update_run(
                run_id,
                status=status,
                source_fingerprint=source_fingerprint,
                target_fingerprint_before=target_fingerprint,
                sheet_name=plan.selected_sheet,
                total_items=source_row_count + len(invalid_rows),
                conflict_count=plan.conflict_count,
            )
            return plan
        except Exception as exc:
            self._finish_failed(run_id, exc)
            raise
        finally:
            if source_book is not None:
                source_book.close()
            if target_book is not None:
                target_book.close()

    def apply(
        self,
        plan: SyncPlan,
        resolutions: Mapping[str, Any] | Sequence[SyncResolution] | None = None,
        progress_callback: ProgressCallback = None,
    ) -> SyncResult:
        resolved = resolution_map(resolutions)
        try:
            candidate = self._selected_candidate(plan, resolved)
        except Exception as exc:
            self._finish_failed(plan.run_id, exc)
            raise
        if candidate is None:
            exc = DailySyncError("Chưa chọn sheet BK cần đồng bộ.")
            self._finish_failed(plan.run_id, exc)
            raise exc

        target_name = candidate.target_sheet
        blocking = [
            conflict
            for conflict in plan.conflicts
            if conflict.conflict_type is ConflictType.SYNC_GROUP_COUNT_MISMATCH
            and conflict.details.get("target_sheet") == target_name
        ]
        if blocking:
            exc = DailySyncError(
                "Không thể đồng bộ vì số dòng của cùng một SQT không khớp "
                "giữa file Hàng ngày và BK."
            )
            self._finish_failed(plan.run_id, exc)
            raise exc

        actions = list(plan.actions_by_target.get(target_name, ()))
        counts = self._action_counts(actions)
        inserted = counts[SyncActionType.INSERT]
        updated = counts[SyncActionType.UPDATE]
        unchanged = counts[SyncActionType.UNCHANGED]
        target_only = counts[SyncActionType.TARGET_ONLY]
        invalid = candidate.invalid_count
        if not inserted and not updated:
            result = SyncResult(
                status=ExcelRunStatus.NO_CHANGES,
                target_path=plan.target_path,
                sheet_name=target_name,
                unchanged_rows=unchanged,
                target_only_rows=target_only,
                invalid_rows=invalid,
                skipped_rows=invalid,
                conflict_count=len(blocking),
                fingerprint_before=plan.target_fingerprint,
                fingerprint_after=plan.target_fingerprint,
                run_id=plan.run_id,
                message=(
                    "Dữ liệu A–K/P đã đồng bộ; không có ô nào cần ghi."
                ),
            )
            self._finish_result(result)
            return result

        self._update_run(plan.run_id, status=ExcelRunStatus.APPLYING)
        _progress(progress_callback, "Đang kiểm tra file chưa thay đổi…")
        try:
            self.gateway.assert_unchanged(
                plan.source_path,
                plan.source_fingerprint,
                label="File Hàng ngày",
            )
            self.gateway.assert_unchanged(
                plan.target_path,
                plan.target_fingerprint,
                label="File BK",
            )
        except Exception as exc:
            self._finish_failed(plan.run_id, exc)
            raise

        backup_path: Path | None = None
        working_path: Path | None = None
        try:
            with self.lock_service.acquire(plan.target_path):
                pass
            self.gateway.assert_unchanged(
                plan.target_path, plan.target_fingerprint, label="File BK"
            )
            working_path = self.backups.create_working_copy(
                plan.target_path, run_id=plan.run_id
            )
            workbook = self.gateway.load(working_path, read_only=False)
            expected_cells: dict[tuple[int, int], Any] = {}
            protected_cells: dict[tuple[int, int], Any] = {}
            try:
                created = target_name not in workbook.sheetnames
                if created:
                    worksheet = self._create_month_sheet(
                        workbook,
                        candidate.month,
                        int(candidate.year),
                        target_name,
                    )
                else:
                    worksheet = workbook[target_name]
                header = self._resolve_target_headers(worksheet)

                for action in actions:
                    if (
                        action.action is not SyncActionType.UPDATE
                        or action.source is None
                        or action.target_row is None
                    ):
                        continue
                    self._write_sync_values(
                        worksheet,
                        action.target_row,
                        action.source.values,
                        expected_cells,
                    )
                    for column, value in action.protected_values.items():
                        protected_cells[(action.target_row, column)] = value

                insert_actions = [
                    action
                    for action in actions
                    if action.action is SyncActionType.INSERT
                    and action.source is not None
                ]
                append_at = self._last_data_row(worksheet, header) + 1
                template_row = max(header.row_end + 1, append_at - 1)
                max_style_column = self._actual_max_column(worksheet)
                for offset, action in enumerate(insert_actions):
                    target_row = append_at + offset
                    self._copy_row_style(
                        worksheet,
                        template_row,
                        target_row,
                        max_column=max_style_column,
                    )
                    self._write_sync_values(
                        worksheet,
                        target_row,
                        action.source.values,
                        expected_cells,
                    )
                from .payment_sync import (
                    find_summary_start,
                    refresh_bk_summary_formulas,
                )

                if find_summary_start(worksheet) is not None:
                    refresh_bk_summary_formulas(worksheet)
                self.gateway.save(workbook, working_path)
            finally:
                workbook.close()

            _progress(progress_callback, "Đang kiểm tra bản BK mới…")
            self._verify_saved_sync(
                working_path,
                target_name,
                expected_cells,
                protected_cells,
            )
            _progress(progress_callback, "Đang cập nhật backup gần nhất…")
            backup_path = self.backups.create_backup(plan.target_path)
            after = self.gateway.atomic_replace(
                working_path,
                plan.target_path,
                expected=plan.target_fingerprint,
            )
            working_path = None
            result = SyncResult(
                status=ExcelRunStatus.SUCCEEDED,
                target_path=plan.target_path,
                sheet_name=target_name,
                added_rows=inserted,
                inserted_rows=inserted,
                updated_rows=updated,
                unchanged_rows=unchanged,
                target_only_rows=target_only,
                invalid_rows=invalid,
                skipped_rows=invalid,
                conflict_count=len(blocking),
                backup_path=backup_path,
                fingerprint_before=plan.target_fingerprint,
                fingerprint_after=after,
                run_id=plan.run_id,
                message=(
                    f"Đã cập nhật {updated} dòng, thêm {inserted} dòng, "
                    f"giữ {target_only} dòng chỉ có ở BK."
                ),
            )
            self._finish_result(result)
            return result
        except Exception as exc:
            self._finish_failed(plan.run_id, exc)
            raise
        finally:
            if working_path is not None and working_path.exists():
                working_path.unlink()

    def cancel(self, plan: SyncPlan) -> None:
        if self.run_repository is not None and plan.run_id is not None:
            self.run_repository.finish_run(
                plan.run_id,
                status=ExcelRunStatus.CANCELLED,
                total_items=(
                    plan.insert_count
                    + plan.update_count
                    + plan.unchanged_count
                    + plan.target_only_count
                    + plan.invalid_count
                ),
                changed_items=0,
                skipped_items=plan.invalid_count,
                conflict_count=plan.conflict_count,
            )

    def _source_rows(
        self,
        worksheet: Worksheet,
        header: HeaderResolution,
        *,
        month: int,
    ) -> tuple[list[SyncRow], list[int]]:
        rows: list[SyncRow] = []
        invalid_rows: list[int] = []
        columns = tuple(header.columns[field] for field in SYNC_FIELDS)
        for row_index in self._populated_rows(
            worksheet, columns, min_row=header.row_end + 1
        ):
            values = tuple(
                worksheet.cell(row_index, header.columns[field]).value
                for field in SYNC_FIELDS
            )
            sqt = parse_sqt(values[0])
            if sqt is None:
                invalid_rows.append(row_index)
                continue
            rows.append(
                SyncRow(
                    source_sheet=worksheet.title,
                    source_row=row_index,
                    month=month,
                    sqt=sqt,
                    values=values,
                    duplicate_key=(sqt,),
                )
            )
        return rows, invalid_rows

    def _target_rows(
        self,
        worksheet: Worksheet,
        header: HeaderResolution,
        *,
        month: int,
    ) -> list[SyncRow]:
        rows: list[SyncRow] = []
        columns = tuple(header.columns[field] for field in SYNC_FIELDS)
        for row_index in self._populated_rows(
            worksheet, columns, min_row=header.row_end + 1
        ):
            values = tuple(
                worksheet.cell(row_index, header.columns[field]).value
                for field in SYNC_FIELDS
            )
            sqt = parse_sqt(values[0])
            if sqt is None:
                continue
            rows.append(
                SyncRow(
                    source_sheet=worksheet.title,
                    source_row=row_index,
                    month=month,
                    sqt=sqt,
                    values=values,
                    duplicate_key=(sqt,),
                )
            )
        return rows

    def _build_actions(
        self,
        source_rows: Sequence[SyncRow],
        target_rows: Sequence[SyncRow],
        *,
        source_fingerprint: str,
        target_sheet: str,
        target_worksheet: Worksheet | None,
    ) -> tuple[list[SyncAction], list[SyncConflict]]:
        source_groups: dict[int, list[SyncRow]] = defaultdict(list)
        target_groups: dict[int, list[SyncRow]] = defaultdict(list)
        for row in source_rows:
            source_groups[row.sqt].append(row)
        for row in target_rows:
            target_groups[row.sqt].append(row)

        actions: list[SyncAction] = []
        conflicts: list[SyncConflict] = []
        seen_sqt: set[int] = set()
        for source in source_rows:
            if source.sqt in seen_sqt:
                continue
            seen_sqt.add(source.sqt)
            source_group = source_groups[source.sqt]
            target_group = target_groups.get(source.sqt, [])
            if not target_group:
                actions.extend(
                    SyncAction(
                        action=SyncActionType.INSERT,
                        sqt=item.sqt,
                        source=item,
                    )
                    for item in source_group
                )
                continue
            if len(source_group) != len(target_group):
                conflicts.append(
                    SyncConflict(
                        conflict_id=_stable_id(
                            "sync-count",
                            source_fingerprint,
                            target_sheet,
                            source.sqt,
                            len(source_group),
                            len(target_group),
                        ),
                        conflict_type=ConflictType.SYNC_GROUP_COUNT_MISMATCH,
                        message=(
                            f"SQT {source.sqt}: nguồn có {len(source_group)} dòng, "
                            f"BK có {len(target_group)} dòng."
                        ),
                        source_sheet=source.source_sheet,
                        source_row=source.source_row,
                        sqt=source.sqt,
                        allowed_actions=(ResolutionAction.CANCEL_ALL,),
                        details={
                            "target_sheet": target_sheet,
                            "source_count": len(source_group),
                            "target_count": len(target_group),
                        },
                    )
                )
                continue
            for source_item, target_item in zip(source_group, target_group):
                action_type = (
                    SyncActionType.UNCHANGED
                    if source_item.values == target_item.values
                    else SyncActionType.UPDATE
                )
                actions.append(
                    SyncAction(
                        action=action_type,
                        sqt=source_item.sqt,
                        source=source_item,
                        target_row=target_item.source_row,
                        target_values=target_item.values,
                        protected_values=self._protected_values(
                            target_worksheet,
                            target_item.source_row,
                        ),
                    )
                )

        source_sqts = set(source_groups)
        for target in target_rows:
            if target.sqt in source_sqts:
                continue
            actions.append(
                SyncAction(
                    action=SyncActionType.TARGET_ONLY,
                    sqt=target.sqt,
                    target_row=target.source_row,
                    target_values=target.values,
                )
            )
        source_actions = sorted(
            (action for action in actions if action.source is not None),
            key=lambda action: int(action.source.source_row),
        )
        target_only_actions = sorted(
            (action for action in actions if action.source is None),
            key=lambda action: int(action.target_row or 0),
        )
        return [*source_actions, *target_only_actions], conflicts

    @staticmethod
    def _protected_values(
        worksheet: Worksheet | None, row: int
    ) -> dict[int, Any]:
        if worksheet is None:
            return {}
        return {
            column: cell.value
            for (cell_row, column), cell in worksheet._cells.items()
            if cell_row == row
            and column not in SYNC_TARGET_COLUMNS
            and cell.value is not None
        }

    @staticmethod
    def _action_counts(
        actions: Sequence[SyncAction],
    ) -> dict[SyncActionType, int]:
        return {
            kind: sum(action.action is kind for action in actions)
            for kind in SyncActionType
        }

    def _resolve_target_headers(self, worksheet: Worksheet) -> HeaderResolution:
        resolution = self.headers.resolve(
            worksheet, SOURCE_HEADER_ALIASES, required=SYNC_FIELDS
        )
        wrong = {
            field: (resolution.columns[field], expected)
            for field, expected in TARGET_EXPECTED_COLUMNS.items()
            if resolution.columns[field] != expected
        }
        if wrong:
            details = ", ".join(
                f"{field}: cột {actual}, cần {expected}"
                for field, (actual, expected) in wrong.items()
            )
            raise HeaderResolutionError(
                f"Header sheet BK không đúng vị trí A–K/P ({details})."
            )
        return resolution

    def _target_sheets(
        self, sheet_names: Sequence[str]
    ) -> dict[int, list[tuple[int, str]]]:
        result: dict[int, list[tuple[int, str]]] = {}
        seen: set[tuple[int, int]] = set()
        for name in sheet_names:
            parsed = self.months.parse_target_sheet(name)
            if parsed is None:
                continue
            month, year = parsed
            key = (month, year)
            if key in seen:
                raise DailySyncError(f"Có nhiều sheet BK cho tháng {month}/{year}.")
            seen.add(key)
            result.setdefault(month, []).append((year, name))
        return result

    @staticmethod
    def _populated_rows(
        worksheet: Worksheet,
        columns: Sequence[int],
        *,
        min_row: int,
    ) -> list[int]:
        wanted = set(columns)
        cells = getattr(worksheet, "_cells", {})
        return sorted(
            {
                row
                for (row, column), cell in cells.items()
                if row >= min_row
                and column in wanted
                and cell.value not in (None, "")
            }
        )

    @classmethod
    def _last_data_row(
        cls, worksheet: Worksheet, header: HeaderResolution
    ) -> int:
        rows = cls._populated_rows(
            worksheet,
            tuple(TARGET_EXPECTED_COLUMNS.values()),
            min_row=header.row_end + 1,
        )
        return rows[-1] if rows else header.row_end

    @staticmethod
    def _actual_max_column(worksheet: Worksheet) -> int:
        cells = getattr(worksheet, "_cells", {})
        used = [
            column
            for (_row, column), cell in cells.items()
            if cell.value is not None or cell.has_style
        ]
        return max(used, default=max(TARGET_EXPECTED_COLUMNS.values()))

    def _create_month_sheet(
        self, workbook: Any, month: int, year: int, target_name: str
    ) -> Worksheet:
        template_name = self.months.nearest_previous_template(
            workbook.sheetnames, month, year
        )
        if template_name is None:
            raise DailySyncError(
                f"Không có sheet tháng trước làm mẫu cho {target_name}."
            )
        template = workbook[template_name]
        header = self._resolve_target_headers(template)
        data_row = self._last_data_row(template, header)
        if data_row <= header.row_end:
            raise DailySyncError(f"Sheet mẫu {template_name} không có dòng dữ liệu.")

        worksheet = workbook.create_sheet(target_name)
        for attribute in (
            "freeze_panes",
            "sheet_format",
            "sheet_properties",
            "page_margins",
            "page_setup",
            "print_options",
            "auto_filter",
            "data_validations",
            "conditional_formatting",
        ):
            try:
                setattr(
                    worksheet,
                    attribute,
                    copy.copy(getattr(template, attribute)),
                )
            except (AttributeError, TypeError):
                pass
        for attribute in ("print_area", "print_title_rows", "print_title_cols"):
            try:
                setattr(worksheet, attribute, getattr(template, attribute))
            except (AttributeError, TypeError, ValueError):
                pass
        for key, dimension in template.column_dimensions.items():
            worksheet.column_dimensions[key] = copy.copy(dimension)

        for (row, column), source in list(template._cells.items()):
            if row > header.row_end:
                continue
            if isinstance(source, MergedCell):
                continue
            target = worksheet.cell(row, column)
            self._copy_cell(source, target, copy_value=True)
        for merged in template.merged_cells.ranges:
            if merged.max_row <= header.row_end:
                worksheet.merge_cells(str(merged))
        for row in range(1, header.row_end + 1):
            if row in template.row_dimensions:
                worksheet.row_dimensions[row] = copy.copy(
                    template.row_dimensions[row]
                )

        template_target_row = header.row_end + 1
        max_column = self._actual_max_column(template)
        if data_row in template.row_dimensions:
            worksheet.row_dimensions[template_target_row] = copy.copy(
                template.row_dimensions[data_row]
            )
        for column in range(1, max_column + 1):
            source = template.cell(data_row, column)
            target = worksheet.cell(template_target_row, column)
            self._copy_cell(source, target, copy_value=False)
        return worksheet

    @staticmethod
    def _copy_cell(source: Any, target: Any, *, copy_value: bool) -> None:
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.hyperlink is not None and copy_value:
            target._hyperlink = copy.copy(source.hyperlink)
        if source.comment is not None and copy_value:
            target.comment = copy.copy(source.comment)
        target.value = source.value if copy_value else None

    @classmethod
    def _copy_row_style(
        cls,
        worksheet: Worksheet,
        source_row: int,
        target_row: int,
        *,
        max_column: int | None = None,
    ) -> None:
        if source_row <= 0:
            return
        worksheet.row_dimensions[target_row].height = (
            worksheet.row_dimensions[source_row].height
        )
        for column in range(
            1, (max_column or cls._actual_max_column(worksheet)) + 1
        ):
            source = worksheet.cell(source_row, column)
            target = worksheet.cell(target_row, column)
            cls._copy_cell(source, target, copy_value=False)

    @staticmethod
    def _write_sync_values(
        worksheet: Worksheet,
        target_row: int,
        values: Sequence[Any],
        expected: dict[tuple[int, int], Any],
    ) -> None:
        for index, field in enumerate(SYNC_FIELDS):
            column = TARGET_EXPECTED_COLUMNS[field]
            value = values[index]
            worksheet.cell(target_row, column).value = value
            expected[(target_row, column)] = value

    def _verify_saved_sync(
        self,
        path: Path,
        sheet_name: str,
        expected_cells: Mapping[tuple[int, int], Any],
        protected_cells: Mapping[tuple[int, int], Any],
    ) -> None:
        workbook = self.gateway.load(path, read_only=False)
        try:
            if sheet_name not in workbook.sheetnames:
                raise DailySyncError("Bản lưu thiếu sheet đích.")
            worksheet = workbook[sheet_name]
            for (row, column), value in expected_cells.items():
                if worksheet.cell(row, column).value != value:
                    raise DailySyncError(
                        f"Ô {worksheet.cell(row, column).coordinate} "
                        "không giữ đúng giá trị đồng bộ."
                    )
            for (row, column), value in protected_cells.items():
                if worksheet.cell(row, column).value != value:
                    raise DailySyncError(
                        f"Ô được bảo vệ {worksheet.cell(row, column).coordinate} "
                        "đã bị thay đổi."
                    )
        finally:
            workbook.close()

    def _selected_candidate(
        self, plan: SyncPlan, resolutions: Mapping[str, Any]
    ) -> MonthCandidate | None:
        if plan.selected_sheet is not None:
            return next(
                (
                    candidate
                    for candidate in plan.month_candidates
                    if candidate.target_sheet == plan.selected_sheet
                ),
                None,
            )
        for conflict in plan.conflicts:
            if conflict.conflict_type is not ConflictType.TARGET_MONTH_AMBIGUOUS:
                continue
            value = resolutions.get(conflict.conflict_id)
            if value is None:
                continue
            action = self._action(value)
            if action in {ResolutionAction.CANCEL, ResolutionAction.CANCEL_ALL}:
                raise DailySyncError("Người dùng đã hủy đồng bộ.")
            month = getattr(value, "selected_month", None)
            sheet = getattr(value, "selected_sheet", None)
            if isinstance(value, Mapping):
                month = value.get("selected_month", month)
                sheet = value.get(
                    "selected_sheet",
                    value.get("selected_sheet_name", sheet),
                )
            if action is ResolutionAction.SELECT_MONTH:
                if sheet:
                    return next(
                        (
                            candidate
                            for candidate in plan.month_candidates
                            if candidate.target_sheet == str(sheet)
                        ),
                        None,
                    )
                if month is not None:
                    matching = [
                        candidate
                        for candidate in plan.month_candidates
                        if candidate.month == int(month)
                    ]
                    if len(matching) == 1:
                        return matching[0]
        return None

    @staticmethod
    def _action(value: Any) -> ResolutionAction:
        if isinstance(value, (ResolutionAction, str)):
            return ResolutionAction(value)
        action = getattr(value, "action", None)
        if isinstance(value, Mapping):
            action = value.get("action", action)
        return ResolutionAction(action)

    def _create_run(self, source: Path, target: Path) -> int | None:
        if self.run_repository is None:
            return None
        record = self.run_repository.create_run(
            operation=ExcelOperation.DAILY_SYNC,
            source_path=source,
            target_path=target,
            status=ExcelRunStatus.ANALYZING,
        )
        return int(getattr(record, "id", record))

    def _update_run(self, run_id: int | None, **changes: Any) -> None:
        if self.run_repository is not None and run_id is not None:
            self.run_repository.update_run(run_id, **changes)

    def _finish_result(self, result: SyncResult) -> None:
        if self.run_repository is None or result.run_id is None:
            return
        total = (
            result.inserted_rows
            + result.updated_rows
            + result.unchanged_rows
            + result.target_only_rows
            + result.invalid_rows
        )
        self.run_repository.finish_run(
            result.run_id,
            status=result.status,
            sheet_name=result.sheet_name,
            backup_path=result.backup_path,
            target_fingerprint_after=result.fingerprint_after,
            total_items=total,
            changed_items=result.inserted_rows + result.updated_rows,
            skipped_items=result.invalid_rows,
            conflict_count=result.conflict_count,
        )

    def _finish_failed(self, run_id: int | None, exc: Exception) -> None:
        if self.run_repository is None or run_id is None:
            return
        cancelled = "người dùng đã hủy" in str(exc).casefold()
        self.run_repository.finish_run(
            run_id,
            status=(
                ExcelRunStatus.CANCELLED
                if cancelled
                else ExcelRunStatus.FAILED
            ),
            error_message=None if cancelled else str(exc),
        )
