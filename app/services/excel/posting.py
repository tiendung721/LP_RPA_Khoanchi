"""Two-phase posting of reviewed expense JSON into the BK workbook."""

from __future__ import annotations

import copy
import hashlib
import json
from collections import defaultdict
from collections.abc import Callable, Mapping, Sequence
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter

from app.constants import FEE_CODES, RULE_CODES
from app.services.json_codec import JsonCodec, JsonCodecError
from app.services.validation_service import normalize_bl, normalize_container

from .headers import HeaderResolution, HeaderResolutionError, HeaderResolver, normalize_header
from .daily_sync import (
    SOURCE_HEADER_ALIASES,
    SYNC_FIELDS,
    TARGET_EXPECTED_COLUMNS,
    DailySyncService,
)
from .models import (
    ConflictType,
    ExcelOperation,
    ExcelRunStatus,
    MonthCandidate,
    PostingConflict,
    PostingItem,
    PostingItemStatus,
    PostingPlan,
    PostingResolution,
    PostingResult,
    ResolutionAction,
    RowCandidate,
    TargetCellKind,
    TargetCellState,
    resolution_map,
)
from .resolvers import MonthSheetService, YearResolver
from .workbook import (
    ExcelBackupService,
    ExcelLockService,
    WorkbookGateway,
    ensure_supported_workbook,
)


ProgressCallback = Callable[[str], None] | None

BASE_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "sqt": ("SQT PM", "SQT", "Số thứ tự PM"),
    "closing_date": ("Ngày Đóng", "Ngày đóng hàng"),
    "container": ("Số Container", "Container", "Số cont"),
    "cargo_type": ("Loại hàng", "Tên hàng"),
    "vessel": ("Tên tàu", "Tàu"),
    "recipient": ("Người nhận", "Khách hàng"),
    "notes": ("Ghi chú", "Ghi Chú"),
}

FEE_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "CB": ("Cước biển QT",),
    "CBDH": ("Cước bộ đóng hàng", "ĐƠN GIÁ"),
    "VTN": ("Cước VTN",),
    "NV": ("Nâng vỏ",),
    "HH": ("Hạ Hàng",),
    "NH": ("Nâng Hàng",),
    "HV": ("Hạ vỏ",),
    "VSDL": ("VS D/O LỆNH", "VS DO LỆNH", "VS D/O"),
    "LC": ("Lưu cont",),
    "QT": ("Quá tải",),
    "LL": ("LÀM LỆNH",),
    "SC": ("SỬA CHỮA",),
}

INVOICE_HEADER_NAMES = frozenset(
    normalize_header(value)
    for value in (
        "Hóa đơn",
        "Số hóa đơn",
        "Ngày hóa đơn",
        "Tiền hóa đơn",
        "VAT",
        "Thuế GTGT",
        "Số HĐ",
        "Hóa đơn cước biển",
        "HD",
        "HĐ",
        "hoá đon",
    )
)
INVOICE_NUMBER_HEADER_NAMES = frozenset(
    normalize_header(value)
    for value in (
        "Hóa đơn",
        "Số hóa đơn",
        "Số HĐ",
        "Hóa đơn cước biển",
        "HD",
        "HĐ",
        "hoá đon",
    )
)
UPDATE_HEADER_NAMES = frozenset(
    normalize_header(value) for value in ("Date cập nhật", "Data cập nhật")
)
UPDATE_HEADER_CANONICAL = "Date cập nhật"
UPDATE_NUMBER_FORMAT = "dd/mm/yyyy hh:mm:ss"


class ExpensePostingError(RuntimeError):
    pass


def _stable_id(*parts: Any) -> str:
    payload = json.dumps(parts, ensure_ascii=False, default=str, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:24]


def _invoice_text(value: Any) -> str | None:
    if value in (None, ""):
        return None
    return str(value).strip() or None


def _invoice_key(value: Any) -> str | None:
    text = _invoice_text(value)
    return " ".join(text.casefold().split()) if text is not None else None


def _unique_invoice_values(values: Sequence[Any]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = _invoice_text(value)
        key = _invoice_key(text)
        if text is None or key is None or key in seen:
            continue
        seen.add(key)
        result.append(text)
    return result


def _period_offset(month: int, year: int, offset: int) -> tuple[int, int]:
    absolute = year * 12 + month - 1 - offset
    target_year, zero_based_month = divmod(absolute, 12)
    return zero_based_month + 1, target_year


def _progress(callback: ProgressCallback, message: str) -> None:
    if callback is not None:
        callback(message)


def classify_target_cell(cell: Any, amount: int) -> TargetCellState:
    value = cell.value
    coordinate = cell.coordinate
    data_type = getattr(cell, "data_type", None)
    if data_type == "f" or (isinstance(value, str) and value.startswith("=")):
        kind = TargetCellKind.FORMULA
    elif value in (None, ""):
        kind = TargetCellKind.EMPTY
    elif isinstance(value, bool):
        kind = TargetCellKind.TEXT
    elif isinstance(value, (int, float)):
        if value == amount:
            kind = TargetCellKind.SAME_VALUE
        elif value == 0:
            kind = TargetCellKind.ZERO
        else:
            kind = TargetCellKind.NUMBER
    else:
        kind = TargetCellKind.TEXT
    return TargetCellState(kind, value, coordinate, data_type)


class ExpensePostingService:
    def __init__(
        self,
        provider: Any | None = None,
        settings: Any | None = None,
        *,
        reviewed_batch_provider: Any | None = None,
        bk_path: str | Path | None = None,
        temp_dir: str | Path | None = None,
        backup_dir: str | Path | None = None,
        gateway: WorkbookGateway | None = None,
        lock_service: ExcelLockService | None = None,
        header_resolver: HeaderResolver | None = None,
        month_service: MonthSheetService | None = None,
        year_resolver: YearResolver | None = None,
        run_repository: Any | None = None,
        posting_repository: Any | None = None,
        clock: Callable[[], datetime] | None = None,
    ) -> None:
        self.provider = reviewed_batch_provider or provider
        if self.provider is None:
            raise TypeError("ExpensePostingService cần ReviewedBatchProvider.")
        paths = getattr(settings, "paths", None)
        self.bk_path = Path(
            bk_path or getattr(settings, "bk_workbook_path", "")
        )
        system_dir = getattr(paths, "system_dir", Path("Output") / "_system")
        self.temp_dir = Path(
            temp_dir
            or getattr(paths, "excel_temp_dir", system_dir / "Excel" / "Temp")
        )
        self.backup_dir = Path(
            backup_dir
            or getattr(paths, "excel_backup_dir", system_dir / "Excel" / "Backup")
        )
        self.gateway = gateway or WorkbookGateway()
        self.lock_service = lock_service or ExcelLockService()
        self.headers = header_resolver or HeaderResolver()
        self.months = month_service or MonthSheetService()
        self.years = year_resolver or YearResolver()
        self.run_repository = run_repository
        self.posting_repository = posting_repository
        self.clock = clock or (lambda: datetime.now().astimezone().replace(tzinfo=None))
        # Working copy nằm cạnh BK để atomic replace không bao giờ cross-volume.
        self.backups = ExcelBackupService(self.backup_dir)

    def analyze(
        self,
        batch_id: int | None = None,
        sheet_name: str | None = None,
        repost_source_indices: Sequence[int] | None = None,
        progress_callback: ProgressCallback = None,
    ) -> PostingPlan:
        target = ensure_supported_workbook(self.bk_path)
        if not target.is_file():
            raise ExpensePostingError(f"Không tìm thấy file BK: {target}")
        batch_path = self._ready_path(batch_id)
        if batch_path is None:
            raise ExpensePostingError(
                "Không có JSON hiện hành đã xác nhận để nhập khoản chi."
            )
        batch_path = Path(batch_path)
        raw = batch_path.read_bytes()
        batch_hash = hashlib.sha256(raw).hexdigest()
        resolved_batch_id = batch_id or self._batch_id_for_path(batch_path)
        document_rows = self._validate_json(raw)
        already_posted = self._successful_indices(batch_hash)
        repost_indices = set(repost_source_indices or ())
        invalid_reposts = repost_indices.difference(already_posted)
        if invalid_reposts:
            raise ExpensePostingError(
                "Danh sách khoản nhập lại không còn hợp lệ; vui lòng đọc lại."
            )
        repost_selection_done = repost_source_indices is not None
        previously_posted = self._previously_posted_items(
            batch_hash,
            document_rows,
            already_posted,
        )
        run_id = self._create_run(batch_path, target)
        try:
            _progress(progress_callback, "Đang đọc cấu trúc file BK…")
            self.lock_service.ensure_readable(target)
            fingerprint = self.gateway.fingerprint(target)
            workbook = self.gateway.load(target, read_only=False)
            try:
                sheet_names = [
                    name
                    for name in workbook.sheetnames
                    if self.months.parse_target_sheet(name) is not None
                ]
                if not sheet_names:
                    raise ExpensePostingError("File BK không có sheet tháng TMM YY.")
                groups = self._group_rows(
                    document_rows,
                    already_posted,
                    repost_indices,
                )
                candidates = [
                    MonthCandidate(
                        month=parsed[0],
                        year=parsed[1],
                        source_sheet="",
                        target_sheet=name,
                    )
                    for name in sheet_names
                    if (parsed := self.months.parse_target_sheet(name)) is not None
                ]
                selected = self._choose_sheet(sheet_name, candidates, groups, None)
                items = self._items_from_groups(groups, repost_indices)
                conflicts: list[PostingConflict] = []
                if selected is not None:
                    sheet_items, item_conflicts = self._analyze_items(
                        workbook,
                        selected,
                        items,
                        batch_hash=batch_hash,
                    )
                    items = sheet_items
                    conflicts.extend(item_conflicts)
            finally:
                workbook.close()

            plan = PostingPlan(
                batch_id=resolved_batch_id,
                batch_path=batch_path,
                batch_hash=batch_hash,
                target_path=target,
                target_fingerprint=fingerprint,
                items=items,
                conflicts=conflicts,
                sheet_candidates=candidates,
                selected_sheet=selected,
                source_item_count=len(document_rows),
                already_posted_indices=already_posted,
                previously_posted_items=previously_posted,
                repost_source_indices=repost_indices,
                repost_selection_done=repost_selection_done,
                run_id=run_id,
            )
            status = (
                ExcelRunStatus.WAITING_USER
                if plan.requires_user_input
                else ExcelRunStatus.NO_CHANGES
                if not items
                else ExcelRunStatus.ANALYZING
            )
            self._update_run(
                run_id,
                status=status,
                source_fingerprint={
                    "size": len(raw),
                    "mtime_ns": batch_path.stat().st_mtime_ns,
                    "sha256": batch_hash,
                },
                target_fingerprint_before=fingerprint,
                total_items=len(document_rows),
                conflict_count=len(conflicts),
                sheet_name=selected,
            )
            return plan
        except Exception as exc:
            self._finish_failed(run_id, exc)
            raise

    def apply(
        self,
        plan: PostingPlan,
        resolutions: Mapping[str, Any] | Sequence[PostingResolution] | None = None,
        progress_callback: ProgressCallback = None,
    ) -> PostingResult:
        resolved = resolution_map(resolutions)
        if self._has_selector_resolution(plan, resolved):
            # SELECT_SHEET/SELECT_ROW/SELECT_FEE changes the target cell itself. Calling
            # apply directly would bypass the mandatory second analysis of
            # that cell and could silently keep an occupied/formula/text value.
            raise ExpensePostingError(
                "Lựa chọn sheet, dòng hoặc mã phí phải được phân tích lại bằng "
                "ExpensePostingService.refine() trước khi apply."
            )
        try:
            self._check_batch_resolution(plan, resolved)
            selected_sheet = self._selected_sheet(plan, resolved)
            if selected_sheet is None and plan.items:
                raise ExpensePostingError(
                    "Chưa chọn sheet tháng cần nhập khoản chi."
                )
            self.gateway.assert_unchanged(
                plan.target_path, plan.target_fingerprint, label="File BK"
            )
            if (
                hashlib.sha256(plan.batch_path.read_bytes()).hexdigest()
                != plan.batch_hash
            ):
                raise ExpensePostingError(
                    "JSON đã xác nhận đã thay đổi sau khi phân tích."
                )

            workbook = self.gateway.load(plan.target_path, read_only=False)
            try:
                if (
                    selected_sheet is not None
                    and selected_sheet not in workbook.sheetnames
                ):
                    raise ExpensePostingError(
                        f"Không tìm thấy sheet {selected_sheet}."
                    )
                base_items = copy.deepcopy(plan.items)
                if (
                    selected_sheet is not None
                    and plan.selected_sheet != selected_sheet
                ):
                    base_items, dynamic_conflicts = self._analyze_items(
                        workbook,
                        selected_sheet,
                        base_items,
                        batch_hash=plan.batch_hash,
                    )
                else:
                    dynamic_conflicts = []
                actions, history = self._resolve_apply_actions(
                    workbook[selected_sheet] if selected_sheet else None,
                    base_items,
                    [*plan.conflicts, *dynamic_conflicts],
                    resolved,
                    batch_hash=plan.batch_hash,
                )
            finally:
                workbook.close()
        except Exception as exc:
            self._finish_failed(plan.run_id, exc)
            raise

        write_actions = [
            action
            for action in actions
            if action.get("amount_write") or action.get("invoice_write")
        ]
        already_actions = [
            action
            for action in actions
            if action["status"] is PostingItemStatus.ALREADY_EXISTS
        ]
        skipped_actions = [
            action
            for action in actions
            if action["status"]
            in {
                PostingItemStatus.USER_SKIPPED,
                PostingItemStatus.NOT_MATCHED,
                PostingItemStatus.UNRESOLVED,
            }
        ]
        working_path: Path | None = None
        backup_path: Path | None = None
        after = plan.target_fingerprint
        update_timestamp: datetime | None = None
        update_column: int | None = None
        self._update_run(plan.run_id, status=ExcelRunStatus.APPLYING)
        try:
            if write_actions:
                with self.lock_service.acquire(plan.target_path):
                    pass
                self.gateway.assert_unchanged(
                    plan.target_path,
                    plan.target_fingerprint,
                    label="File BK",
                )
                working_path = self.backups.create_working_copy(
                    plan.target_path, run_id=plan.run_id
                )
                write_book = self.gateway.load(working_path, read_only=False)
                try:
                    worksheet = write_book[selected_sheet]
                    carried_plan_rows = self._write_carried_plan_rows(
                        worksheet, write_actions
                    )
                    base = self._resolve_base_headers(worksheet)
                    fee_columns = self._resolve_fee_columns(worksheet, base)
                    invoice_columns = self._resolve_invoice_columns(
                        base, fee_columns
                    )
                    update_column = self._ensure_update_column(worksheet)
                    update_timestamp = self.clock().replace(
                        tzinfo=None,
                        microsecond=0,
                    )
                    updated_rows: set[int] = set()
                    for action in write_actions:
                        column = int(action["target_column"])
                        fee = str(action["fee_selected"])
                        if fee_columns.get(fee) != column:
                            raise ExpensePostingError(
                                f"Cột phí {fee} không còn khớp header."
                            )
                        if action.get("amount_write"):
                            worksheet.cell(
                                int(action["target_row"]), column
                            ).value = action["value_after"]
                        if action.get("invoice_write"):
                            invoice_column = int(action["invoice_target_column"])
                            if invoice_columns.get(fee) != invoice_column:
                                raise ExpensePostingError(
                                    f"Cột Số HĐ của phí {fee} không còn khớp header."
                                )
                            worksheet.cell(
                                int(action["target_row"]), invoice_column
                            ).value = action["invoice_value_after"]
                        updated_rows.add(int(action["target_row"]))
                    for target_row in updated_rows:
                        update_cell = worksheet.cell(target_row, update_column)
                        update_cell.value = update_timestamp
                        update_cell.number_format = UPDATE_NUMBER_FORMAT
                    if carried_plan_rows:
                        from .payment_sync import (
                            find_summary_start,
                            refresh_bk_summary_formulas,
                        )

                        if find_summary_start(worksheet) is not None:
                            refresh_bk_summary_formulas(worksheet)
                    self.gateway.save(write_book, working_path)
                finally:
                    write_book.close()
                self._verify_posting(
                    working_path,
                    selected_sheet,
                    write_actions,
                    update_column=update_column,
                    update_timestamp=update_timestamp,
                )
                backup_path = self.backups.create_backup(plan.target_path)
                after = self.gateway.atomic_replace(
                    working_path,
                    plan.target_path,
                    expected=plan.target_fingerprint,
                )
                working_path = None
                self._record_carry_forwards(plan, write_actions)

            self._record_history(plan, history)
            status = (
                ExcelRunStatus.SUCCEEDED
                if write_actions
                else ExcelRunStatus.NO_CHANGES
            )
            posted_count = sum(
                len(action["source_indices"]) for action in write_actions
            )
            already_count = sum(
                len(action["source_indices"]) for action in already_actions
            )
            skipped_count = sum(
                len(action["source_indices"]) for action in skipped_actions
            )
            amount_written_count = sum(
                bool(action.get("amount_write")) for action in write_actions
            )
            invoice_written_count = sum(
                bool(action.get("invoice_write")) for action in write_actions
            )
            result = PostingResult(
                status=status,
                target_path=plan.target_path,
                sheet_name=selected_sheet,
                posted_source_items=posted_count,
                written_cells=amount_written_count,
                invoice_written_cells=invoice_written_count,
                skipped_source_items=skipped_count,
                already_existing_items=already_count,
                conflict_count=len(plan.conflicts),
                backup_path=backup_path,
                fingerprint_before=plan.target_fingerprint,
                fingerprint_after=after,
                run_id=plan.run_id,
                message=(
                    f"Đã nhập {posted_count} khoản vào {amount_written_count} ô tiền; "
                    f"cập nhật {invoice_written_count} ô Số HĐ."
                    if write_actions
                    else "Không có ô cần ghi."
                ),
            )
            self._finish_result(result)
            return result
        except Exception as exc:
            self._finish_failed(plan.run_id, exc)
            raise
        finally:
            if working_path is not None and working_path.exists():
                working_path.unlink()

    def refine(
        self,
        plan: PostingPlan,
        resolutions: Mapping[str, Any] | Sequence[PostingResolution] | None,
        progress_callback: ProgressCallback = None,
    ) -> PostingPlan:
        try:
            return self._refine_plan(
                plan,
                resolutions,
                progress_callback=progress_callback,
            )
        except Exception as exc:
            self._finish_failed(plan.run_id, exc)
            raise

    def _refine_plan(
        self,
        plan: PostingPlan,
        resolutions: Mapping[str, Any] | Sequence[PostingResolution] | None,
        progress_callback: ProgressCallback = None,
    ) -> PostingPlan:
        """Áp dụng lựa chọn dòng/mã phí rồi phân tích lại ô đích.

        Pha này chỉ đọc workbook. Nó bảo đảm một lựa chọn ``SELECT_ROW`` hoặc
        ``SELECT_FEE`` không thể đi thẳng tới apply trước khi các xung đột
        number/formula/text phát sinh được đưa lại cho dialog tổng hợp.
        """

        resolved = resolution_map(resolutions)
        self._check_batch_resolution(plan, resolved)
        selected_sheet = self._selected_sheet(plan, resolved)
        if selected_sheet is None:
            raise ExpensePostingError("Chưa chọn sheet tháng cần nhập khoản chi.")
        self.gateway.assert_unchanged(
            plan.target_path, plan.target_fingerprint, label="File BK"
        )
        if hashlib.sha256(plan.batch_path.read_bytes()).hexdigest() != plan.batch_hash:
            raise ExpensePostingError(
                "JSON đã xác nhận đã thay đổi sau khi phân tích."
            )

        items = copy.deepcopy(plan.items)
        for conflict in plan.conflicts:
            value = resolved.get(conflict.conflict_id)
            if conflict.conflict_type is ConflictType.MULTIPLE_EXPENSE_SAME_CELL:
                if value is None:
                    continue
                action = self._action(value)
                if action in {ResolutionAction.CANCEL, ResolutionAction.CANCEL_ALL}:
                    raise ExpensePostingError("Người dùng đã hủy nhập khoản chi.")
                if action is not ResolutionAction.SELECT_SOURCE_ITEM:
                    raise ExpensePostingError("Chưa chọn dòng JSON cần ghi vào ô phí.")
                selected_source_index = self._resolution_attr(
                    value, "selected_source_item_index"
                )
                valid_indexes = {
                    int(option["source_item_index"])
                    for option in conflict.details.get("source_item_options", ())
                }
                if (
                    selected_source_index is None
                    or int(selected_source_index) not in valid_indexes
                ):
                    raise ExpensePostingError("Dòng JSON được chọn không hợp lệ.")
                selected_source_index = int(selected_source_index)
                for grouped_index in conflict.details.get("item_indexes", ()):
                    grouped_item = items[int(grouped_index)]
                    if selected_source_index in grouped_item.source_indices:
                        grouped_item.status = PostingItemStatus.PLANNED
                        grouped_item.action = None
                    else:
                        grouped_item.status = PostingItemStatus.USER_SKIPPED
                        grouped_item.action = ResolutionAction.SKIP
                continue
            if conflict.item_index is None:
                continue
            if value is None:
                continue
            action = self._action(value)
            item = items[conflict.item_index]
            invoice_conflict = conflict.conflict_type in {
                ConflictType.INVOICE_COLUMN_MISSING,
                ConflictType.INVOICE_VALUE_CONFLICT,
                ConflictType.MULTIPLE_SOURCE_INVOICES,
            }
            if invoice_conflict:
                if action is ResolutionAction.SELECT_INVOICE:
                    selected_invoice = self._resolution_attr(
                        value, "selected_invoice"
                    )
                    if _invoice_key(selected_invoice) not in {
                        _invoice_key(candidate)
                        for candidate in item.invoice_candidates
                    }:
                        raise ExpensePostingError("Số HĐ được chọn không hợp lệ.")
                    item.selected_invoice = str(selected_invoice)
                    item.invoice_action = None
                elif action in {
                    ResolutionAction.SKIP_INVOICE,
                    ResolutionAction.KEEP_EXISTING,
                    ResolutionAction.OVERWRITE,
                }:
                    item.invoice_action = action
                elif action in {
                    ResolutionAction.CANCEL,
                    ResolutionAction.CANCEL_ALL,
                }:
                    raise ExpensePostingError("Người dùng đã hủy nhập khoản chi.")
                continue
            if action is ResolutionAction.SELECT_FEE:
                selected_fee = self._resolution_attr(value, "selected_fee")
                if selected_fee not in FEE_HEADER_ALIASES:
                    raise ExpensePostingError("Mã phí được chọn không hợp lệ.")
                item.selected_fee = str(selected_fee)
                item.selected_invoice = None
                item.invoice_action = None
            elif action is ResolutionAction.SELECT_ROW:
                selected_row = self._resolution_attr(value, "selected_row")
                selected_source_sheet = self._resolution_attr(
                    value, "selected_source_sheet"
                )
                if selected_row is None:
                    raise ExpensePostingError("Chưa chọn dòng BK.")
                item.selected_source_row = int(selected_row)
                item.selected_source_sheet = str(
                    selected_source_sheet or selected_sheet
                )
                item.target_row = None
                item.invoice_action = None
            elif action in {
                ResolutionAction.OVERWRITE,
            }:
                item.action = action
            elif action in {
                ResolutionAction.SKIP,
                ResolutionAction.KEEP_EXISTING,
                ResolutionAction.KEEP_FORMULA,
            }:
                item.action = action
                item.status = PostingItemStatus.USER_SKIPPED
            elif action in {ResolutionAction.CANCEL, ResolutionAction.CANCEL_ALL}:
                raise ExpensePostingError("Người dùng đã hủy nhập khoản chi.")

        _progress(progress_callback, "Đang kiểm tra lại dòng và ô đã chọn…")
        workbook = self.gateway.load(plan.target_path, read_only=False)
        try:
            if selected_sheet not in workbook.sheetnames:
                raise ExpensePostingError(f"Không tìm thấy sheet {selected_sheet}.")
            items, conflicts = self._analyze_items(
                workbook,
                selected_sheet,
                items,
                batch_hash=plan.batch_hash,
            )
        finally:
            workbook.close()
        # Không hỏi lại conflict đã được giải quyết. Conflict mới (đặc biệt ô
        # sau SELECT_ROW/SELECT_FEE) có ID khác và vẫn được giữ lại.
        conflicts = [
            conflict
            for conflict in conflicts
            if conflict.conflict_id not in resolved
        ]
        refined = PostingPlan(
            batch_id=plan.batch_id,
            batch_path=plan.batch_path,
            batch_hash=plan.batch_hash,
            target_path=plan.target_path,
            target_fingerprint=plan.target_fingerprint,
            items=items,
            conflicts=conflicts,
            sheet_candidates=plan.sheet_candidates,
            selected_sheet=selected_sheet,
            source_item_count=plan.source_item_count,
            already_posted_indices=plan.already_posted_indices,
            previously_posted_items=plan.previously_posted_items,
            repost_source_indices=plan.repost_source_indices,
            repost_selection_done=plan.repost_selection_done,
            run_id=plan.run_id,
        )
        self._update_run(
            plan.run_id,
            status=(
                ExcelRunStatus.WAITING_USER
                if conflicts
                else ExcelRunStatus.ANALYZING
            ),
            sheet_name=selected_sheet,
            conflict_count=len(conflicts),
        )
        return refined

    def _validate_json(self, raw: bytes) -> list[dict[str, Any]]:
        try:
            document = JsonCodec().loads(raw)
        except JsonCodecError as exc:
            raise ExpensePostingError(
                f"JSON đã xác nhận không đọc được: {exc}"
            ) from exc
        result: list[dict[str, Any]] = []
        for index, row in enumerate(document.rows):
            container = row.cont
            bl = row.bl
            fee = row.fee
            rule = row.rule
            invoice_no = row.invoice_no
            carrier = row.carrier
            amount = row.amount
            if container is not None and not isinstance(container, str):
                raise ExpensePostingError(f"Container dòng {index + 1} không hợp lệ.")
            if bl is not None and not isinstance(bl, str):
                raise ExpensePostingError(f"B/L dòng {index + 1} không hợp lệ.")
            if invoice_no is not None and not isinstance(invoice_no, str):
                raise ExpensePostingError(f"Số HĐ dòng {index + 1} không hợp lệ.")
            if carrier is not None and not isinstance(carrier, str):
                raise ExpensePostingError(
                    f"Bên vận tải dòng {index + 1} không hợp lệ."
                )
            if not isinstance(fee, str) or fee.strip().upper() not in FEE_CODES:
                raise ExpensePostingError(f"Mã phí dòng {index + 1} không hợp lệ.")
            if rule is not None and (
                not isinstance(rule, str) or rule.strip().upper() not in RULE_CODES
            ):
                raise ExpensePostingError(f"Rule dòng {index + 1} không hợp lệ.")
            if isinstance(amount, bool) or type(amount) is not int or amount < 0:
                raise ExpensePostingError(f"Số tiền dòng {index + 1} không hợp lệ.")
            result.append(
                {
                    "source_item_index": index,
                    "container": normalize_container(container),
                    "bl": normalize_bl(bl),
                    "fee": fee.strip().upper(),
                    "rule": rule.strip().upper() if isinstance(rule, str) else None,
                    "invoice_no": invoice_no,
                    "carrier": carrier,
                    "amount": amount,
                }
            )
        return result

    def cancel(self, plan: PostingPlan) -> None:
        """Đánh dấu lần nhập đang chờ lựa chọn là đã hủy, không sửa BK."""

        if self.run_repository is not None and plan.run_id is not None:
            self.run_repository.finish_run(
                plan.run_id,
                status=ExcelRunStatus.CANCELLED,
                sheet_name=plan.selected_sheet,
                total_items=plan.source_item_count,
                changed_items=0,
                skipped_items=0,
                conflict_count=len(plan.conflicts),
            )

    @staticmethod
    def _group_rows(
        rows: Sequence[dict[str, Any]],
        already_posted: set[int],
        repost_source_indices: set[int] | None = None,
    ) -> list[list[dict[str, Any]]]:
        repost = repost_source_indices or set()
        # Mỗi dòng hóa đơn phải còn độc lập cho tới khi xác định được dòng BK.
        # Cùng container/mã phí chưa đủ để kết luận các khoản thuộc cùng một SQT.
        return [
            [row]
            for row in rows
            if row["source_item_index"] not in already_posted
            or row["source_item_index"] in repost
        ]

    @staticmethod
    def _items_from_groups(
        groups: Sequence[Sequence[dict[str, Any]]],
        repost_source_indices: set[int] | None = None,
    ) -> list[PostingItem]:
        repost = repost_source_indices or set()
        return [
            PostingItem(
                source_indices=[row["source_item_index"] for row in group],
                container=group[0]["container"],
                bl=group[0]["bl"],
                original_fee=group[0]["fee"],
                selected_fee=group[0]["fee"],
                rule=group[0]["rule"],
                amount=sum(row["amount"] for row in group),
                source_items=[dict(row) for row in group],
                invoice_candidates=_unique_invoice_values(
                    [row.get("invoice_no") for row in group]
                ),
                force_repost=any(
                    row["source_item_index"] in repost for row in group
                ),
            )
            for group in groups
        ]

    def _resolve_base_headers(self, worksheet: Any) -> HeaderResolution:
        return self.headers.resolve(
            worksheet,
            BASE_HEADER_ALIASES,
            required=("sqt", "closing_date", "container"),
        )

    def _resolve_fee_columns(
        self, worksheet: Any, base: HeaderResolution
    ) -> dict[str, int]:
        from .payment_sync import find_summary_start

        normalized = base.headers
        summary_start = find_summary_start(worksheet)
        result: dict[str, int] = {}
        notes_candidates = [
            column
            for column, header in normalized.items()
            if header in {
                normalize_header(alias)
                for alias in BASE_HEADER_ALIASES["notes"]
            }
        ]
        notes_column = min(notes_candidates) if notes_candidates else None
        for fee, aliases in FEE_HEADER_ALIASES.items():
            alias_keys = {normalize_header(alias) for alias in aliases}
            matches = [
                column
                for column, header in normalized.items()
                if header in alias_keys
                and (summary_start is None or column < summary_start)
            ]
            if len(matches) != 1:
                continue
            column = matches[0]
            if normalized[column] in INVOICE_HEADER_NAMES:
                continue
            if (
                fee == "CBDH"
                and normalized[column] == normalize_header("ĐƠN GIÁ")
                and notes_column is not None
                and column >= notes_column
            ):
                # "ĐƠN GIÁ" trong vùng sau GHI CHÚ thuộc thông tin hóa đơn,
                # không phải cột Cước bộ đóng hàng.
                continue
            if fee == "LL" and (
                notes_column is None or column >= notes_column
            ):
                # Không đoán cột LÀM LỆNH nếu không xác định được ranh giới
                # vùng cước chính trước GHI CHÚ.
                continue
            result[fee] = column
        return result

    @staticmethod
    def _resolve_invoice_columns(
        base: HeaderResolution,
        fee_columns: Mapping[str, int],
    ) -> dict[str, int]:
        result: dict[str, int] = {}
        for fee, fee_column in fee_columns.items():
            if fee == "LL":
                continue
            invoice_column = fee_column + 1
            if base.headers.get(invoice_column) in INVOICE_NUMBER_HEADER_NAMES:
                result[fee] = invoice_column
        return result

    @staticmethod
    def _update_column(
        base: HeaderResolution,
    ) -> int | None:
        matches = [
            column
            for column, header in base.headers.items()
            if header in UPDATE_HEADER_NAMES
        ]
        if len(matches) > 1:
            raise ExpensePostingError(
                "Sheet BK có nhiều cột Date cập nhật; không thể chọn an toàn."
            )
        return matches[0] if matches else None

    def _ensure_update_column(self, worksheet: Any) -> int:
        base = self._resolve_base_headers(worksheet)
        existing = self._update_column(base)
        if existing is not None:
            return existing
        used_headers = [
            column for column, header in base.headers.items() if header
        ]
        column = (max(used_headers) if used_headers else worksheet.max_column) + 1
        header_cell = worksheet.cell(base.row_end, column)
        previous = worksheet.cell(base.row_end, column - 1)
        if previous.has_style:
            header_cell._style = copy.copy(previous._style)
        header_cell.font = copy.copy(previous.font)
        header_cell.fill = copy.copy(previous.fill)
        header_cell.border = copy.copy(previous.border)
        header_cell.alignment = copy.copy(previous.alignment)
        header_cell.protection = copy.copy(previous.protection)
        header_cell.value = UPDATE_HEADER_CANONICAL
        return column

    def _container_index(
        self,
        worksheet: Any,
        base: HeaderResolution,
        *,
        plan_header: HeaderResolution | None = None,
    ) -> dict[str, list[RowCandidate]]:
        result: dict[str, list[RowCandidate]] = defaultdict(list)
        columns = base.columns
        for row in range(base.row_end + 1, worksheet.max_row + 1):
            raw = worksheet.cell(row, columns["container"]).value
            try:
                container = normalize_container(
                    str(raw) if raw not in (None, "") else None
                )
            except TypeError:
                continue
            if not container:
                continue
            cargo = (
                worksheet.cell(row, columns["cargo_type"]).value
                if "cargo_type" in columns
                else None
            )
            plan_values = (
                tuple(
                    worksheet.cell(row, plan_header.columns[field]).value
                    for field in SYNC_FIELDS
                )
                if plan_header is not None
                else ()
            )
            result[container].append(
                RowCandidate(
                    row=row,
                    sqt=self._positive_int(
                        worksheet.cell(row, columns["sqt"]).value
                    ),
                    container=container,
                    source_sheet=worksheet.title,
                    source_row=row,
                    plan_values=plan_values,
                    source_signature=(
                        self._plan_signature(plan_values) if plan_values else ""
                    ),
                    cargo_type=cargo,
                    closing_date=worksheet.cell(
                        row, columns["closing_date"]
                    ).value,
                    vessel=(
                        worksheet.cell(row, columns["vessel"]).value
                        if "vessel" in columns
                        else None
                    ),
                    recipient=(
                        worksheet.cell(row, columns["recipient"]).value
                        if "recipient" in columns
                        else None
                    ),
                    is_ron="ron" in normalize_header(cargo),
                )
            )
        return dict(result)

    @staticmethod
    def _plan_signature(values: Sequence[Any]) -> str:
        payload = json.dumps(
            list(values),
            ensure_ascii=False,
            default=str,
            separators=(",", ":"),
        )
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    def _plan_header(self, worksheet: Any) -> HeaderResolution:
        return self.headers.resolve(
            worksheet,
            SOURCE_HEADER_ALIASES,
            required=SYNC_FIELDS,
        )

    def _source_window_names(self, workbook: Any, target_sheet: str) -> list[str]:
        parsed_target = self.months.parse_target_sheet(target_sheet)
        if parsed_target is None:
            raise ExpensePostingError(f"Sheet {target_sheet!r} không có dạng TMM YY.")
        month, year = parsed_target
        by_period: dict[tuple[int, int], str] = {}
        for name in workbook.sheetnames:
            parsed = self.months.parse_target_sheet(name)
            if parsed is None:
                continue
            if parsed in by_period:
                raise ExpensePostingError(
                    f"Có nhiều sheet BK cho tháng {parsed[0]:02d}/{parsed[1]}."
                )
            by_period[parsed] = name
        return [
            by_period[period]
            for offset in range(3)
            if (period := _period_offset(month, year, offset)) in by_period
        ]

    def _carry_forward_record(
        self,
        candidate: RowCandidate,
        target_sheet: str,
    ) -> Any | None:
        getter = getattr(self.posting_repository, "get_carry_forward", None)
        if not callable(getter) or candidate.sqt is None or not candidate.container:
            return None
        return getter(
            workbook_path=self.bk_path,
            source_sheet=candidate.source_sheet,
            source_row=int(candidate.source_row or candidate.row),
            source_sqt=candidate.sqt,
            container=candidate.container,
            target_sheet=target_sheet,
        )

    @staticmethod
    def _row_plan_values(
        worksheet: Any,
        header: HeaderResolution,
        row: int,
    ) -> tuple[Any, ...]:
        return tuple(
            worksheet.cell(row, header.columns[field]).value
            for field in SYNC_FIELDS
        )

    def _window_index(
        self,
        workbook: Any,
        target_sheet: str,
    ) -> tuple[dict[str, list[RowCandidate]], list[RowCandidate], HeaderResolution | None]:
        window_names = self._source_window_names(workbook, target_sheet)
        indexes: dict[str, dict[str, list[RowCandidate]]] = {}
        plan_headers: dict[str, HeaderResolution | None] = {}
        for name in window_names:
            worksheet = workbook[name]
            base = self._resolve_base_headers(worksheet)
            try:
                plan_header = self._plan_header(worksheet)
            except HeaderResolutionError:
                plan_header = None
                if name != target_sheet:
                    raise ExpensePostingError(
                        f"Sheet nguồn {name} không đủ 12 cột thông tin kế hoạch."
                    )
            plan_headers[name] = plan_header
            indexes[name] = self._container_index(
                worksheet,
                base,
                plan_header=plan_header,
            )

        target_worksheet = workbook[target_sheet]
        target_plan_header = plan_headers.get(target_sheet)
        target_candidates = [
            candidate
            for candidates in indexes[target_sheet].values()
            for candidate in candidates
        ]
        carried_target_rows: set[int] = set()
        for name in window_names[1:]:
            for candidates in indexes[name].values():
                for candidate in candidates:
                    record = self._carry_forward_record(candidate, target_sheet)
                    if record is None:
                        continue
                    if str(getattr(record, "source_signature", "")) != candidate.source_signature:
                        candidate.mapping_invalid = True
                        continue
                    target_row = int(getattr(record, "target_row"))
                    matches_recorded_row = (
                        target_plan_header is not None
                        and self._row_plan_values(
                            target_worksheet, target_plan_header, target_row
                        )
                        == candidate.plan_values
                    )
                    if matches_recorded_row:
                        candidate.carried_target_row = target_row
                        carried_target_rows.add(target_row)
                        continue
                    signature_matches = [
                        item.row
                        for item in target_candidates
                        if item.source_signature == candidate.source_signature
                    ]
                    if len(signature_matches) == 1:
                        candidate.carried_target_row = signature_matches[0]
                        carried_target_rows.add(signature_matches[0])
                    else:
                        candidate.mapping_invalid = True

        combined: dict[str, list[RowCandidate]] = defaultdict(list)
        all_candidates: list[RowCandidate] = []
        for name in window_names:
            for container, candidates in indexes[name].items():
                for candidate in candidates:
                    if name == target_sheet and candidate.row in carried_target_rows:
                        continue
                    combined[container].append(candidate)
                    all_candidates.append(candidate)
        return dict(combined), all_candidates, target_plan_header

    @staticmethod
    def _origin_key(candidate: RowCandidate) -> tuple[str, int, int | None, str | None]:
        return (
            candidate.source_sheet,
            int(candidate.source_row or candidate.row),
            candidate.sqt,
            candidate.container,
        )

    def _sheet_candidates(
        self,
        workbook: Any,
        sheet_names: Sequence[str],
        groups: Sequence[Sequence[dict[str, Any]]],
        latest_sheet: str | None,
    ) -> tuple[list[MonthCandidate], dict[str, dict[str, list[RowCandidate]]]]:
        containers = {
            row["container"]
            for group in groups
            for row in group
            if row["container"]
        }
        candidates: list[MonthCandidate] = []
        indexes: dict[str, dict[str, list[RowCandidate]]] = {}
        for name in sheet_names:
            base = self._resolve_base_headers(workbook[name])
            index = self._container_index(workbook[name], base)
            indexes[name] = index
            parsed = self.months.parse_target_sheet(name)
            candidates.append(
                MonthCandidate(
                    month=parsed[0] if parsed else 0,
                    source_sheet="",
                    target_sheet=name,
                    match_count=sum(container in index for container in containers),
                    recently_synced=name == latest_sheet,
                )
            )
        return candidates, indexes

    @staticmethod
    def _choose_sheet(
        requested: str | None,
        candidates: Sequence[MonthCandidate],
        groups: Sequence[Sequence[dict[str, Any]]],
        latest_sheet: str | None,
    ) -> str | None:
        names = {candidate.target_sheet for candidate in candidates}
        if requested is not None:
            if requested not in names:
                raise ExpensePostingError(f"Sheet {requested!r} không hợp lệ.")
            return requested
        return None

    def _analyze_items(
        self,
        workbook: Any,
        target_sheet: str,
        items: list[PostingItem],
        *,
        batch_hash: str,
    ) -> tuple[list[PostingItem], list[PostingConflict]]:
        worksheet = workbook[target_sheet]
        base = self._resolve_base_headers(worksheet)
        fee_columns = self._resolve_fee_columns(worksheet, base)
        invoice_columns = self._resolve_invoice_columns(base, fee_columns)
        index, manual_candidates, target_plan_header = self._window_index(
            workbook, target_sheet
        )
        manual_candidates = sorted(
            manual_candidates,
            key=lambda candidate: (
                self._source_window_names(workbook, target_sheet).index(
                    candidate.source_sheet
                ),
                candidate.row,
            ),
        )
        next_target_row = (
            DailySyncService._last_data_row(worksheet, target_plan_header) + 1
            if target_plan_header is not None
            else base.row_end + 1
        )
        allocated_rows: dict[tuple[str, int, int | None, str | None], int] = {}
        analyzed_items: list[PostingItem] = []
        conflicts: list[PostingConflict] = []
        target_groups: dict[tuple[str, int, int], list[int]] = defaultdict(list)

        for item in items:
            if item.status is not PostingItemStatus.PLANNED:
                analyzed_items.append(item)
                continue
            item.sheet_name = target_sheet
            item.target_column = None
            item.target_cell = None
            item.current_value = None
            item.cell_state = None
            item.carry_forward_required = False
            if item.selected_fee not in FEE_HEADER_ALIASES:
                item_index = len(analyzed_items)
                analyzed_items.append(item)
                conflicts.append(
                    self._item_conflict(
                        batch_hash,
                        item_index,
                        item,
                        ConflictType.UNKNOWN_FEE_CODE,
                        "Mã phí chưa xác định; hãy chọn một trong 12 mã phí.",
                        (ResolutionAction.SELECT_FEE, ResolutionAction.SKIP),
                        details={"fees": sorted(FEE_HEADER_ALIASES)},
                    )
                )
                continue

            manually_selected: RowCandidate | None = None
            selected_source_row = item.selected_source_row
            selected_source_sheet = item.selected_source_sheet
            if selected_source_row is None and item.target_row is not None:
                selected_source_row = item.target_row
                selected_source_sheet = target_sheet
            if selected_source_row is not None:
                manually_selected = next(
                    (
                        candidate
                        for candidate in manual_candidates
                        if candidate.row == int(selected_source_row)
                        and candidate.source_sheet
                        == str(selected_source_sheet or target_sheet)
                    ),
                    None,
                )
                if manually_selected is None:
                    raise ExpensePostingError(
                        f"Dòng nguồn {selected_source_sheet or target_sheet}!"
                        f"{selected_source_row} không còn hợp lệ."
                    )

            if item.container is None and manually_selected is None:
                item_index = len(analyzed_items)
                analyzed_items.append(item)
                conflict_type = (
                    ConflictType.BL_ONLY_NO_CONTAINER
                    if item.bl and item.selected_fee == "CB"
                    else ConflictType.CONTAINER_NOT_FOUND
                )
                conflicts.append(
                    self._item_conflict(
                        batch_hash,
                        item_index,
                        item,
                        conflict_type,
                        "Khoản chi chưa có container; hãy chọn đúng dòng kế hoạch.",
                        (ResolutionAction.SELECT_ROW, ResolutionAction.SKIP),
                        row_candidates=manual_candidates,
                    )
                )
                continue

            candidates = (
                [manually_selected]
                if manually_selected is not None
                else list(index.get(item.container or "", ()))
            )
            item.row_candidates = candidates
            selected = manually_selected or self._automatic_row(candidates)
            if not candidates:
                item_index = len(analyzed_items)
                analyzed_items.append(item)
                conflicts.append(
                    self._item_conflict(
                        batch_hash,
                        item_index,
                        item,
                        ConflictType.CONTAINER_NOT_FOUND,
                        f"Không tìm thấy container {item.container} trong tháng đích và hai tháng trước.",
                        (ResolutionAction.SELECT_ROW, ResolutionAction.SKIP),
                        row_candidates=manual_candidates,
                    )
                )
                continue
            if selected is None:
                item_index = len(analyzed_items)
                analyzed_items.append(item)
                conflicts.append(
                    self._item_conflict(
                        batch_hash,
                        item_index,
                        item,
                        ConflictType.MULTIPLE_CONTAINER_MATCH,
                        f"Container {item.container} có nhiều dòng kế hoạch; hãy chọn đúng tháng và SQT.",
                        (ResolutionAction.SELECT_ROW, ResolutionAction.SKIP),
                        row_candidates=candidates,
                    )
                )
                continue

            item.selected_source_sheet = selected.source_sheet
            item.selected_source_row = int(selected.source_row or selected.row)
            item.source_sqt = selected.sqt
            item.plan_values = tuple(selected.plan_values)
            item.source_signature = selected.source_signature
            if selected.mapping_invalid:
                item_index = len(analyzed_items)
                analyzed_items.append(item)
                conflicts.append(
                    self._item_conflict(
                        batch_hash,
                        item_index,
                        item,
                        ConflictType.CARRY_FORWARD_MAPPING_INVALID,
                        "Ánh xạ dòng đã mang sang không còn khớp; cần kiểm tra BK trước khi tiếp tục.",
                        (ResolutionAction.SKIP, ResolutionAction.CANCEL_ALL),
                        default=ResolutionAction.SKIP,
                    )
                )
                continue

            if selected.source_sheet == target_sheet:
                item.target_row = selected.row
            elif selected.carried_target_row is not None:
                item.target_row = selected.carried_target_row
            else:
                if target_plan_header is None or not selected.plan_values:
                    raise ExpensePostingError(
                        f"Không thể mang dòng {selected.source_sheet}!{selected.row}: "
                        "sheet đích không đủ 12 cột kế hoạch."
                    )
                origin_key = self._origin_key(selected)
                if origin_key not in allocated_rows:
                    allocated_rows[origin_key] = next_target_row
                    next_target_row += 1
                item.target_row = allocated_rows[origin_key]
                item.carry_forward_required = True

            item.target_column = fee_columns.get(item.selected_fee)
            if item.target_column is None:
                item_index = len(analyzed_items)
                analyzed_items.append(item)
                conflicts.append(
                    self._item_conflict(
                        batch_hash,
                        item_index,
                        item,
                        ConflictType.FEE_COLUMN_MISSING,
                        f"Không nhận diện duy nhất cột phí {item.selected_fee}.",
                        (ResolutionAction.SKIP,),
                        default=ResolutionAction.SKIP,
                    )
                )
                continue

            item_index = len(analyzed_items)
            analyzed_items.append(item)
            target_groups[(target_sheet, item.target_row, item.target_column)].append(
                item_index
            )

        for (_sheet, _row, _column), item_indexes in target_groups.items():
            if len(item_indexes) > 1:
                first_index = item_indexes[0]
                first = analyzed_items[first_index]
                first.target_cell = f"{get_column_letter(int(first.target_column))}{first.target_row}"
                conflicts.append(
                    self._item_conflict(
                        batch_hash,
                        first_index,
                        first,
                        ConflictType.MULTIPLE_EXPENSE_SAME_CELL,
                        "Nhiều dòng JSON cùng trỏ tới một ô phí; hãy chọn đúng một dòng để ghi.",
                        (ResolutionAction.SELECT_SOURCE_ITEM,),
                        default=ResolutionAction.SELECT_SOURCE_ITEM,
                        details={
                            "item_indexes": list(item_indexes),
                            "source_item_options": [
                                {
                                    "source_item_index": analyzed_items[index].source_indices[0],
                                    "amount": analyzed_items[index].amount,
                                    "invoice_no": (
                                        analyzed_items[index].invoice_candidates[0]
                                        if len(analyzed_items[index].invoice_candidates) == 1
                                        else None
                                    ),
                                }
                                for index in item_indexes
                            ],
                        },
                    )
                )
                continue

            item_index = item_indexes[0]
            item = analyzed_items[item_index]
            self._set_cell_state(worksheet, item)
            cell_conflict = self._cell_conflict(batch_hash, item_index, item)
            if cell_conflict is not None:
                conflicts.append(cell_conflict)
            invoice_conflict = self._set_invoice_state(
                worksheet,
                item,
                item_index=item_index,
                batch_hash=batch_hash,
                invoice_columns=invoice_columns,
            )
            if invoice_conflict is not None:
                conflicts.append(invoice_conflict)
        return analyzed_items, conflicts

    def _set_invoice_state(
        self,
        worksheet: Any,
        item: PostingItem,
        *,
        item_index: int,
        batch_hash: str,
        invoice_columns: Mapping[str, int],
    ) -> PostingConflict | None:
        item.invoice_column = None
        item.invoice_cell = None
        item.invoice_current_value = None
        item.invoice_value_after = None

        if item.selected_fee == "LL" or not item.invoice_candidates:
            item.selected_invoice = None
            item.invoice_action = None
            return None

        if item.selected_invoice is not None and _invoice_key(
            item.selected_invoice
        ) not in {_invoice_key(value) for value in item.invoice_candidates}:
            item.selected_invoice = None

        invoice_column = invoice_columns.get(item.selected_fee)
        if invoice_column is None:
            item.invoice_action = None
            return self._item_conflict(
                batch_hash,
                item_index,
                item,
                ConflictType.INVOICE_COLUMN_MISSING,
                "Không nhận diện được cột Số HĐ tương ứng với loại phí.",
                (ResolutionAction.SKIP_INVOICE, ResolutionAction.CANCEL_ALL),
                default=ResolutionAction.SKIP_INVOICE,
                details={
                    "scope": "invoice",
                    "invoice_candidates": list(item.invoice_candidates),
                },
            )

        item.invoice_column = invoice_column
        invoice_cell = worksheet.cell(int(item.target_row), invoice_column)
        item.invoice_cell = invoice_cell.coordinate
        item.invoice_current_value = invoice_cell.value

        if len(item.invoice_candidates) > 1 and item.selected_invoice is None:
            item.invoice_action = None
            return self._item_conflict(
                batch_hash,
                item_index,
                item,
                ConflictType.MULTIPLE_SOURCE_INVOICES,
                "Nhiều Số HĐ khác nhau cùng trỏ tới một ô BK; hãy chọn một Số HĐ.",
                (ResolutionAction.SELECT_INVOICE,),
                default=ResolutionAction.SELECT_INVOICE,
                details={
                    "scope": "invoice",
                    "invoice_candidates": list(item.invoice_candidates),
                    "invoice_column": invoice_column,
                    "invoice_cell": invoice_cell.coordinate,
                    "invoice_current_value": invoice_cell.value,
                },
            )

        selected_invoice = item.selected_invoice or item.invoice_candidates[0]
        item.selected_invoice = selected_invoice
        if _invoice_text(invoice_cell.value) is None:
            item.invoice_action = ResolutionAction.OVERWRITE
            item.invoice_value_after = selected_invoice
            return None
        if _invoice_key(invoice_cell.value) == _invoice_key(selected_invoice):
            item.invoice_action = ResolutionAction.KEEP_EXISTING
            item.invoice_value_after = invoice_cell.value
            return None

        item.invoice_action = None
        return self._item_conflict(
            batch_hash,
            item_index,
            item,
            ConflictType.INVOICE_VALUE_CONFLICT,
            "Ô Số HĐ đang có giá trị khác với Số HĐ từ JSON.",
            (ResolutionAction.KEEP_EXISTING, ResolutionAction.OVERWRITE),
            default=ResolutionAction.KEEP_EXISTING,
            details={
                "scope": "invoice",
                "invoice_candidates": [selected_invoice],
                "selected_invoice": selected_invoice,
                "invoice_column": invoice_column,
                "invoice_cell": invoice_cell.coordinate,
                "invoice_current_value": invoice_cell.value,
            },
        )

    @staticmethod
    def _automatic_row(candidates: Sequence[RowCandidate]) -> RowCandidate | None:
        return candidates[0] if len(candidates) == 1 else None

    @staticmethod
    def _set_cell_state(worksheet: Any, item: PostingItem) -> None:
        cell = worksheet.cell(item.target_row, item.target_column)
        item.target_cell = cell.coordinate
        item.current_value = cell.value
        item.cell_state = classify_target_cell(cell, item.amount)
        if item.cell_state.kind in {TargetCellKind.EMPTY, TargetCellKind.ZERO}:
            item.action = ResolutionAction.OVERWRITE
        elif item.cell_state.kind is TargetCellKind.SAME_VALUE:
            if item.force_repost:
                item.action = ResolutionAction.OVERWRITE
            else:
                item.status = PostingItemStatus.ALREADY_EXISTS

    def _cell_conflict(
        self, batch_hash: str, item_index: int, item: PostingItem
    ) -> PostingConflict | None:
        if item.cell_state is None:
            return None
        if item.cell_state.kind is TargetCellKind.NUMBER:
            return self._item_conflict(
                batch_hash,
                item_index,
                item,
                ConflictType.TARGET_CELL_OCCUPIED,
                "Ô đích đang có số khác.",
                (
                    ResolutionAction.KEEP_EXISTING,
                    ResolutionAction.OVERWRITE,
                    ResolutionAction.SKIP,
                ),
                default=ResolutionAction.KEEP_EXISTING,
            )
        if item.cell_state.kind is TargetCellKind.FORMULA:
            return self._item_conflict(
                batch_hash,
                item_index,
                item,
                ConflictType.TARGET_CELL_FORMULA,
                "Ô đích chứa công thức.",
                (
                    ResolutionAction.KEEP_FORMULA,
                    ResolutionAction.OVERWRITE,
                    ResolutionAction.SKIP,
                ),
                default=ResolutionAction.KEEP_FORMULA,
            )
        if item.cell_state.kind is TargetCellKind.TEXT:
            return self._item_conflict(
                batch_hash,
                item_index,
                item,
                ConflictType.TARGET_CELL_TEXT,
                "Ô đích chứa văn bản.",
                (
                    ResolutionAction.KEEP_EXISTING,
                    ResolutionAction.OVERWRITE,
                    ResolutionAction.SKIP,
                ),
                default=ResolutionAction.KEEP_EXISTING,
            )
        return None

    def _item_conflict(
        self,
        batch_hash: str,
        item_index: int,
        item: PostingItem,
        conflict_type: ConflictType,
        message: str,
        actions: tuple[ResolutionAction, ...],
        *,
        default: ResolutionAction | None = None,
        row_candidates: Sequence[RowCandidate] = (),
        details: Mapping[str, Any] | None = None,
    ) -> PostingConflict:
        conflict_details = dict(details or {})
        invoice_scope = conflict_details.get("scope") == "invoice"
        target_column = (
            conflict_details.get("invoice_column")
            if invoice_scope
            else item.target_column
        )
        target_cell = (
            conflict_details.get("invoice_cell")
            if invoice_scope
            else item.target_cell
        )
        current_value = (
            conflict_details.get("invoice_current_value")
            if invoice_scope
            else item.current_value
        )
        return PostingConflict(
            conflict_id=_stable_id(
                conflict_type.value,
                batch_hash,
                item.source_indices,
                item.container,
                item.selected_fee,
                target_cell,
            ),
            conflict_type=conflict_type,
            message=message,
            item_index=item_index,
            container=item.container,
            bl=item.bl,
            fee=item.selected_fee,
            amount=item.amount,
            sheet_name=item.sheet_name,
            target_row=item.target_row,
            target_column=target_column,
            target_cell=target_cell,
            current_value=current_value,
            allowed_actions=actions,
            default_action=default,
            row_candidates=list(row_candidates),
            details=conflict_details,
        )

    def _resolve_apply_actions(
        self,
        worksheet: Any,
        items: list[PostingItem],
        conflicts: Sequence[PostingConflict],
        resolutions: Mapping[str, Any],
        *,
        batch_hash: str,
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        if worksheet is None:
            return [], []
        base = self._resolve_base_headers(worksheet)
        fee_columns = self._resolve_fee_columns(worksheet, base)
        invoice_columns = self._resolve_invoice_columns(base, fee_columns)
        index = self._container_index(worksheet, base)
        conflicts_by_item: dict[int, list[PostingConflict]] = defaultdict(list)
        for conflict in conflicts:
            if conflict.item_index is not None:
                conflicts_by_item[conflict.item_index].append(conflict)
        if any(
            conflict.conflict_type is ConflictType.MULTIPLE_EXPENSE_SAME_CELL
            for conflict in conflicts
        ):
            raise ExpensePostingError(
                "Xung đột nhiều dòng JSON cùng ô phí phải được refine trước khi ghi."
            )

        invoice_conflict_types = {
            ConflictType.MULTIPLE_SOURCE_INVOICES,
            ConflictType.INVOICE_VALUE_CONFLICT,
            ConflictType.INVOICE_COLUMN_MISSING,
        }
        amount_conflict_types = {
            ConflictType.TARGET_CELL_OCCUPIED,
            ConflictType.TARGET_CELL_FORMULA,
            ConflictType.TARGET_CELL_TEXT,
        }
        actions: list[dict[str, Any]] = []
        history: list[dict[str, Any]] = []
        for item_index, item in enumerate(items):
            selected_fee = item.selected_fee
            selected_row = item.target_row
            selected_invoice = item.selected_invoice
            amount_action = item.action
            invoice_action = item.invoice_action
            skip_status: PostingItemStatus | None = (
                item.status
                if item.status
                in {
                    PostingItemStatus.USER_SKIPPED,
                    PostingItemStatus.NOT_MATCHED,
                    PostingItemStatus.UNRESOLVED,
                }
                else None
            )

            for conflict in conflicts_by_item.get(item_index, ()):
                value = resolutions.get(conflict.conflict_id)
                action = (
                    self._action(value)
                    if value is not None
                    else conflict.default_action
                )
                if action in {ResolutionAction.CANCEL, ResolutionAction.CANCEL_ALL}:
                    raise ExpensePostingError("Người dùng đã hủy nhập khoản chi.")
                if conflict.conflict_type in invoice_conflict_types:
                    if action is ResolutionAction.SELECT_INVOICE:
                        selected = self._resolution_attr(value, "selected_invoice")
                        if _invoice_key(selected) not in {
                            _invoice_key(candidate)
                            for candidate in item.invoice_candidates
                        }:
                            raise ExpensePostingError("Số HĐ được chọn không hợp lệ.")
                        selected_invoice = str(selected)
                    elif action in {
                        ResolutionAction.KEEP_EXISTING,
                        ResolutionAction.OVERWRITE,
                        ResolutionAction.SKIP_INVOICE,
                    }:
                        invoice_action = action
                    continue
                if action is ResolutionAction.SELECT_FEE:
                    selected_fee = self._resolution_attr(value, "selected_fee")
                elif action is ResolutionAction.SELECT_ROW:
                    row_value = self._resolution_attr(value, "selected_row")
                    selected_row = int(row_value) if row_value is not None else None
                elif conflict.conflict_type in amount_conflict_types and action in {
                    ResolutionAction.OVERWRITE,
                    ResolutionAction.KEEP_EXISTING,
                    ResolutionAction.KEEP_FORMULA,
                    ResolutionAction.SKIP,
                }:
                    amount_action = action
                if conflict.conflict_type not in amount_conflict_types and action in {
                    None,
                    ResolutionAction.SKIP,
                    ResolutionAction.KEEP_EXISTING,
                    ResolutionAction.KEEP_FORMULA,
                }:
                    skip_status = PostingItemStatus.USER_SKIPPED

            if selected_fee not in fee_columns:
                skip_status = PostingItemStatus.UNRESOLVED
            if selected_row is None:
                selected = self._automatic_row(index.get(item.container or "", ()))
                selected_row = selected.row if selected else None
            if selected_row is None:
                skip_status = PostingItemStatus.NOT_MATCHED

            column = fee_columns.get(selected_fee)
            current_value = (
                worksheet.cell(selected_row, column).value
                if selected_row is not None and column is not None
                else None
            )
            amount_after = current_value
            amount_status = skip_status
            recorded_amount_action = amount_action or ResolutionAction.SKIP
            amount_write = False
            if skip_status is None and column is not None and selected_row is not None:
                cell = worksheet.cell(selected_row, column)
                state = classify_target_cell(cell, item.amount)
                if state.kind is TargetCellKind.SAME_VALUE and not item.force_repost:
                    recorded_amount_action = ResolutionAction.KEEP_EXISTING
                    amount_status = PostingItemStatus.ALREADY_EXISTS
                else:
                    chosen = amount_action
                    if state.kind in {TargetCellKind.EMPTY, TargetCellKind.ZERO} or (
                        state.kind is TargetCellKind.SAME_VALUE and item.force_repost
                    ):
                        chosen = ResolutionAction.OVERWRITE
                    if chosen is ResolutionAction.OVERWRITE:
                        amount_after = item.amount
                        amount_write = True
                        amount_status = PostingItemStatus.POSTED
                    else:
                        recorded_amount_action = chosen or ResolutionAction.SKIP
                        amount_status = PostingItemStatus.USER_SKIPPED
                    recorded_amount_action = chosen or recorded_amount_action

            invoice_column: int | None = None
            invoice_cell: Any = None
            invoice_before: Any = None
            invoice_after: Any = None
            invoice_write = False
            recorded_invoice_action = invoice_action
            if (
                skip_status is None
                and selected_row is not None
                and selected_fee != "LL"
                and item.invoice_candidates
            ):
                if selected_invoice is None:
                    if len(item.invoice_candidates) > 1:
                        raise ExpensePostingError(
                            "Chưa chọn Số HĐ cho khoản có nhiều hóa đơn."
                        )
                    selected_invoice = item.invoice_candidates[0]
                invoice_column = invoice_columns.get(str(selected_fee))
                if invoice_column is not None:
                    invoice_cell = worksheet.cell(selected_row, invoice_column)
                    invoice_before = invoice_cell.value
                    invoice_after = invoice_before
                    if _invoice_key(invoice_before) == _invoice_key(selected_invoice):
                        recorded_invoice_action = ResolutionAction.KEEP_EXISTING
                    elif _invoice_text(invoice_before) is None:
                        recorded_invoice_action = ResolutionAction.OVERWRITE
                        invoice_after = selected_invoice
                        invoice_write = True
                    elif invoice_action is ResolutionAction.OVERWRITE:
                        recorded_invoice_action = ResolutionAction.OVERWRITE
                        invoice_after = selected_invoice
                        invoice_write = True
                    else:
                        recorded_invoice_action = (
                            ResolutionAction.SKIP_INVOICE
                            if invoice_action is ResolutionAction.SKIP_INVOICE
                            else ResolutionAction.KEEP_EXISTING
                        )
                else:
                    recorded_invoice_action = ResolutionAction.SKIP_INVOICE

            final_status = amount_status or PostingItemStatus.USER_SKIPPED
            if amount_write or invoice_write:
                final_status = PostingItemStatus.POSTED
            action_record = self._action_record(
                item,
                str(selected_fee),
                selected_row,
                column,
                current_value,
                amount_after,
                recorded_amount_action,
                final_status,
            )
            action_record.update(
                {
                    "amount_write": amount_write,
                    "invoice_no": item.invoice_candidates[0]
                    if len(item.invoice_candidates) == 1
                    else None,
                    "invoice_selected": selected_invoice,
                    "invoice_target_column": invoice_column,
                    "invoice_target_cell": (
                        invoice_cell.coordinate if invoice_cell is not None else None
                    ),
                    "invoice_value_before": invoice_before,
                    "invoice_value_after": invoice_after,
                    "invoice_action": recorded_invoice_action,
                    "invoice_write": invoice_write,
                }
            )
            actions.append(action_record)
            history.extend(self._history_rows(action_record, item))
        return actions, history

    @staticmethod
    def _action_record(
        item: PostingItem,
        fee: str,
        row: int | None,
        column: int | None,
        before: Any,
        after: Any,
        action: ResolutionAction,
        status: PostingItemStatus,
    ) -> dict[str, Any]:
        return {
            "source_indices": list(item.source_indices),
            "container": item.container,
            "fee_original": item.original_fee,
            "fee_selected": fee,
            "sheet_name": item.sheet_name,
            "selected_source_sheet": item.selected_source_sheet,
            "selected_source_row": item.selected_source_row,
            "source_sqt": item.source_sqt,
            "plan_values": tuple(item.plan_values),
            "source_signature": item.source_signature,
            "carry_forward_required": item.carry_forward_required,
            "target_row": row,
            "target_column": column,
            "target_cell": (
                f"{get_column_letter(column)}{row}" if row and column else None
            ),
            "value_before": before,
            "value_after": after,
            "action": action,
            "status": status,
        }

    @staticmethod
    def _history_rows(
        action: Mapping[str, Any], item: PostingItem
    ) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        sources = {
            source["source_item_index"]: source for source in item.source_items
        }
        for source_index in item.source_indices:
            source = sources.get(source_index, {})
            result.append(
                {
                    "source_item_index": source_index,
                    "container": source.get("container", item.container),
                    "bl": source.get("bl", item.bl),
                    "fee_original": source.get("fee", item.original_fee),
                    "fee_selected": action["fee_selected"],
                    "rule": source.get("rule", item.rule),
                    "amount": source.get("amount", item.amount),
                    "sheet_name": action["sheet_name"],
                    "target_row": action["target_row"],
                    "target_column": action["target_column"],
                    "target_cell": action["target_cell"],
                    "value_before": action["value_before"],
                    "value_after": action["value_after"],
                    "action": action["action"],
                    "invoice_no": source.get("invoice_no"),
                    "invoice_selected": action.get("invoice_selected"),
                    "invoice_target_column": action.get("invoice_target_column"),
                    "invoice_target_cell": action.get("invoice_target_cell"),
                    "invoice_value_before": action.get("invoice_value_before"),
                    "invoice_value_after": action.get("invoice_value_after"),
                    "invoice_action": action.get("invoice_action"),
                    "status": action["status"],
                }
            )
        return result

    def _write_carried_plan_rows(
        self,
        worksheet: Any,
        actions: Sequence[Mapping[str, Any]],
    ) -> list[Mapping[str, Any]]:
        unique: dict[tuple[str, int, int | None, str | None], Mapping[str, Any]] = {}
        for action in actions:
            if not action.get("carry_forward_required"):
                continue
            key = (
                str(action.get("selected_source_sheet") or ""),
                int(action.get("selected_source_row") or 0),
                action.get("source_sqt"),
                action.get("container"),
            )
            unique.setdefault(key, action)
        carried = sorted(unique.values(), key=lambda value: int(value["target_row"]))
        if not carried:
            return []

        target_header = self._plan_header(worksheet)
        last_data_row = DailySyncService._last_data_row(worksheet, target_header)
        expected_rows = list(
            range(last_data_row + 1, last_data_row + len(carried) + 1)
        )
        actual_rows = [int(action["target_row"]) for action in carried]
        if actual_rows != expected_rows:
            raise ExpensePostingError(
                "Vị trí dòng kế hoạch cần mang sang đã thay đổi; hãy phân tích lại."
            )
        template_row = max(target_header.row_end + 1, last_data_row)
        max_column = DailySyncService._actual_max_column(worksheet)
        for action in carried:
            target_row = int(action["target_row"])
            values = tuple(action.get("plan_values") or ())
            if len(values) != len(SYNC_FIELDS):
                raise ExpensePostingError("Dòng nguồn không đủ 12 trường kế hoạch.")
            DailySyncService._copy_row_style(
                worksheet,
                template_row,
                target_row,
                max_column=max_column,
            )
            for index, field in enumerate(SYNC_FIELDS):
                worksheet.cell(
                    target_row, TARGET_EXPECTED_COLUMNS[field]
                ).value = values[index]
        return carried

    def _record_carry_forwards(
        self,
        plan: PostingPlan,
        actions: Sequence[Mapping[str, Any]],
    ) -> None:
        saver = getattr(self.posting_repository, "save_carry_forward", None)
        if not callable(saver):
            return
        saved: set[tuple[str, int, int | None, str | None]] = set()
        for action in actions:
            source_sheet = str(action.get("selected_source_sheet") or "")
            source_row = int(action.get("selected_source_row") or 0)
            source_sqt = action.get("source_sqt")
            container = action.get("container")
            if (
                not source_sheet
                or source_sheet == action.get("sheet_name")
                or source_row <= 0
                or not isinstance(source_sqt, int)
                or source_sqt <= 0
                or not container
            ):
                continue
            key = (source_sheet, source_row, source_sqt, str(container))
            if key in saved:
                continue
            saver(
                workbook_path=plan.target_path,
                source_sheet=source_sheet,
                source_row=source_row,
                source_sqt=source_sqt,
                container=str(container),
                source_signature=str(action.get("source_signature") or ""),
                target_sheet=str(action.get("sheet_name") or ""),
                target_row=int(action["target_row"]),
            )
            saved.add(key)

    def _verify_posting(
        self,
        path: Path,
        sheet_name: str,
        actions: Sequence[Mapping[str, Any]],
        *,
        update_column: int | None = None,
        update_timestamp: datetime | None = None,
    ) -> None:
        workbook = self.gateway.load(path, read_only=False)
        try:
            worksheet = workbook[sheet_name]
            base = self._resolve_base_headers(worksheet)
            fee_columns = self._resolve_fee_columns(worksheet, base)
            invoice_columns = self._resolve_invoice_columns(base, fee_columns)
            target_plan_header = (
                self._plan_header(worksheet)
                if any(action.get("carry_forward_required") for action in actions)
                else None
            )
            for action in actions:
                if action.get("carry_forward_required"):
                    actual_plan_values = self._row_plan_values(
                        worksheet,
                        target_plan_header,
                        int(action["target_row"]),
                    )
                    if actual_plan_values != tuple(action.get("plan_values") or ()):
                        raise ExpensePostingError(
                            f"Dòng kế hoạch {action['target_row']} không qua verify."
                        )
                fee = action["fee_selected"]
                column = int(action["target_column"])
                if fee_columns.get(fee) != column:
                    raise ExpensePostingError(f"Cột phí {fee} không qua verify.")
                actual = worksheet.cell(int(action["target_row"]), column).value
                if action.get("amount_write") and actual != action["value_after"]:
                    raise ExpensePostingError(
                        f"Ô {action['target_cell']} không có giá trị dự kiến."
                    )
                if action.get("invoice_write"):
                    invoice_column = int(action["invoice_target_column"])
                    if invoice_columns.get(fee) != invoice_column:
                        raise ExpensePostingError(
                            f"Cột Số HĐ của phí {fee} không qua verify."
                        )
                    invoice_actual = worksheet.cell(
                        int(action["target_row"]), invoice_column
                    ).value
                    if invoice_actual != action["invoice_value_after"]:
                        raise ExpensePostingError(
                            f"Ô {action['invoice_target_cell']} không có Số HĐ dự kiến."
                        )
            if update_column is None or update_timestamp is None:
                raise ExpensePostingError("Bản lưu thiếu thông tin Date cập nhật.")
            actual_update_column = self._update_column(base)
            if actual_update_column != update_column:
                raise ExpensePostingError("Cột Date cập nhật không qua verify.")
            for target_row in {
                int(action["target_row"]) for action in actions
            }:
                actual = worksheet.cell(target_row, update_column).value
                if actual != update_timestamp:
                    raise ExpensePostingError(
                        f"Date cập nhật tại dòng {target_row} không đúng."
                    )
        finally:
            workbook.close()

    def _record_history(
        self, plan: PostingPlan, history: Sequence[Mapping[str, Any]]
    ) -> None:
        if (
            self.posting_repository is None
            or plan.run_id is None
            or plan.batch_id is None
            or not history
        ):
            return
        self.posting_repository.create_items(
            history,
            run_id=plan.run_id,
            batch_id=plan.batch_id,
            batch_hash=plan.batch_hash,
        )

    def _ready_path(self, batch_id: int | None) -> Path | None:
        if batch_id is None:
            return self.provider.get_latest_ready_json_path()
        return self.provider.get_ready_json_path(batch_id)

    def _batch_id_for_path(self, path: Path) -> int | None:
        list_method = getattr(self.provider, "list_ready_batches", None)
        if not callable(list_method):
            return None
        resolved = path.resolve()
        for metadata in list_method():
            current_path = getattr(metadata, "source_output_path", None)
            if current_path is not None and Path(current_path).resolve() == resolved:
                return int(getattr(metadata, "id", getattr(metadata, "batch_id")))
        return None

    def _successful_indices(self, batch_hash: str) -> set[int]:
        if self.posting_repository is None:
            return set()
        return set(self.posting_repository.successful_source_indices(batch_hash))

    def _previously_posted_items(
        self,
        batch_hash: str,
        document_rows: Sequence[Mapping[str, Any]],
        successful_indices: set[int],
    ) -> list[dict[str, Any]]:
        records: Sequence[Any] = ()
        latest = getattr(
            self.posting_repository,
            "latest_successful_items",
            None,
        )
        if callable(latest):
            records = latest(batch_hash)
        by_index = {
            int(getattr(record, "source_item_index")): record
            for record in records
        }
        result: list[dict[str, Any]] = []
        for source_index in sorted(successful_indices):
            if source_index >= len(document_rows):
                continue
            source = document_rows[source_index]
            record = by_index.get(source_index)

            def previous(name: str, default: Any = None) -> Any:
                return getattr(record, name, default) if record is not None else default

            result.append(
                {
                    "source_item_index": source_index,
                    "container": previous("container", source.get("container")),
                    "bl": previous("bl", source.get("bl")),
                    "fee": previous(
                        "fee_selected",
                        source.get("fee"),
                    )
                    or source.get("fee"),
                    "amount": previous("amount", source.get("amount")),
                    "sheet_name": previous("sheet_name"),
                    "target_row": previous("target_row"),
                    "target_cell": previous("target_cell"),
                    "created_at": previous("created_at"),
                }
            )
        return result

    def _latest_sync_sheet(self) -> str | None:
        if self.run_repository is None:
            return None
        return self.run_repository.get_latest_sync_sheet()

    @staticmethod
    def _positive_int(value: Any) -> int | None:
        if isinstance(value, bool):
            return None
        if isinstance(value, int) and value > 0:
            return value
        if isinstance(value, float) and value.is_integer() and value > 0:
            return int(value)
        if isinstance(value, str) and value.strip().isdigit():
            parsed = int(value.strip())
            return parsed if parsed > 0 else None
        return None

    def _selected_sheet(
        self, plan: PostingPlan, resolutions: Mapping[str, Any]
    ) -> str | None:
        if plan.selected_sheet is not None:
            return plan.selected_sheet
        for conflict in plan.conflicts:
            if conflict.conflict_type is ConflictType.TARGET_SHEET_AMBIGUOUS:
                value = resolutions.get(conflict.conflict_id)
                if value is None:
                    continue
                action = self._action(value)
                if action in {ResolutionAction.CANCEL, ResolutionAction.CANCEL_ALL}:
                    raise ExpensePostingError("Người dùng đã hủy nhập khoản chi.")
                if action is ResolutionAction.SELECT_SHEET:
                    selected = self._resolution_attr(value, "selected_sheet")
                    valid = {c.target_sheet for c in plan.sheet_candidates}
                    if selected not in valid:
                        raise ExpensePostingError("Sheet được chọn không hợp lệ.")
                    return str(selected)
        return None

    def _has_selector_resolution(
        self,
        plan: PostingPlan,
        resolutions: Mapping[str, Any],
    ) -> bool:
        for conflict in plan.conflicts:
            value = resolutions.get(conflict.conflict_id)
            if value is None:
                continue
            if self._action(value) in {
                ResolutionAction.SELECT_SHEET,
                ResolutionAction.SELECT_ROW,
                ResolutionAction.SELECT_FEE,
                ResolutionAction.SELECT_INVOICE,
                ResolutionAction.SELECT_SOURCE_ITEM,
            }:
                return True
        return False

    def _check_batch_resolution(
        self, plan: PostingPlan, resolutions: Mapping[str, Any]
    ) -> None:
        for conflict in plan.conflicts:
            if conflict.conflict_type is not ConflictType.BATCH_ALREADY_POSTED:
                continue
            value = resolutions.get(conflict.conflict_id)
            action = self._action(value) if value is not None else conflict.default_action
            if action in {ResolutionAction.CANCEL, ResolutionAction.CANCEL_ALL}:
                raise ExpensePostingError("Người dùng đã hủy nhập lại batch.")

    @staticmethod
    def _action(value: Any) -> ResolutionAction:
        if isinstance(value, (ResolutionAction, str)):
            return ResolutionAction(value)
        action = getattr(value, "action", None)
        if isinstance(value, Mapping):
            action = value.get("action", action)
        return ResolutionAction(action)

    @staticmethod
    def _resolution_attr(value: Any, name: str) -> Any:
        result = getattr(value, name, None)
        if isinstance(value, Mapping):
            result = value.get(name, result)
        return result

    def _create_run(self, source: Path, target: Path) -> int | None:
        if self.run_repository is None:
            return None
        record = self.run_repository.create_run(
            operation=ExcelOperation.EXPENSE_POSTING,
            source_path=source,
            target_path=target,
            status=ExcelRunStatus.ANALYZING,
        )
        return int(getattr(record, "id", record))

    def _update_run(self, run_id: int | None, **changes: Any) -> None:
        if self.run_repository is not None and run_id is not None:
            self.run_repository.update_run(run_id, **changes)

    def _finish_result(self, result: PostingResult) -> None:
        if self.run_repository is None or result.run_id is None:
            return
        self.run_repository.finish_run(
            result.run_id,
            status=result.status,
            sheet_name=result.sheet_name,
            backup_path=result.backup_path,
            target_fingerprint_after=result.fingerprint_after,
            total_items=(
                result.posted_source_items
                + result.skipped_source_items
                + result.already_existing_items
            ),
            changed_items=result.posted_source_items,
            skipped_items=result.skipped_source_items,
            conflict_count=result.conflict_count,
        )

    def _finish_failed(self, run_id: int | None, exc: Exception) -> None:
        if self.run_repository is not None and run_id is not None:
            cancelled = "người dùng đã hủy" in str(exc).casefold()
            self.run_repository.finish_run(
                run_id,
                status=(
                    ExcelRunStatus.CANCELLED
                    if cancelled
                    else ExcelRunStatus.FAILED
                ),
                error_message=None if cancelled else str(exc),
            )
