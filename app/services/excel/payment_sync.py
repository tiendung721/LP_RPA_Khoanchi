"""Chuẩn hóa BK và đồng bộ các khoản chi được phép sang workbook Thanh toán."""

from __future__ import annotations

import ast
import copy
import hashlib
import json
import re
import shutil
from collections import defaultdict
from collections.abc import Callable, Mapping, Sequence
from contextlib import ExitStack
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, DivisionByZero, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, column_index_from_string

from app.services.file_stability import FileStabilityChecker

from .headers import normalize_header
from .models import (
    ConflictType,
    ExcelOperation,
    ExcelRunStatus,
    PaymentSyncConflict,
    PaymentSyncItem,
    PaymentSyncPlan,
    PaymentSyncResult,
    PaymentTargetPlan,
    PaymentTargetResult,
    ResolutionAction,
    RowCandidate,
    SourceSheetCandidate,
)
from .resolvers import MonthSheetService
from .workbook import (
    ExcelBackupService,
    ExcelLockService,
    WorkbookGateway,
    ensure_supported_workbook,
    workbook_has_vba,
)


ProgressCallback = Callable[[str], None] | None
DATE_HEADER = "Date cập nhật"
DATE_NUMBER_FORMAT = "dd/mm/yyyy hh:mm:ss"

SUMMARY_HEADERS: tuple[str, ...] = (
    "QT",
    "CƯỚC MB",
    "N.HẠ MB",
    "CƯỚC BIỂN",
    "N.HA VS D/O LỆNH",
    "CƯỚC MN",
    "Lưu cont",
    "Sửa chữa Cont",
    "QUÁ TẢI",
    "",
)
_SUMMARY_SIGNATURE = tuple(normalize_header(value) for value in SUMMARY_HEADERS[:-1])

BK_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "sqt": ("SQT", "SQT PM", "Số thứ tự PM"),
    "container": ("Số Container", "Container", "Số cont"),
    "sea_freight": ("Cước biển",),
    "north_freight": ("Cước bộ đóng hàng", "ĐƠN GIÁ"),
    "empty_lift": ("Nâng vỏ",),
    "loaded_drop": ("Hạ Hàng",),
    "loaded_lift": ("Nâng Hàng",),
    "empty_drop": ("Hạ vỏ",),
    "south_freight": ("Cước VTN",),
    "storage": ("Lưu cont",),
    "overweight": ("Quá tải",),
    "vs_do": ("VS + D/O", "VS D/O LỆNH", "VS DO LỆNH"),
    "command_fee": ("LÀM LỆNH", "Làm lệnh"),
    "repair": ("SỬA CHỮA", "Sửa chữa"),
}

PAYMENT_FIELDS: tuple[str, ...] = (
    "sea_freight",
    "north_freight",
    "empty_lift",
    "loaded_drop",
    "loaded_lift",
    "empty_drop",
    "south_freight",
    "storage",
    "overweight",
    "vs_do",
    "command_fee",
    "repair",
)

HP_FIELDS: tuple[str, ...] = ("loaded_drop",)
NAM_FIELDS: tuple[str, ...] = (
    "empty_lift",
    "loaded_lift",
    "empty_drop",
    "vs_do",
    "command_fee",
    "storage",
    "repair",
    "overweight",
)
SYNC_SOURCE_FIELDS: tuple[str, ...] = ("sqt", "container", *HP_FIELDS, *NAM_FIELDS)

TARGET_DETAIL_HEADERS: dict[str, tuple[str, ...]] = {
    "sqt": ("QT",),
    "container": ("SỐ CONT", "Số Container"),
    "empty_lift": ("NÂNG VỎ",),
    "loaded_drop": ("HẠ HÀNG",),
    "sea_freight": ("Cước biển",),
    "loaded_lift": ("NÂNG HÀNG",),
    "empty_drop": ("HẠ VỎ",),
    "vs_do": ("VS + D/O", "VS D/O"),
    "command_fee": ("LÀM LỆNH",),
    "storage": ("Lưu Cont", "Lưu container"),
    "repair": ("Sửa chữa Cont", "SỬA CHỮA"),
    "overweight": ("QUÁ TẢI",),
}

_CELL_REF_RE = re.compile(
    r"(?<![A-Z0-9_])(?P<col_abs>\$?)(?P<col>[A-Z]{1,3})"
    r"(?P<row_abs>\$?)(?P<row>[1-9][0-9]*)",
    re.IGNORECASE,
)
_RANGE_RE = re.compile(
    r"(?P<left>\$?[A-Z]{1,3}\$?[1-9][0-9]*):"
    r"(?P<right>\$?[A-Z]{1,3}\$?[1-9][0-9]*)",
    re.IGNORECASE,
)
_PAYMENT_TITLE_MONTH_RE = re.compile(
    r"(tháng\s*)0?(?:1[0-2]|[1-9])\s*/\s*(?:\d{2}|\d{4})",
    re.IGNORECASE,
)
_PAYMENT_MANUAL_SECTION = normalize_header("Số tiền thanh toán")


class PaymentSyncError(RuntimeError):
    pass


@dataclass(slots=True)
class NormalizationIssue:
    sheet_name: str
    row: int
    message: str


@dataclass(slots=True)
class NormalizationReport:
    changed: bool = False
    changed_sheets: set[str] = field(default_factory=set)
    issues: list[NormalizationIssue] = field(default_factory=list)

    def add_change(self, sheet_name: str) -> None:
        self.changed = True
        self.changed_sheets.add(sheet_name)


def _progress(callback: ProgressCallback, message: str) -> None:
    if callback is not None:
        callback(message)


def _setting(settings: Any | None, name: str, default: Any = "") -> Any:
    return getattr(settings, name, default) if settings is not None else default


def _parse_sqt(value: Any) -> int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    if isinstance(value, float) and value.is_integer():
        parsed = int(value)
        return parsed if parsed > 0 else None
    if isinstance(value, str):
        text = "".join(value.split())
        if text.isdigit():
            parsed = int(text)
            return parsed if parsed > 0 else None
    return None


def _container_key(value: Any) -> str:
    if value is None:
        return ""
    return "".join(str(value).split()).upper()


def _stable_id(*parts: Any) -> str:
    raw = json.dumps(parts, ensure_ascii=False, default=str, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]


def _numeric_equal(left: Any, right: Any) -> bool:
    if left in (None, ""):
        left = 0
    if right in (None, ""):
        right = 0
    try:
        return Decimal(str(left)) == Decimal(str(right))
    except (InvalidOperation, ValueError):
        return left == right


def _number(value: Decimal) -> int | float:
    integral = value.to_integral_value()
    if value == integral:
        return int(integral)
    return float(value)


class ArithmeticFormulaEvaluator:
    """Tính công thức số học đơn giản và tham chiếu ô mà không dùng ``eval``."""

    def __init__(self, worksheet: Any) -> None:
        self.worksheet = worksheet
        self._cache: dict[str, Decimal] = {}
        self._active: set[str] = set()

    def value(self, cell_or_value: Any) -> int | float | None:
        if hasattr(cell_or_value, "coordinate"):
            value = self._cell_decimal(cell_or_value.coordinate)
            raw = cell_or_value.value
            if raw in (None, ""):
                return None
            return _number(value)
        if cell_or_value in (None, ""):
            return None
        return _number(self._scalar_decimal(cell_or_value))

    def _cell_decimal(self, coordinate: str) -> Decimal:
        key = coordinate.upper()
        if key in self._cache:
            return self._cache[key]
        if key in self._active:
            raise PaymentSyncError(f"Công thức tham chiếu vòng tại ô {coordinate}.")
        self._active.add(key)
        try:
            result = self._scalar_decimal(self.worksheet[coordinate].value)
        finally:
            self._active.remove(key)
        self._cache[key] = result
        return result

    def _scalar_decimal(self, value: Any) -> Decimal:
        if value in (None, ""):
            return Decimal(0)
        if isinstance(value, bool):
            raise PaymentSyncError("Giá trị boolean không phải số tiền.")
        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value))
        if not isinstance(value, str) or not value.startswith("="):
            raise PaymentSyncError(f"Giá trị {value!r} không phải số hoặc công thức.")
        expression = value[1:].strip()
        if not expression:
            return Decimal(0)
        if re.search(r"\b(?:SUM|SUBTOTAL|IF)\s*\(", expression, re.IGNORECASE):
            return self._function_formula(expression)

        def replace_reference(match: re.Match[str]) -> str:
            coordinate = f"{match.group('col')}{match.group('row')}"
            return str(self._cell_decimal(coordinate))

        translated = _CELL_REF_RE.sub(replace_reference, expression)
        try:
            node = ast.parse(translated, mode="eval")
        except SyntaxError as exc:
            raise PaymentSyncError(f"Công thức không được hỗ trợ: ={expression}") from exc
        try:
            return self._ast_decimal(node.body)
        except (DivisionByZero, ZeroDivisionError, InvalidOperation) as exc:
            raise PaymentSyncError(f"Công thức không tính được: ={expression}") from exc

    def _function_formula(self, expression: str) -> Decimal:
        sum_match = re.fullmatch(
            r"SUM\(\s*(\$?[A-Z]{1,3}\$?[1-9][0-9]*)"
            r"\s*:\s*(\$?[A-Z]{1,3}\$?[1-9][0-9]*)\s*\)",
            expression,
            re.IGNORECASE,
        )
        if sum_match:
            start, end = sum_match.groups()
            total = Decimal(0)
            for row in self.worksheet[start:end]:
                for cell in row:
                    total += self._cell_decimal(cell.coordinate)
            return total
        raise PaymentSyncError(f"Hàm Excel chưa được hỗ trợ: ={expression}")

    def _ast_decimal(self, node: ast.AST) -> Decimal:
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return Decimal(str(node.value))
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
            value = self._ast_decimal(node.operand)
            return value if isinstance(node.op, ast.UAdd) else -value
        if isinstance(node, ast.BinOp):
            left = self._ast_decimal(node.left)
            right = self._ast_decimal(node.right)
            if isinstance(node.op, ast.Add):
                return left + right
            if isinstance(node.op, ast.Sub):
                return left - right
            if isinstance(node.op, ast.Mult):
                return left * right
            if isinstance(node.op, ast.Div):
                return left / right
        raise PaymentSyncError("Công thức chứa phép toán không được hỗ trợ.")


def _header_values(worksheet: Any, row: int = 1) -> dict[int, str]:
    return {
        column: normalize_header(worksheet.cell(row, column).value)
        for column in range(1, _effective_max_column(worksheet) + 1)
    }


def find_summary_start(worksheet: Any, *, header_row: int = 1) -> int | None:
    headers = _header_values(worksheet, header_row)
    limit = max(headers, default=0) - len(_SUMMARY_SIGNATURE) + 1
    for start in range(1, limit + 1):
        if tuple(
            headers.get(start + offset, "")
            for offset in range(len(_SUMMARY_SIGNATURE))
        ) == _SUMMARY_SIGNATURE:
            return start
    return None


def _matching_columns(
    headers: Mapping[int, str],
    aliases: Sequence[str],
    *,
    before: int | None = None,
) -> list[int]:
    alias_keys = {normalize_header(value) for value in aliases}
    return [
        column
        for column, value in headers.items()
        if value in alias_keys and (before is None or column < before)
    ]


def _required_column(
    headers: Mapping[int, str],
    aliases: Sequence[str],
    field_label: str,
    *,
    before: int | None = None,
) -> int:
    matches = _matching_columns(headers, aliases, before=before)
    if len(matches) != 1:
        raise PaymentSyncError(
            f"Không xác định duy nhất cột {field_label}: "
            f"{len(matches)} cột phù hợp."
        )
    return matches[0]


def _copy_cell_style(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy.copy(source._style)
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)
    target.alignment = copy.copy(source.alignment)
    target.protection = copy.copy(source.protection)
    target.number_format = source.number_format


def _map_formula_columns(worksheet: Any, mapper: Callable[[int], int]) -> None:
    for cell in tuple(getattr(worksheet, "_cells", {}).values()):
        value = getattr(cell, "value", None)
        if not isinstance(value, str) or not value.startswith("="):
            continue

        def replace(match: re.Match[str]) -> str:
            column = column_index_from_string(match.group("col").upper())
            mapped = mapper(column)
            return (
                f"{match.group('col_abs')}{get_column_letter(mapped)}"
                f"{match.group('row_abs')}{match.group('row')}"
            )

        cell.value = _CELL_REF_RE.sub(replace, value)


def _insert_column(worksheet: Any, column: int) -> None:
    _map_formula_columns(
        worksheet, lambda current: current + 1 if current >= column else current
    )
    worksheet.insert_cols(column)


def _delete_column(worksheet: Any, column: int, *, replacement: int) -> None:
    def mapper(current: int) -> int:
        if current == column:
            return replacement
        return current - 1 if current > column else current

    _map_formula_columns(worksheet, mapper)
    worksheet.delete_cols(column)


def _data_rows(
    worksheet: Any,
    *,
    sqt_column: int,
    container_column: int,
    header_row: int = 1,
) -> list[int]:
    rows: list[int] = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        if (
            _parse_sqt(worksheet.cell(row, sqt_column).value) is not None
            and _container_key(worksheet.cell(row, container_column).value)
        ):
            rows.append(row)
    return rows


def _migrated_vs_formula(value: Any) -> Any:
    if isinstance(value, str) and value.startswith("="):
        return f"=({value[1:]})-150000"
    if value in (None, ""):
        return value
    return _number(Decimal(str(value)) - Decimal(150000))


def _ensure_bk_structure(
    worksheet: Any,
    report: NormalizationReport,
) -> None:
    summary_start = find_summary_start(worksheet)
    headers = _header_values(worksheet)
    boundary = summary_start
    sqt_col = _required_column(headers, BK_HEADER_ALIASES["sqt"], "SQT", before=boundary)
    container_col = _required_column(
        headers, BK_HEADER_ALIASES["container"], "Số Container", before=boundary
    )
    vs_candidates = _matching_columns(
        headers, BK_HEADER_ALIASES["vs_do"], before=boundary
    )
    if len(vs_candidates) != 1:
        raise PaymentSyncError(
            f"Sheet {worksheet.title} không xác định duy nhất cột VS/D/O."
        )
    vs_col = vs_candidates[0]
    ll_columns = _matching_columns(
        headers, BK_HEADER_ALIASES["command_fee"], before=boundary
    )
    desired_ll = vs_col + 2
    if (
        ll_columns == [desired_ll]
        and normalize_header(worksheet.cell(1, desired_ll).value)
        == normalize_header("LÀM LỆNH")
    ):
        # The workbook is already on the split VS + D/O / LÀM LỆNH layout.
        # Values are business data and must not be forced to the legacy
        # migration assumption of exactly 150,000.
        _ensure_summary_block(worksheet, report)
        return
    rows = _data_rows(
        worksheet, sqt_column=sqt_col, container_column=container_col
    )
    evaluator = ArithmeticFormulaEvaluator(worksheet)
    migrations: dict[int, tuple[Any, Any]] = {}
    old_ll_values: dict[int, list[Any]] = {}
    for row in rows:
        values: list[Any] = []
        for column in ll_columns:
            raw = worksheet.cell(row, column).value
            if raw in (None, "", 0):
                continue
            try:
                parsed = evaluator.value(worksheet.cell(row, column))
            except PaymentSyncError as exc:
                report.issues.append(
                    NormalizationIssue(worksheet.title, row, str(exc))
                )
                continue
            if parsed not in (None, 0):
                values.append(parsed)
        old_ll_values[row] = values
        distinct = {Decimal(str(value)) for value in values}
        combined_cell = worksheet.cell(row, vs_col)
        combined_raw = combined_cell.value
        try:
            combined = evaluator.value(combined_cell)
        except PaymentSyncError as exc:
            report.issues.append(NormalizationIssue(worksheet.title, row, str(exc)))
            migrations[row] = (combined_raw, values[0] if values else None)
            continue
        if len(distinct) > 1 or (distinct and distinct != {Decimal(150000)}):
            report.issues.append(
                NormalizationIssue(
                    worksheet.title,
                    row,
                    "Cột LÀM LỆNH cũ có giá trị khác 150.000 hoặc mâu thuẫn.",
                )
            )
            migrations[row] = (combined_raw, values[0] if values else None)
        elif distinct == {Decimal(150000)}:
            migrations[row] = (combined_raw, 150000)
        elif combined in (None, 0):
            migrations[row] = (combined_raw, None if combined_raw in (None, "") else 0)
        elif Decimal(str(combined)) >= Decimal(150000):
            migrations[row] = (_migrated_vs_formula(combined_raw), 150000)
        else:
            report.issues.append(
                NormalizationIssue(
                    worksheet.title,
                    row,
                    "Tổng VS/D/O/Lệnh nhỏ hơn 150.000 nên không thể tự tách.",
                )
            )
            migrations[row] = (combined_raw, None)

    desired_already = (
        desired_ll in ll_columns
        and normalize_header(worksheet.cell(1, desired_ll).value)
        == normalize_header("LÀM LỆNH")
    )
    original_ll_columns = list(ll_columns)
    if not desired_already:
        _insert_column(worksheet, desired_ll)
        report.add_change(worksheet.title)
        original_ll_columns = [
            column + 1 if column >= desired_ll else column
            for column in original_ll_columns
        ]
        if summary_start is not None and summary_start >= desired_ll:
            summary_start += 1

    worksheet.cell(1, vs_col).value = "VS + D/O"
    worksheet.cell(1, desired_ll).value = "LÀM LỆNH"
    style_source_col = original_ll_columns[0] if original_ll_columns else vs_col
    _copy_cell_style(
        worksheet.cell(1, style_source_col), worksheet.cell(1, desired_ll)
    )

    delete_columns = sorted(
        {column for column in original_ll_columns if column != desired_ll},
        reverse=True,
    )
    for column in delete_columns:
        replacement = desired_ll
        _delete_column(worksheet, column, replacement=replacement)
        report.add_change(worksheet.title)
        if column < desired_ll:
            desired_ll -= 1
            vs_col -= 1
        if summary_start is not None and column < summary_start:
            summary_start -= 1

    for row, (vs_value, ll_value) in migrations.items():
        vs_cell = worksheet.cell(row, vs_col)
        ll_cell = worksheet.cell(row, desired_ll)
        if vs_cell.value != vs_value:
            vs_cell.value = vs_value
            report.add_change(worksheet.title)
        if ll_cell.value != ll_value:
            ll_cell.value = ll_value
            report.add_change(worksheet.title)
        _copy_cell_style(vs_cell, ll_cell)

    _ensure_summary_block(worksheet, report)


def _ensure_summary_block(
    worksheet: Any,
    report: NormalizationReport,
) -> int:
    summary_start = find_summary_start(worksheet)
    headers = _header_values(worksheet)
    if summary_start is None:
        used_headers = [column for column, value in headers.items() if value]
        summary_start = (max(used_headers) if used_headers else 0) + 1
        previous = worksheet.cell(1, max(1, summary_start - 1))
        for offset, label in enumerate(SUMMARY_HEADERS):
            cell = worksheet.cell(1, summary_start + offset)
            _copy_cell_style(previous, cell)
            cell.value = label or None
        report.add_change(worksheet.title)

    headers = _header_values(worksheet)
    raw: dict[str, int] = {}
    for field, aliases in BK_HEADER_ALIASES.items():
        raw[field] = _required_column(
            headers, aliases, field, before=summary_start
        )
    summary_columns = {
        "summary_sqt": summary_start,
        "north_freight": summary_start + 1,
        "north_handling": summary_start + 2,
        "sea_freight": summary_start + 3,
        "handling": summary_start + 4,
        "south_freight": summary_start + 5,
        "storage": summary_start + 6,
        "repair": summary_start + 7,
        "overweight": summary_start + 8,
        "storage_overweight": summary_start + 9,
    }
    rows = _data_rows(
        worksheet,
        sqt_column=raw["sqt"],
        container_column=raw["container"],
    )
    for row in rows:
        formulas = {
            "summary_sqt": f"={get_column_letter(raw['sqt'])}{row}",
            "north_freight": f"={get_column_letter(raw['north_freight'])}{row}",
            "north_handling": (
                f"={get_column_letter(raw['empty_lift'])}{row}+"
                f"{get_column_letter(raw['loaded_drop'])}{row}"
            ),
            "sea_freight": f"={get_column_letter(raw['sea_freight'])}{row}",
            "handling": (
                f"={get_column_letter(raw['loaded_lift'])}{row}+"
                f"{get_column_letter(raw['empty_drop'])}{row}+"
                f"{get_column_letter(raw['vs_do'])}{row}+"
                f"{get_column_letter(raw['command_fee'])}{row}"
            ),
            "south_freight": f"={get_column_letter(raw['south_freight'])}{row}",
            "storage": f"={get_column_letter(raw['storage'])}{row}",
            "repair": f"={get_column_letter(raw['repair'])}{row}",
            "overweight": f"={get_column_letter(raw['overweight'])}{row}",
            "storage_overweight": (
                f"={get_column_letter(summary_columns['storage'])}{row}+"
                f"{get_column_letter(summary_columns['overweight'])}{row}"
            ),
        }
        for field, formula in formulas.items():
            cell = worksheet.cell(row, summary_columns[field])
            if cell.value != formula:
                cell.value = formula
                report.add_change(worksheet.title)
            source = worksheet.cell(
                row,
                raw.get(
                    field,
                    raw.get(
                        "storage" if field == "storage_overweight" else field,
                        raw["sqt"],
                    ),
                ),
            )
            _copy_cell_style(source, cell)
            if field != "summary_sqt":
                cell.number_format = '#,##0'
    return summary_start


def normalize_bk_workbook(
    workbook: Any,
    *,
    month_service: MonthSheetService | None = None,
) -> NormalizationReport:
    months = month_service or MonthSheetService()
    report = NormalizationReport()
    for worksheet in workbook.worksheets:
        if months.parse_target_sheet(worksheet.title) is None:
            continue
        _ensure_bk_structure(worksheet, report)
    calculation = getattr(workbook, "calculation", None)
    if calculation is not None:
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True
        calculation.calcMode = "auto"
    return report


def refresh_bk_summary_formulas(worksheet: Any) -> bool:
    """Public helper used after Daily Sync adds BK rows."""

    report = NormalizationReport()
    _ensure_summary_block(worksheet, report)
    return report.changed


def _find_target_header_row(worksheet: Any) -> int:
    required = {
        normalize_header("QT"),
        normalize_header("SỐ CONT"),
        normalize_header("NÂNG VỎ"),
        normalize_header("HẠ HÀNG"),
    }
    for row in range(1, min(30, worksheet.max_row) + 1):
        values = {
            normalize_header(worksheet.cell(row, column).value)
            for column in range(1, min(80, worksheet.max_column) + 1)
        }
        if required.issubset(values):
            return row
    raise PaymentSyncError(
        f"Sheet {worksheet.title} không có vùng tiêu đề Thanh toán phù hợp."
    )


def _resolve_target_columns(worksheet: Any) -> tuple[int, dict[str, int], int]:
    header_row = _find_target_header_row(worksheet)
    summary_start = find_summary_start(worksheet, header_row=header_row)
    if summary_start is None:
        raise PaymentSyncError(
            f"Sheet {worksheet.title} thiếu khối cột tổng chuẩn."
        )
    headers = _header_values(worksheet, header_row)
    result: dict[str, int] = {}
    for field, aliases in TARGET_DETAIL_HEADERS.items():
        result[field] = _required_column(
            headers, aliases, field, before=summary_start
        )
    cước_bộ = _matching_columns(
        headers, ("Cước bộ",), before=summary_start
    )
    if len(cước_bộ) < 2:
        raise PaymentSyncError(
            f"Sheet {worksheet.title} phải có hai cột Cước bộ MB/MN."
        )
    result["north_freight"] = cước_bộ[0]
    result["south_freight"] = cước_bộ[1]
    result.update(
        {
            "summary_sqt": summary_start,
            "summary_north_freight": summary_start + 1,
            "summary_north_handling": summary_start + 2,
            "summary_sea_freight": summary_start + 3,
            "summary_handling": summary_start + 4,
            "summary_south_freight": summary_start + 5,
            "summary_storage": summary_start + 6,
            "summary_repair": summary_start + 7,
            "summary_overweight": summary_start + 8,
            "summary_storage_overweight": summary_start + 9,
        }
    )
    return header_row, result, summary_start


def _target_total_row(
    worksheet: Any,
    *,
    header_row: int,
    columns: Mapping[str, int],
) -> int:
    watched = {
        columns[field]
        for field in (
            "empty_lift",
            "loaded_drop",
            "loaded_lift",
            "empty_drop",
            "storage",
            "repair",
            "overweight",
        )
    }
    for row in range(header_row + 1, worksheet.max_row + 1):
        if _parse_sqt(worksheet.cell(row, columns["sqt"]).value) is not None:
            continue
        if _container_key(worksheet.cell(row, columns["container"]).value):
            continue
        for column in watched:
            value = worksheet.cell(row, column).value
            if isinstance(value, str) and re.match(
                r"^=(?:SUM|SUBTOTAL)\(", value, re.IGNORECASE
            ):
                return row
    value_rows = [
        row
        for (row, _column), cell in getattr(worksheet, "_cells", {}).items()
        if row > header_row and cell.value not in (None, "")
    ]
    return (max(value_rows) if value_rows else header_row) + 1


def _summary_formulas(row: int, columns: Mapping[str, int]) -> dict[int, str]:
    letter = get_column_letter
    return {
        columns["summary_sqt"]: f"={letter(columns['sqt'])}{row}",
        columns["summary_north_freight"]: (
            f"={letter(columns['north_freight'])}{row}"
        ),
        columns["summary_north_handling"]: (
            f"={letter(columns['empty_lift'])}{row}+"
            f"{letter(columns['loaded_drop'])}{row}"
        ),
        columns["summary_sea_freight"]: f"={letter(columns['sea_freight'])}{row}",
        columns["summary_handling"]: (
            f"={letter(columns['loaded_lift'])}{row}+"
            f"{letter(columns['empty_drop'])}{row}+"
            f"{letter(columns['vs_do'])}{row}+"
            f"{letter(columns['command_fee'])}{row}"
        ),
        columns["summary_south_freight"]: (
            f"={letter(columns['south_freight'])}{row}"
        ),
        columns["summary_storage"]: f"={letter(columns['storage'])}{row}",
        columns["summary_repair"]: f"={letter(columns['repair'])}{row}",
        columns["summary_overweight"]: f"={letter(columns['overweight'])}{row}",
        columns["summary_storage_overweight"]: (
            f"={letter(columns['summary_storage'])}{row}+"
            f"{letter(columns['summary_overweight'])}{row}"
        ),
    }


def _target_index(
    worksheet: Any,
    *,
    header_row: int,
    total_row: int,
    columns: Mapping[str, int],
) -> tuple[
    dict[tuple[int, str], list[int]],
    dict[int, list[int]],
    dict[str, list[int]],
]:
    exact: dict[tuple[int, str], list[int]] = defaultdict(list)
    by_sqt: dict[int, list[int]] = defaultdict(list)
    by_container: dict[str, list[int]] = defaultdict(list)
    for row in range(header_row + 1, total_row):
        sqt = _parse_sqt(worksheet.cell(row, columns["sqt"]).value)
        container = _container_key(
            worksheet.cell(row, columns["container"]).value
        )
        if sqt is None and not container:
            continue
        if sqt is not None:
            by_sqt[sqt].append(row)
        if container:
            by_container[container].append(row)
        if sqt is not None and container:
            exact[(sqt, container)].append(row)
    return exact, by_sqt, by_container


def _source_columns(worksheet: Any) -> tuple[dict[str, int], int]:
    summary_start = find_summary_start(worksheet)
    if summary_start is None:
        raise PaymentSyncError(
            f"Sheet {worksheet.title} chưa có khối cột tổng."
        )
    headers = _header_values(worksheet)
    columns = {
        field: _required_column(headers, aliases, field, before=summary_start)
        for field, aliases in BK_HEADER_ALIASES.items()
    }
    return columns, summary_start


def _source_items(
    worksheet: Any,
    *,
    normalization_issues: Sequence[NormalizationIssue],
) -> tuple[list[PaymentSyncItem], dict[str, str]]:
    columns, _summary_start = _source_columns(worksheet)
    evaluator = ArithmeticFormulaEvaluator(worksheet)
    grouped: dict[tuple[int, str], list[int]] = defaultdict(list)
    for row in _data_rows(
        worksheet,
        sqt_column=columns["sqt"],
        container_column=columns["container"],
    ):
        sqt = _parse_sqt(worksheet.cell(row, columns["sqt"]).value)
        container = _container_key(worksheet.cell(row, columns["container"]).value)
        if sqt is not None and container:
            grouped[(sqt, container)].append(row)
    issues_by_row = {
        issue.row: issue.message
        for issue in normalization_issues
        if issue.sheet_name == worksheet.title
    }
    items: list[PaymentSyncItem] = []
    item_errors: dict[str, str] = {}
    for (sqt, container), rows in grouped.items():
        item_id = _stable_id(worksheet.title, sqt, container, rows)
        values: dict[str, Any] = {}
        errors: list[str] = [
            issues_by_row[row] for row in rows if row in issues_by_row
        ]
        for field in PAYMENT_FIELDS:
            nonzero: list[Any] = []
            saw_zero = False
            for row in rows:
                cell = worksheet.cell(row, columns[field])
                if cell.value in (None, ""):
                    continue
                try:
                    value = evaluator.value(cell)
                except PaymentSyncError as exc:
                    errors.append(f"{field}: {exc}")
                    continue
                if value in (None, 0):
                    saw_zero = True
                else:
                    nonzero.append(value)
            if len(nonzero) > 1:
                errors.append(
                    f"{field}: có nhiều dòng cùng SQT/container chứa số tiền."
                )
            values[field] = nonzero[0] if nonzero else (0 if saw_zero else None)
        item = PaymentSyncItem(
            item_id=item_id,
            source_row=rows[0],
            source_rows=tuple(rows),
            sqt=sqt,
            container=container,
            values=values,
        )
        if errors:
            item.status = "INVALID"
            item_errors[item_id] = " • ".join(dict.fromkeys(errors))
        items.append(item)
    return items, item_errors


def _analyze_target(
    worksheet: Any,
    items: list[PaymentSyncItem],
    item_errors: Mapping[str, str],
) -> list[PaymentSyncConflict]:
    header_row, columns, _summary_start = _resolve_target_columns(worksheet)
    total_row = _target_total_row(
        worksheet, header_row=header_row, columns=columns
    )
    exact, by_sqt, by_container = _target_index(
        worksheet,
        header_row=header_row,
        total_row=total_row,
        columns=columns,
    )
    evaluator = ArithmeticFormulaEvaluator(worksheet)
    conflicts: list[PaymentSyncConflict] = []
    for item in items:
        if item.item_id in item_errors:
            conflicts.append(
                PaymentSyncConflict(
                    conflict_id=_stable_id("invalid", item.item_id),
                    conflict_type=ConflictType.PAYMENT_SOURCE_INVALID,
                    message=item_errors[item.item_id],
                    item_id=item.item_id,
                    source_row=item.source_row,
                    sqt=item.sqt,
                    container=item.container,
                    allowed_actions=(
                        ResolutionAction.SKIP,
                        ResolutionAction.CANCEL_ALL,
                    ),
                    default_action=ResolutionAction.SKIP,
                )
            )
            continue
        matches = exact.get((item.sqt, item.container), ())
        if len(matches) == 1:
            row = matches[0]
            item.target_row = row
            differences: dict[str, tuple[Any, Any]] = {}
            for field in PAYMENT_FIELDS:
                cell = worksheet.cell(row, columns[field])
                try:
                    current = evaluator.value(cell)
                except PaymentSyncError:
                    current = cell.value
                incoming = item.values[field]
                if not _numeric_equal(current, incoming):
                    differences[field] = (current, incoming)
            for column, formula in _summary_formulas(row, columns).items():
                if worksheet.cell(row, column).value != formula:
                    differences[f"formula_{column}"] = (
                        worksheet.cell(row, column).value,
                        formula,
                    )
            item.differences = differences
            item.status = "UPDATE" if differences else "UNCHANGED"
            continue
        candidates = sorted(
            set(by_sqt.get(item.sqt, ()))
            | set(by_container.get(item.container, ()))
        )
        if candidates or len(matches) > 1:
            item.status = "CONFLICT"
            row_candidates = [
                RowCandidate(
                    row=row,
                    sqt=_parse_sqt(worksheet.cell(row, columns["sqt"]).value),
                    container=_container_key(
                        worksheet.cell(row, columns["container"]).value
                    ),
                )
                for row in (matches if len(matches) > 1 else candidates)
            ]
            conflicts.append(
                PaymentSyncConflict(
                    conflict_id=_stable_id("partial", item.item_id, candidates),
                    conflict_type=ConflictType.PARTIAL_KEY_MATCH,
                    message=(
                        "Chỉ khớp SQT hoặc container, hoặc có nhiều dòng đích. "
                        "Hãy chọn dòng Thanh toán chính xác."
                    ),
                    item_id=item.item_id,
                    source_row=item.source_row,
                    sqt=item.sqt,
                    container=item.container,
                    row_candidates=row_candidates,
                    details={"sheet_name": worksheet.title},
                )
            )
        else:
            item.status = "NEW"
    return conflicts


def _resolution_action(value: Any, default: ResolutionAction) -> ResolutionAction:
    if value is None:
        return default
    if isinstance(value, Mapping):
        value = value.get("action", default)
    else:
        value = getattr(value, "action", value)
    return ResolutionAction(getattr(value, "value", value))


def _resolution_value(value: Any, name: str) -> Any:
    if isinstance(value, Mapping):
        return value.get(name)
    return getattr(value, name, None)


def _actual_last_value_column(worksheet: Any) -> int:
    columns = [
        column
        for (_row, column), cell in getattr(worksheet, "_cells", {}).items()
        if cell.value not in (None, "")
    ]
    return max(columns, default=worksheet.max_column)


def _actual_last_value_row(worksheet: Any) -> int:
    rows = [
        row
        for (row, _column), cell in getattr(worksheet, "_cells", {}).items()
        if cell.value not in (None, "")
    ]
    return max(rows, default=worksheet.max_row)


def _copy_template_cell(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy.copy(source._style)
    target.number_format = source.number_format
    target.value = source.value
    if source.hyperlink is not None:
        target._hyperlink = copy.copy(source.hyperlink)
    if source.comment is not None:
        target.comment = copy.copy(source.comment)


def _clear_template_values(
    worksheet: Any,
    *,
    first_row: int,
    last_row: int,
    last_column: int,
    preserve_formulas: bool = False,
) -> None:
    if last_row < first_row:
        return
    for row in range(first_row, last_row + 1):
        for column in range(1, last_column + 1):
            cell = worksheet.cell(row, column)
            if isinstance(cell, MergedCell):
                continue
            if (
                preserve_formulas
                and isinstance(cell.value, str)
                and cell.value.startswith("=")
            ):
                continue
            cell.value = None
            cell.comment = None
            if cell.hyperlink is not None:
                cell._hyperlink = None


def _create_payment_month_sheet(
    workbook: Any,
    *,
    month_service: MonthSheetService,
    month: int,
    year: int,
    target_name: str,
    template_name: str | None = None,
) -> Any:
    """Create a clean payment month from the nearest previous valid sheet."""

    if target_name in workbook.sheetnames:
        raise PaymentSyncError(f"Sheet Thanh toán {target_name} đã tồn tại.")
    template_name = template_name or month_service.nearest_previous_template(
        workbook.sheetnames, month, year
    )
    if template_name is None or template_name not in workbook.sheetnames:
        raise PaymentSyncError(
            f"Không có sheet Thanh toán tháng trước làm mẫu cho {target_name}."
        )
    parsed_template = month_service.parse_target_sheet(template_name)
    if parsed_template is None or (
        parsed_template[1], parsed_template[0]
    ) >= (year, month):
        raise PaymentSyncError(
            f"Sheet {template_name} không phải tháng trước của {target_name}."
        )

    template = workbook[template_name]
    header_row, columns, summary_start = _resolve_target_columns(template)
    total_row = _target_total_row(
        template,
        header_row=header_row,
        columns=columns,
    )
    if total_row <= header_row + 1:
        raise PaymentSyncError(
            f"Sheet mẫu {template_name} không có vùng chi tiết hợp lệ."
        )

    last_value_column = _actual_last_value_column(template)
    last_column = max(
        last_value_column,
        summary_start + len(SUMMARY_HEADERS) - 1,
        max(columns.values()),
    )
    last_row = max(
        _actual_last_value_row(template),
        total_row,
        max(
            (merged.max_row for merged in template.merged_cells.ranges),
            default=total_row,
        ),
    )
    template_index = workbook.sheetnames.index(template_name)
    worksheet = workbook.create_sheet(target_name, index=template_index)

    for attribute in (
        "freeze_panes",
        "sheet_format",
        "sheet_properties",
        "page_margins",
        "page_setup",
        "print_options",
        "data_validations",
        "conditional_formatting",
    ):
        try:
            setattr(worksheet, attribute, copy.copy(getattr(template, attribute)))
        except (AttributeError, TypeError):
            pass
    for attribute in ("print_title_rows", "print_title_cols"):
        try:
            setattr(worksheet, attribute, getattr(template, attribute))
        except (AttributeError, TypeError, ValueError):
            pass

    for key, dimension in template.column_dimensions.items():
        try:
            column = column_index_from_string(str(key))
        except ValueError:
            continue
        if column <= last_column:
            worksheet.column_dimensions[key] = copy.copy(dimension)
    for row, dimension in template.row_dimensions.items():
        if row <= last_row:
            worksheet.row_dimensions[row] = copy.copy(dimension)

    for row in range(1, last_row + 1):
        for column in range(1, last_column + 1):
            source = template.cell(row, column)
            if isinstance(source, MergedCell):
                continue
            _copy_template_cell(source, worksheet.cell(row, column))
    for merged in template.merged_cells.ranges:
        if merged.max_row <= last_row and merged.max_col <= last_column:
            worksheet.merge_cells(str(merged))

    _template_month, template_year = parsed_template
    for row in range(1, header_row):
        for column in range(1, last_column + 1):
            cell = worksheet.cell(row, column)
            if not isinstance(cell.value, str):
                continue
            value = _PAYMENT_TITLE_MONTH_RE.sub(
                lambda match: f"{match.group(1)}{month:02d}/{year % 100:02d}",
                cell.value,
            )
            if template_year != year:
                value = re.sub(
                    rf"(?<!\d){template_year}(?!\d)",
                    str(year),
                    value,
                )
            cell.value = value

    _clear_template_values(
        worksheet,
        first_row=header_row + 1,
        last_row=total_row - 1,
        last_column=last_column,
    )
    manual_start = next(
        (
            row
            for row in range(total_row + 1, last_row + 1)
            if any(
                normalize_header(worksheet.cell(row, column).value)
                == _PAYMENT_MANUAL_SECTION
                for column in range(1, last_column + 1)
            )
        ),
        None,
    )
    if manual_start is not None:
        _clear_template_values(
            worksheet,
            first_row=manual_start,
            last_row=last_row,
            last_column=last_column,
        )

    if template.auto_filter.ref:
        worksheet.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(last_column)}{total_row - 1}"
        )
    try:
        if template.print_area:
            worksheet.print_area = (
                f"A1:{get_column_letter(last_column)}{last_row}"
            )
    except (AttributeError, TypeError, ValueError):
        pass

    _resolve_target_columns(worksheet)
    created_total = _target_total_row(
        worksheet,
        header_row=header_row,
        columns=columns,
    )
    if created_total != total_row:
        raise PaymentSyncError(
            f"Không bảo toàn được dòng tổng của sheet mẫu {template_name}."
        )
    return worksheet


def _ensure_date_column(
    worksheet: Any,
    *,
    header_row: int,
    summary_start: int,
) -> int:
    header_key = normalize_header(DATE_HEADER)
    matches = [
        column
        for column, value in _header_values(worksheet, header_row).items()
        if value == header_key
    ]
    if len(matches) > 1:
        raise PaymentSyncError("Sheet Thanh toán có nhiều cột Date cập nhật.")
    if matches:
        return matches[0]
    column = max(_actual_last_value_column(worksheet), summary_start + 9) + 1
    source = worksheet.cell(header_row, max(1, column - 1))
    cell = worksheet.cell(header_row, column)
    _copy_cell_style(source, cell)
    cell.value = DATE_HEADER
    return column


def _map_formula_rows_for_insert(
    worksheet: Any,
    *,
    insert_row: int,
    amount: int,
) -> None:
    for cell in tuple(getattr(worksheet, "_cells", {}).values()):
        value = getattr(cell, "value", None)
        if not isinstance(value, str) or not value.startswith("="):
            continue
        placeholders: dict[str, str] = {}

        def replace_range(match: re.Match[str]) -> str:
            if match.start() > 0 and value[match.start() - 1] == "!":
                return match.group(0)
            left = match.group("left")
            right = match.group("right")
            left_match = _CELL_REF_RE.fullmatch(left)
            right_match = _CELL_REF_RE.fullmatch(right)
            if left_match is None or right_match is None:
                return match.group(0)
            left_row = int(left_match.group("row"))
            right_row = int(right_match.group("row"))
            if left_row < insert_row and right_row == insert_row - 1:
                right = re.sub(
                    r"[1-9][0-9]*$",
                    str(right_row + amount),
                    right,
                )
            token = f"__RANGE_{len(placeholders)}__"
            placeholders[token] = f"{left}:{right}"
            return token

        translated = _RANGE_RE.sub(replace_range, value)

        def replace_cell(match: re.Match[str]) -> str:
            if match.start() > 0 and translated[match.start() - 1] == "!":
                return match.group(0)
            row = int(match.group("row"))
            mapped = row + amount if row >= insert_row else row
            return (
                f"{match.group('col_abs')}{match.group('col')}"
                f"{match.group('row_abs')}{mapped}"
            )

        translated = _CELL_REF_RE.sub(replace_cell, translated)
        for token, replacement in placeholders.items():
            translated = translated.replace(token, replacement)
        cell.value = translated


def _insert_target_rows(
    worksheet: Any,
    *,
    total_row: int,
    amount: int,
) -> None:
    if amount <= 0:
        return
    template_row = total_row - 1
    _map_formula_rows_for_insert(
        worksheet, insert_row=total_row, amount=amount
    )
    worksheet.insert_rows(total_row, amount)
    for offset in range(amount):
        target_row = total_row + offset
        worksheet.row_dimensions[target_row].height = (
            worksheet.row_dimensions[template_row].height
        )
        for column in range(1, min(worksheet.max_column, 100) + 1):
            source_cell = worksheet.cell(template_row, column)
            target_cell = worksheet.cell(target_row, column)
            _copy_cell_style(
                source_cell,
                target_cell,
            )
            if isinstance(source_cell.value, str) and source_cell.value.startswith("="):
                delta = target_row - template_row

                def translate_reference(match: re.Match[str]) -> str:
                    if match.start() > 0 and source_cell.value[match.start() - 1] == "!":
                        return match.group(0)
                    mapped_row = int(match.group("row"))
                    if not match.group("row_abs"):
                        mapped_row += delta
                    return (
                        f"{match.group('col_abs')}{match.group('col')}"
                        f"{match.group('row_abs')}{mapped_row}"
                    )

                target_cell.value = _CELL_REF_RE.sub(
                    translate_reference, source_cell.value
                )


def _write_target_item(
    worksheet: Any,
    *,
    row: int,
    item: PaymentSyncItem,
    columns: Mapping[str, int],
    date_column: int,
    timestamp: datetime,
    write_identity: bool,
) -> None:
    if write_identity:
        worksheet.cell(row, columns["sqt"]).value = item.sqt
        worksheet.cell(row, columns["container"]).value = item.container
    for field in PAYMENT_FIELDS:
        cell = worksheet.cell(row, columns[field])
        cell.value = item.values[field]
        cell.number_format = '#,##0'
    for column, formula in _summary_formulas(row, columns).items():
        cell = worksheet.cell(row, column)
        cell.value = formula
        if column != columns["summary_sqt"]:
            cell.number_format = '#,##0'
    date_cell = worksheet.cell(row, date_column)
    date_cell.value = timestamp
    date_cell.number_format = DATE_NUMBER_FORMAT


def _effective_max_column(worksheet: Any, *, cap: int = 100) -> int:
    """Return the last meaningful column without trusting phantom styles."""

    candidates = [
        column
        for (_row, column), cell in getattr(worksheet, "_cells", {}).items()
        if getattr(cell, "value", None) not in (None, "")
    ]
    merged_ranges = getattr(getattr(worksheet, "merged_cells", None), "ranges", ())
    candidates.extend(merged.max_col for merged in merged_ranges)
    return min(max(candidates, default=1), cap)


def _effective_max_row(worksheet: Any) -> int:
    candidates = [
        row
        for (row, _column), cell in getattr(worksheet, "_cells", {}).items()
        if getattr(cell, "value", None) not in (None, "")
    ]
    merged_ranges = getattr(getattr(worksheet, "merged_cells", None), "ranges", ())
    candidates.extend(merged.max_row for merged in merged_ranges)
    return max(candidates, default=1)


@dataclass(frozen=True, slots=True)
class ResolvedPaymentProfile:
    target_type: str
    header_row: int
    columns: dict[str, int]
    data_start_row: int
    total_row: int
    effective_max_column: int


class PaymentSheetProfile:
    target_type = ""
    managed_fields: tuple[str, ...] = ()
    aliases: Mapping[str, Sequence[str]] = {}

    def resolve(self, worksheet: Any) -> ResolvedPaymentProfile:
        scan_columns = _effective_max_column(worksheet)
        scan_rows = min(30, _effective_max_row(worksheet))
        required = ("sqt", "container", *self.managed_fields, "date")
        best_row = 1
        best_score = -1
        best_matches: dict[str, list[int]] = {}
        for row in range(1, scan_rows + 1):
            headers = {
                column: normalize_header(worksheet.cell(row, column).value)
                for column in range(1, scan_columns + 1)
            }
            qt_matches = _matching_columns(headers, self.aliases["sqt"])
            # NAM has a summary block to the right that intentionally repeats
            # QT and several fee headers. The second QT is the stable boundary
            # between detail columns and that summary block.
            detail_boundary = qt_matches[1] if len(qt_matches) > 1 else None
            matches = {}
            for field in required:
                matches[field] = _matching_columns(
                    headers,
                    self.aliases[field],
                    before=None if field == "date" else detail_boundary,
                )
            score = sum(bool(value) for value in matches.values())
            if score > best_score:
                best_row, best_score, best_matches = row, score, matches
            if all(len(value) == 1 for value in matches.values()):
                best_row, best_matches = row, matches
                break
        invalid = {
            field: len(best_matches.get(field, ()))
            for field in required
            if len(best_matches.get(field, ())) != 1
        }
        if invalid:
            detail = ", ".join(f"{field}={count}" for field, count in invalid.items())
            raise PaymentSyncError(
                f"Cấu trúc sheet {worksheet.title} ({self.target_type}) không hợp lệ; "
                f"cột bắt buộc thiếu hoặc trùng: {detail}."
            )
        columns = {field: values[0] for field, values in best_matches.items()}
        total_row = self.find_total_row(
            worksheet,
            header_row=best_row,
            columns=columns,
        )
        return ResolvedPaymentProfile(
            target_type=self.target_type,
            header_row=best_row,
            columns=columns,
            data_start_row=best_row + 1,
            total_row=total_row,
            effective_max_column=max(scan_columns, max(columns.values())),
        )

    def find_total_row(
        self,
        worksheet: Any,
        *,
        header_row: int,
        columns: Mapping[str, int],
    ) -> int:
        for row in range(header_row + 1, _effective_max_row(worksheet) + 1):
            if _parse_sqt(worksheet.cell(row, columns["sqt"]).value) is not None:
                continue
            if _container_key(worksheet.cell(row, columns["container"]).value):
                continue
            if any(
                isinstance(worksheet.cell(row, columns[field]).value, str)
                and re.match(
                    r"^=(?:SUM|SUBTOTAL)\(",
                    worksheet.cell(row, columns[field]).value,
                    re.IGNORECASE,
                )
                for field in self.managed_fields
            ):
                return row
        raise PaymentSyncError(
            f"Không tìm thấy dòng tổng của sheet {worksheet.title} ({self.target_type})."
        )

    def is_blank_row(
        self,
        worksheet: Any,
        row: int,
        resolved: ResolvedPaymentProfile,
    ) -> bool:
        fields = ("sqt", "container", *self.managed_fields)
        return all(
            worksheet.cell(row, resolved.columns[field]).value in (None, "", 0, 0.0)
            for field in fields
        )


class HPSheetProfile(PaymentSheetProfile):
    target_type = "HP"
    managed_fields = HP_FIELDS
    aliases = {
        "sqt": ("QT",),
        "container": ("SỐ CONT", "Số Container"),
        "loaded_drop": ("HẠ HÀNG",),
        "date": (DATE_HEADER,),
    }


class NAMSheetProfile(PaymentSheetProfile):
    target_type = "NAM"
    managed_fields = NAM_FIELDS
    aliases = {
        "sqt": ("QT",),
        "container": ("SỐ CONT", "Số Container"),
        "empty_lift": ("NÂNG VỎ",),
        "loaded_lift": ("NÂNG HÀNG",),
        "empty_drop": ("HẠ VỎ",),
        "vs_do": ("VS + D/O", "VS D/O"),
        "command_fee": ("LÀM LỆNH",),
        "storage": ("Lưu Cont", "Lưu container"),
        "repair": ("Sửa chữa Cont", "SỬA CHỮA"),
        "overweight": ("QUÁ TẢI",),
        "date": (DATE_HEADER,),
    }


PAYMENT_PROFILES: dict[str, PaymentSheetProfile] = {
    "HP": HPSheetProfile(),
    "NAM": NAMSheetProfile(),
}


def _has_money(value: Any) -> bool:
    if value in (None, "", 0, 0.0):
        return False
    try:
        return Decimal(str(value)) != 0
    except (InvalidOperation, ValueError):
        return True


def _source_sync_columns(worksheet: Any) -> dict[str, int]:
    summary_start = find_summary_start(worksheet)
    headers = _header_values(worksheet)
    return {
        field: _required_column(
            headers,
            BK_HEADER_ALIASES[field],
            field,
            before=summary_start,
        )
        for field in SYNC_SOURCE_FIELDS
    }


def _source_target_items(
    worksheet: Any,
    *,
    normalization_issues: Sequence[NormalizationIssue],
) -> tuple[dict[str, list[PaymentSyncItem]], dict[str, str]]:
    columns = _source_sync_columns(worksheet)
    evaluator = ArithmeticFormulaEvaluator(worksheet)
    grouped: dict[tuple[int, str], list[int]] = defaultdict(list)
    for row in _data_rows(
        worksheet,
        sqt_column=columns["sqt"],
        container_column=columns["container"],
    ):
        sqt = _parse_sqt(worksheet.cell(row, columns["sqt"]).value)
        container = _container_key(worksheet.cell(row, columns["container"]).value)
        if sqt is not None and container:
            grouped[(sqt, container)].append(row)
    issues_by_row = {
        issue.row: issue.message
        for issue in normalization_issues
        if issue.sheet_name == worksheet.title
    }
    result: dict[str, list[PaymentSyncItem]] = {"HP": [], "NAM": []}
    errors: dict[str, str] = {}
    for (sqt, container), rows in grouped.items():
        values: dict[str, Any] = {}
        row_errors = [issues_by_row[row] for row in rows if row in issues_by_row]
        for field in (*HP_FIELDS, *NAM_FIELDS):
            nonzero: list[Any] = []
            saw_zero = False
            for row in rows:
                cell = worksheet.cell(row, columns[field])
                if cell.value in (None, ""):
                    continue
                try:
                    value = evaluator.value(cell)
                except PaymentSyncError as exc:
                    row_errors.append(f"{field}: {exc}")
                    continue
                if _has_money(value):
                    nonzero.append(value)
                else:
                    saw_zero = True
            if len(nonzero) > 1:
                row_errors.append(
                    f"{field}: có nhiều dòng cùng SQT/container chứa số tiền."
                )
            values[field] = nonzero[0] if nonzero else (0 if saw_zero else None)
        for target_type, fields in (("HP", HP_FIELDS), ("NAM", NAM_FIELDS)):
            if not any(_has_money(values[field]) for field in fields):
                continue
            item_id = _stable_id(worksheet.title, target_type, sqt, container, rows)
            item = PaymentSyncItem(
                item_id=item_id,
                source_row=rows[0],
                source_rows=tuple(rows),
                sqt=sqt,
                container=container,
                values={field: values[field] for field in fields},
                target_type=target_type,
            )
            if row_errors:
                item.status = "INVALID"
                errors[item_id] = " • ".join(dict.fromkeys(row_errors))
            result[target_type].append(item)
    return result, errors


def _profile_target_index(
    worksheet: Any,
    resolved: ResolvedPaymentProfile,
) -> tuple[dict[tuple[int, str], list[int]], dict[int, list[int]], dict[str, list[int]]]:
    exact: dict[tuple[int, str], list[int]] = defaultdict(list)
    by_sqt: dict[int, list[int]] = defaultdict(list)
    by_container: dict[str, list[int]] = defaultdict(list)
    for row in range(resolved.data_start_row, resolved.total_row):
        sqt = _parse_sqt(worksheet.cell(row, resolved.columns["sqt"]).value)
        container = _container_key(
            worksheet.cell(row, resolved.columns["container"]).value
        )
        if sqt is not None:
            by_sqt[sqt].append(row)
        if container:
            by_container[container].append(row)
        if sqt is not None and container:
            exact[(sqt, container)].append(row)
    return exact, by_sqt, by_container


def _analyze_profile_target(
    worksheet: Any,
    profile: PaymentSheetProfile,
    items: list[PaymentSyncItem],
    item_errors: Mapping[str, str],
) -> list[PaymentSyncConflict]:
    resolved = profile.resolve(worksheet)
    exact, by_sqt, by_container = _profile_target_index(worksheet, resolved)
    evaluator = ArithmeticFormulaEvaluator(worksheet)
    conflicts: list[PaymentSyncConflict] = []
    for item in items:
        if item.item_id in item_errors:
            conflicts.append(
                PaymentSyncConflict(
                    conflict_id=_stable_id("invalid", item.item_id),
                    conflict_type=ConflictType.PAYMENT_SOURCE_INVALID,
                    message=item_errors[item.item_id],
                    item_id=item.item_id,
                    source_row=item.source_row,
                    sqt=item.sqt,
                    container=item.container,
                    allowed_actions=(ResolutionAction.SKIP, ResolutionAction.CANCEL_ALL),
                    default_action=ResolutionAction.SKIP,
                    details={"sheet_name": worksheet.title, "target_type": profile.target_type},
                )
            )
            continue
        matches = exact.get((item.sqt, item.container), ())
        if len(matches) == 1:
            row = matches[0]
            item.target_row = row
            differences: dict[str, tuple[Any, Any]] = {}
            clear_fields: list[str] = []
            for field in profile.managed_fields:
                cell = worksheet.cell(row, resolved.columns[field])
                try:
                    current = evaluator.value(cell)
                except PaymentSyncError:
                    current = cell.value
                incoming = item.values[field]
                if not _has_money(incoming) and _has_money(current):
                    clear_fields.append(field)
                elif _has_money(incoming) and not _numeric_equal(current, incoming):
                    differences[field] = (current, incoming)
            item.differences = differences
            if clear_fields:
                item.status = "CONFLICT"
                conflicts.append(
                    PaymentSyncConflict(
                        conflict_id=_stable_id("clear", item.item_id, clear_fields),
                        conflict_type=ConflictType.PAYMENT_CLEAR_VALUE,
                        message=(
                            "BK đang trống nhưng Thanh toán đang có tiền. "
                            "Mặc định giữ giá trị Thanh toán."
                        ),
                        item_id=item.item_id,
                        source_row=item.source_row,
                        sqt=item.sqt,
                        container=item.container,
                        allowed_actions=(
                            ResolutionAction.KEEP_EXISTING,
                            ResolutionAction.OVERWRITE,
                            ResolutionAction.SKIP,
                            ResolutionAction.CANCEL_ALL,
                        ),
                        default_action=ResolutionAction.KEEP_EXISTING,
                        details={
                            "sheet_name": worksheet.title,
                            "target_type": profile.target_type,
                            "clear_fields": tuple(clear_fields),
                        },
                    )
                )
            else:
                item.status = "UPDATE" if differences else "UNCHANGED"
            continue
        candidates = sorted(
            set(by_sqt.get(item.sqt, ())) | set(by_container.get(item.container, ()))
        )
        if candidates or len(matches) > 1:
            item.status = "CONFLICT"
            candidate_rows = list(matches) if len(matches) > 1 else candidates
            conflicts.append(
                PaymentSyncConflict(
                    conflict_id=_stable_id("partial", item.item_id, candidate_rows),
                    conflict_type=ConflictType.PARTIAL_KEY_MATCH,
                    message=(
                        "Chỉ khớp SQT hoặc container, hoặc có nhiều dòng cùng khóa. "
                        "Hãy chọn dòng đích chính xác."
                    ),
                    item_id=item.item_id,
                    source_row=item.source_row,
                    sqt=item.sqt,
                    container=item.container,
                    row_candidates=[
                        RowCandidate(
                            row=row,
                            sqt=_parse_sqt(worksheet.cell(row, resolved.columns["sqt"]).value),
                            container=_container_key(
                                worksheet.cell(row, resolved.columns["container"]).value
                            ),
                        )
                        for row in candidate_rows
                    ],
                    details={"sheet_name": worksheet.title, "target_type": profile.target_type},
                )
            )
        else:
            item.status = "NEW"
    return conflicts


def _copy_profile_sheet(
    workbook: Any,
    *,
    month_service: MonthSheetService,
    profile: PaymentSheetProfile,
    month: int,
    year: int,
    target_name: str,
    template_name: str,
) -> Any:
    if target_name in workbook.sheetnames:
        raise PaymentSyncError(f"Sheet Thanh toán {target_name} đã tồn tại.")
    parsed_template = month_service.parse_payment_sheet(template_name)
    if (
        parsed_template is None
        or parsed_template.sheet_type != profile.target_type
        or (parsed_template.year, parsed_template.month) >= (year, month)
    ):
        raise PaymentSyncError(
            f"Sheet {template_name} không phải sheet {profile.target_type} mẫu hợp lệ."
        )
    template = workbook[template_name]
    resolved = profile.resolve(template)
    last_column = max(
        resolved.effective_max_column,
        max((merged.max_col for merged in template.merged_cells.ranges), default=1),
    )
    last_row = max(
        _effective_max_row(template),
        max((merged.max_row for merged in template.merged_cells.ranges), default=1),
    )
    worksheet = workbook.create_sheet(
        target_name,
        index=workbook.sheetnames.index(template_name),
    )
    for attribute in (
        "freeze_panes",
        "sheet_format",
        "sheet_properties",
        "page_margins",
        "page_setup",
        "print_options",
        "data_validations",
        "conditional_formatting",
    ):
        try:
            setattr(worksheet, attribute, copy.copy(getattr(template, attribute)))
        except (AttributeError, TypeError):
            pass
    for attribute in ("print_title_rows", "print_title_cols"):
        try:
            setattr(worksheet, attribute, getattr(template, attribute))
        except (AttributeError, TypeError, ValueError):
            pass
    for key, dimension in template.column_dimensions.items():
        try:
            column = column_index_from_string(str(key))
        except ValueError:
            continue
        if column <= last_column:
            worksheet.column_dimensions[key] = copy.copy(dimension)
    for row, dimension in template.row_dimensions.items():
        if row <= last_row:
            worksheet.row_dimensions[row] = copy.copy(dimension)
    for row in range(1, last_row + 1):
        for column in range(1, last_column + 1):
            source = template.cell(row, column)
            if not isinstance(source, MergedCell):
                _copy_template_cell(source, worksheet.cell(row, column))
    for merged in template.merged_cells.ranges:
        if merged.max_row <= last_row and merged.max_col <= last_column:
            worksheet.merge_cells(str(merged))
    for row in range(1, resolved.header_row):
        for column in range(1, last_column + 1):
            cell = worksheet.cell(row, column)
            if isinstance(cell.value, str):
                cell.value = _PAYMENT_TITLE_MONTH_RE.sub(
                    lambda match: f"{match.group(1)}{month:02d}/{year % 100:02d}",
                    cell.value,
                )
    _clear_template_values(
        worksheet,
        first_row=resolved.data_start_row,
        last_row=resolved.total_row - 1,
        last_column=last_column,
        preserve_formulas=True,
    )
    if template.auto_filter.ref:
        worksheet.auto_filter.ref = (
            f"A{resolved.header_row}:{get_column_letter(last_column)}{resolved.total_row - 1}"
        )
    profile.resolve(worksheet)
    return worksheet


def _write_profile_item(
    worksheet: Any,
    *,
    row: int,
    item: PaymentSyncItem,
    profile: PaymentSheetProfile,
    resolved: ResolvedPaymentProfile,
    timestamp: datetime,
    write_identity: bool,
    clear_fields: set[str] | None = None,
) -> bool:
    clear_fields = clear_fields or set()
    changed = False
    if write_identity:
        for field, incoming in (("sqt", item.sqt), ("container", item.container)):
            cell = worksheet.cell(row, resolved.columns[field])
            normalized_current = (
                _parse_sqt(cell.value) if field == "sqt" else _container_key(cell.value)
            )
            if normalized_current != incoming:
                cell.value = incoming
                changed = True
    for field in profile.managed_fields:
        cell = worksheet.cell(row, resolved.columns[field])
        incoming = item.values[field]
        if not _has_money(incoming):
            if field in clear_fields and cell.value not in (None, "", 0, 0.0):
                cell.value = None
                changed = True
            continue
        if not _numeric_equal(cell.value, incoming):
            cell.value = incoming
            cell.number_format = "#,##0"
            changed = True
    if changed:
        date_cell = worksheet.cell(row, resolved.columns["date"])
        date_cell.value = timestamp
        date_cell.number_format = DATE_NUMBER_FORMAT
    return changed


def _package_has_vba(path: str | Path) -> bool:
    return workbook_has_vba(path)


class PaymentSyncService:
    def __init__(
        self,
        settings: Any | None = None,
        *,
        bk_path: str | Path | None = None,
        payment_path: str | Path | None = None,
        temp_dir: str | Path | None = None,
        backup_dir: str | Path | None = None,
        gateway: WorkbookGateway | None = None,
        lock_service: ExcelLockService | None = None,
        stability_checker: FileStabilityChecker | None = None,
        month_service: MonthSheetService | None = None,
        run_repository: Any | None = None,
        clock: Callable[[], datetime] | None = None,
    ) -> None:
        paths = getattr(settings, "paths", None)
        system_dir = getattr(paths, "system_dir", Path("Output") / "_system")
        self.bk_path = Path(
            bk_path or _setting(settings, "bk_workbook_path", "")
        )
        self.payment_path = Path(
            payment_path or _setting(settings, "payment_workbook_path", "")
        )
        self.temp_dir = Path(
            temp_dir
            or getattr(paths, "excel_temp_dir", system_dir / "Excel" / "Temp")
        )
        self.backup_dir = Path(
            backup_dir
            or getattr(paths, "excel_backup_dir", system_dir / "Excel" / "Backup")
        )
        self.gateway = gateway or WorkbookGateway()
        self.lock_service = lock_service or ExcelLockService()
        self.stability_checker = stability_checker or FileStabilityChecker()
        self.months = month_service or MonthSheetService()
        self.run_repository = run_repository
        self.clock = clock or datetime.now
        self.backups = ExcelBackupService(
            self.backup_dir, working_dir=self.temp_dir
        )

    def source_sheet_candidates(self) -> list[SourceSheetCandidate]:
        path = ensure_supported_workbook(self.bk_path)
        if not path.is_file():
            raise PaymentSyncError(f"Không tìm thấy file BK: {path}")
        workbook = self.gateway.load(path, read_only=True)
        try:
            result: list[SourceSheetCandidate] = []
            for name in workbook.sheetnames:
                parsed = self.months.parse_target_sheet(name)
                if parsed is None:
                    continue
                month, _year = parsed
                result.append(
                    SourceSheetCandidate(month=month, source_sheet=name)
                )
            return result
        finally:
            workbook.close()

    def _analyze_legacy_single_target(
        self,
        *,
        source_sheet_name: str,
        progress_callback: ProgressCallback = None,
    ) -> PaymentSyncPlan:
        source = ensure_supported_workbook(self.bk_path)
        target = ensure_supported_workbook(self.payment_path)
        if not source.is_file():
            raise PaymentSyncError(f"Không tìm thấy file BK: {source}")
        if not target.is_file():
            raise PaymentSyncError(f"Không tìm thấy file Thanh toán: {target}")
        if source.resolve() == target.resolve():
            raise PaymentSyncError("File BK và file Thanh toán không được trùng nhau.")
        _progress(progress_callback, "Đang kiểm tra hai workbook…")
        self.stability_checker.wait(source)
        self.stability_checker.wait(target)
        self.lock_service.ensure_readable(source)
        self.lock_service.ensure_readable(target)
        source_fp = self.gateway.fingerprint(source)
        target_fp = self.gateway.fingerprint(target)
        run_id = self._create_run(source, target, source_fp, target_fp)
        source_book = target_book = None
        try:
            source_book = self.gateway.load(source, read_only=False)
            target_book = self.gateway.load(target, read_only=False)
            if source_sheet_name not in source_book.sheetnames:
                raise PaymentSyncError(
                    f"Không tìm thấy sheet BK {source_sheet_name}."
                )
            parsed = self.months.parse_target_sheet(source_sheet_name)
            if parsed is None:
                raise PaymentSyncError(
                    f"Sheet BK {source_sheet_name} không có dạng TMM YY."
                )
            target_matches = [
                name
                for name in target_book.sheetnames
                if self.months.parse_target_sheet(name) == parsed
            ]
            if len(target_matches) > 1:
                raise PaymentSyncError(
                    "Không xác định duy nhất sheet Thanh toán cùng tháng/năm "
                    f"với {source_sheet_name}."
                )
            target_sheet_created = not target_matches
            template_sheet = None
            if target_matches:
                target_sheet_name = target_matches[0]
            else:
                month, year = parsed
                target_sheet_name = self.months.target_name(month, year)
                template_sheet = self.months.nearest_previous_template(
                    target_book.sheetnames,
                    month,
                    year,
                )
                if template_sheet is None:
                    raise PaymentSyncError(
                        "Không có sheet Thanh toán tháng trước làm mẫu cho "
                        f"{target_sheet_name}."
                    )
                _progress(
                    progress_callback,
                    f"Đang dựng {target_sheet_name} từ mẫu {template_sheet} "
                    "trong bộ nhớ…",
                )
                _create_payment_month_sheet(
                    target_book,
                    month_service=self.months,
                    month=month,
                    year=year,
                    target_name=target_sheet_name,
                    template_name=template_sheet,
                )
            _progress(progress_callback, "Đang chuẩn hóa cấu trúc BK trong bộ nhớ…")
            normalization = normalize_bk_workbook(
                source_book, month_service=self.months
            )
            _progress(
                progress_callback,
                f"Đang đối chiếu {source_sheet_name} với {target_sheet_name}…",
            )
            items, item_errors = _source_items(
                source_book[source_sheet_name],
                normalization_issues=normalization.issues,
            )
            conflicts = _analyze_target(
                target_book[target_sheet_name], items, item_errors
            )
            plan = PaymentSyncPlan(
                source_path=source,
                target_path=target,
                source_fingerprint=source_fp,
                target_fingerprint=target_fp,
                source_sheet=source_sheet_name,
                target_sheet=target_sheet_name,
                items=items,
                conflicts=conflicts,
                normalization_required=normalization.changed,
                normalization_sheet_count=len(normalization.changed_sheets),
                target_sheet_created=target_sheet_created,
                template_sheet=template_sheet,
                run_id=run_id,
            )
            self._update_run(
                run_id,
                status=ExcelRunStatus.WAITING_USER
                if plan.requires_user_input
                else ExcelRunStatus.ANALYZING,
                sheet_name=target_sheet_name,
                total_items=len(items),
                changed_items=plan.update_count + plan.new_count,
                conflict_count=plan.conflict_count,
            )
            return plan
        except Exception as exc:
            self._finish_failed(run_id, exc)
            raise
        finally:
            if source_book is not None:
                source_book.close()
            if target_book is not None:
                target_book.close()

    def _apply_legacy_single_target(
        self,
        plan: PaymentSyncPlan,
        resolutions: Mapping[str, Any] | None = None,
        *,
        progress_callback: ProgressCallback = None,
    ) -> PaymentSyncResult:
        resolved = dict(resolutions or {})
        selected_raw = resolved.get("selected_new_rows")
        selected_new = (
            {str(value) for value in selected_raw}
            if selected_raw is not None
            else {item.item_id for item in plan.new_rows}
        )
        skipped_ids: set[str] = set()
        selected_targets: dict[str, int] = {}
        for conflict in plan.conflicts:
            value = resolved.get(conflict.conflict_id)
            action = _resolution_action(value, conflict.default_action)
            if action in {ResolutionAction.CANCEL, ResolutionAction.CANCEL_ALL}:
                raise PaymentSyncError("Người dùng đã hủy đồng bộ Thanh toán.")
            if action is ResolutionAction.SELECT_ROW:
                row = _resolution_value(value, "selected_row")
                if row is None:
                    raise PaymentSyncError("Xung đột chưa được chọn dòng đích.")
                selected_targets[conflict.item_id] = int(row)
            else:
                skipped_ids.add(conflict.item_id)

        self.gateway.assert_unchanged(
            plan.source_path, plan.source_fingerprint, label="File BK"
        )
        self.gateway.assert_unchanged(
            plan.target_path, plan.target_fingerprint, label="File Thanh toán"
        )
        self._update_run(plan.run_id, status=ExcelRunStatus.APPLYING)
        source_working = target_working = None
        source_backup = target_backup = None
        source_after = plan.source_fingerprint
        target_after = plan.target_fingerprint
        timestamp = self.clock().replace(tzinfo=None, microsecond=0)
        inserted = updated = 0
        sheet_created = False
        try:
            _progress(progress_callback, "Đang tạo bản làm việc an toàn…")
            source_working = self.backups.create_working_copy(
                plan.source_path, run_id=f"{plan.run_id}-bk"
            )
            target_working = self.backups.create_working_copy(
                plan.target_path, run_id=f"{plan.run_id}-payment"
            )
            source_book = self.gateway.load(source_working, read_only=False)
            target_book = self.gateway.load(target_working, read_only=False)
            try:
                normalization = normalize_bk_workbook(
                    source_book, month_service=self.months
                )
                if plan.target_sheet_created:
                    parsed = self.months.parse_target_sheet(plan.target_sheet)
                    if parsed is None:
                        raise PaymentSyncError(
                            f"Sheet Thanh toán {plan.target_sheet} không hợp lệ."
                        )
                    if any(
                        self.months.parse_target_sheet(name) == parsed
                        for name in target_book.sheetnames
                    ):
                        raise PaymentSyncError(
                            "Sheet Thanh toán dự kiến tạo mới đã tồn tại; "
                            "hãy phân tích lại."
                        )
                    _progress(
                        progress_callback,
                        f"Đang tạo {plan.target_sheet} từ mẫu "
                        f"{plan.template_sheet}…",
                    )
                    _create_payment_month_sheet(
                        target_book,
                        month_service=self.months,
                        month=parsed[0],
                        year=parsed[1],
                        target_name=plan.target_sheet,
                        template_name=plan.template_sheet,
                    )
                    sheet_created = True
                elif plan.target_sheet not in target_book.sheetnames:
                    raise PaymentSyncError(
                        f"Không tìm thấy sheet Thanh toán {plan.target_sheet}; "
                        "hãy phân tích lại."
                    )
                worksheet = target_book[plan.target_sheet]
                header_row, columns, summary_start = _resolve_target_columns(
                    worksheet
                )
                total_row = _target_total_row(
                    worksheet, header_row=header_row, columns=columns
                )
                date_column = _ensure_date_column(
                    worksheet,
                    header_row=header_row,
                    summary_start=summary_start,
                )
                items_by_id = {item.item_id: item for item in plan.items}
                write_existing = [
                    item
                    for item in plan.update_rows
                    if item.item_id not in skipped_ids
                ]
                selected_conflicts = [
                    items_by_id[item_id]
                    for item_id in selected_targets
                    if item_id not in skipped_ids
                ]
                new_items = [
                    item
                    for item in plan.new_rows
                    if item.item_id in selected_new
                    and item.item_id not in skipped_ids
                ]
                blank_rows = [
                    row
                    for row in range(header_row + 1, total_row)
                    if worksheet.cell(row, columns["sqt"]).value in (None, "")
                    and worksheet.cell(row, columns["container"]).value in (None, "")
                ]
                if len(blank_rows) < len(new_items):
                    needed = len(new_items) - len(blank_rows)
                    _insert_target_rows(
                        worksheet, total_row=total_row, amount=needed
                    )
                    blank_rows.extend(range(total_row, total_row + needed))
                    total_row += needed

                for item in write_existing:
                    if item.target_row is None:
                        continue
                    _write_target_item(
                        worksheet,
                        row=item.target_row,
                        item=item,
                        columns=columns,
                        date_column=date_column,
                        timestamp=timestamp,
                        write_identity=False,
                    )
                    updated += 1
                for item in selected_conflicts:
                    _write_target_item(
                        worksheet,
                        row=selected_targets[item.item_id],
                        item=item,
                        columns=columns,
                        date_column=date_column,
                        timestamp=timestamp,
                        write_identity=True,
                    )
                    updated += 1
                for item, row in zip(
                    new_items,
                    blank_rows[: len(new_items)],
                    strict=True,
                ):
                    _write_target_item(
                        worksheet,
                        row=row,
                        item=item,
                        columns=columns,
                        date_column=date_column,
                        timestamp=timestamp,
                        write_identity=True,
                    )
                    inserted += 1
                calculation = getattr(target_book, "calculation", None)
                if calculation is not None:
                    calculation.fullCalcOnLoad = True
                    calculation.forceFullCalc = True
                    calculation.calcMode = "auto"
                self.gateway.save(source_book, source_working)
                self.gateway.save(target_book, target_working)
            finally:
                source_book.close()
                target_book.close()

            self._verify_working_files(
                source_working,
                target_working,
                plan,
                selected_new=selected_new,
                skipped_ids=skipped_ids,
                selected_targets=selected_targets,
                timestamp=timestamp,
            )
            source_changed = normalization.changed
            target_changed = bool(inserted or updated or sheet_created)
            if source_changed or target_changed:
                with ExitStack() as stack:
                    for path in sorted(
                        (plan.source_path, plan.target_path),
                        key=lambda value: str(value).casefold(),
                    ):
                        self.lock_service.ensure_writable(path)
                self.gateway.assert_unchanged(
                    plan.source_path, plan.source_fingerprint, label="File BK"
                )
                self.gateway.assert_unchanged(
                    plan.target_path,
                    plan.target_fingerprint,
                    label="File Thanh toán",
                )
                if source_changed:
                    source_backup = self.backups.create_backup(
                        plan.source_path, run_id=f"{plan.run_id}-bk"
                    )
                if target_changed:
                    target_backup = self.backups.create_backup(
                        plan.target_path,
                        run_id=f"{plan.run_id}-payment",
                    )
                if source_changed:
                    source_after = self.gateway.atomic_replace(
                        source_working,
                        plan.source_path,
                        expected=plan.source_fingerprint,
                    )
                    source_working = None
                try:
                    if target_changed:
                        target_after = self.gateway.atomic_replace(
                            target_working,
                            plan.target_path,
                            expected=plan.target_fingerprint,
                        )
                        target_working = None
                except Exception:
                    if source_changed and source_backup is not None:
                        restore = self.backups.create_working_copy(
                            source_backup,
                            run_id=f"{plan.run_id}-rollback",
                        )
                        self.gateway.atomic_replace(
                            restore,
                            plan.source_path,
                            expected=source_after,
                        )
                    raise

            skipped = (
                len(plan.new_rows) - inserted
                + len(skipped_ids)
            )
            status = (
                ExcelRunStatus.SUCCEEDED
                if inserted or updated or normalization.changed or sheet_created
                else ExcelRunStatus.NO_CHANGES
            )
            result = PaymentSyncResult(
                status=status,
                source_path=plan.source_path,
                target_path=plan.target_path,
                sheet_name=plan.target_sheet,
                source_sheet_name=plan.source_sheet,
                sheet_created=sheet_created,
                template_sheet_name=plan.template_sheet,
                inserted_rows=inserted,
                updated_rows=updated,
                unchanged_rows=plan.unchanged_count,
                skipped_rows=skipped,
                conflict_count=plan.conflict_count,
                backup_path=target_backup,
                source_backup_path=source_backup,
                fingerprint_before=plan.target_fingerprint,
                fingerprint_after=target_after,
                source_fingerprint_after=source_after,
                run_id=plan.run_id,
                message=(
                    (
                        f"Đã tạo {plan.target_sheet} từ mẫu "
                        f"{plan.template_sheet}, cập nhật {updated} dòng và "
                        f"thêm {inserted} dòng."
                    )
                    if sheet_created
                    else (
                        f"Đã cập nhật {updated} dòng và thêm {inserted} dòng vào "
                        f"{plan.target_sheet}."
                    )
                    if inserted or updated
                    else (
                        "Đã chuẩn hóa cấu trúc BK; Thanh toán không có dòng cần ghi."
                        if normalization.changed
                        else "Dữ liệu BK và Thanh toán đã đồng bộ."
                    )
                ),
            )
            self._finish_result(result)
            return result
        except Exception as exc:
            self._finish_failed(plan.run_id, exc)
            raise
        finally:
            for path in (source_working, target_working):
                if path is not None:
                    Path(path).unlink(missing_ok=True)

    def _verify_working_files_legacy(
        self,
        source_path: Path,
        target_path: Path,
        plan: PaymentSyncPlan,
        *,
        selected_new: set[str],
        skipped_ids: set[str],
        selected_targets: Mapping[str, int],
        timestamp: datetime,
    ) -> None:
        source_book = self.gateway.load(source_path, read_only=False)
        target_book = self.gateway.load(target_path, read_only=False)
        try:
            for worksheet in source_book.worksheets:
                if self.months.parse_target_sheet(worksheet.title) is None:
                    continue
                if find_summary_start(worksheet) is None:
                    raise PaymentSyncError(
                        f"Khối tổng của {worksheet.title} không qua kiểm tra."
                    )
                headers = _header_values(worksheet)
                boundary = find_summary_start(worksheet)
                _required_column(
                    headers,
                    ("VS + D/O",),
                    "VS + D/O",
                    before=boundary,
                )
                _required_column(
                    headers,
                    ("LÀM LỆNH",),
                    "LÀM LỆNH",
                    before=boundary,
                )
            worksheet = target_book[plan.target_sheet]
            header_row, columns, summary_start = _resolve_target_columns(worksheet)
            date_column = _ensure_date_column(
                worksheet,
                header_row=header_row,
                summary_start=summary_start,
            )
            total_row = _target_total_row(
                worksheet, header_row=header_row, columns=columns
            )
            exact, _by_sqt, _by_cont = _target_index(
                worksheet,
                header_row=header_row,
                total_row=total_row,
                columns=columns,
            )
            expected = [
                item
                for item in plan.items
                if item.item_id not in skipped_ids
                and (
                    item.is_update
                    or item.item_id in selected_new
                    or item.item_id in selected_targets
                )
            ]
            for item in expected:
                matches = exact.get((item.sqt, item.container), ())
                if len(matches) != 1:
                    raise PaymentSyncError(
                        f"Dòng {item.sqt}/{item.container} không qua kiểm tra ghép khóa."
                    )
                row = matches[0]
                for field in PAYMENT_FIELDS:
                    actual = worksheet.cell(row, columns[field]).value
                    if not _numeric_equal(actual, item.values[field]):
                        raise PaymentSyncError(
                            f"Giá trị {field} của dòng {row} không qua kiểm tra."
                        )
                if worksheet.cell(row, date_column).value != timestamp:
                    raise PaymentSyncError(
                        f"Date cập nhật dòng {row} không qua kiểm tra."
                    )
        finally:
            source_book.close()
            target_book.close()

    # The dual-target implementation intentionally lives after the legacy
    # single-sheet methods. These definitions replace them on the class while
    # keeping the BK normalization helpers above available to Daily Sync.
    def analyze(
        self,
        *,
        source_sheet_name: str,
        progress_callback: ProgressCallback = None,
    ) -> PaymentSyncPlan:
        source = ensure_supported_workbook(self.bk_path)
        target = ensure_supported_workbook(self.payment_path)
        if not source.is_file():
            raise PaymentSyncError(f"Không tìm thấy file BK: {source}")
        if not target.is_file():
            raise PaymentSyncError(f"Không tìm thấy file Thanh toán: {target}")
        if source.resolve() == target.resolve():
            raise PaymentSyncError("File BK và file Thanh toán không được trùng nhau.")
        _progress(progress_callback, "Đang kiểm tra hai workbook…")
        self.stability_checker.wait(source)
        self.stability_checker.wait(target)
        self.lock_service.ensure_readable(source)
        self.lock_service.ensure_readable(target)
        source_fp = self.gateway.fingerprint(source)
        target_fp = self.gateway.fingerprint(target)
        run_id = self._create_run(source, target, source_fp, target_fp)
        source_book = target_book = None
        try:
            source_book = self.gateway.load(source, read_only=False)
            target_book = self.gateway.load(target, read_only=False)
            if source_sheet_name not in source_book.sheetnames:
                raise PaymentSyncError(f"Không tìm thấy sheet BK {source_sheet_name}.")
            parsed = self.months.parse_target_sheet(source_sheet_name)
            if parsed is None:
                raise PaymentSyncError(
                    f"Sheet BK {source_sheet_name} không có dạng TMM YY."
                )
            month, year = parsed
            _progress(progress_callback, "Đang chuẩn hóa cấu trúc BK trong bộ nhớ…")
            normalization = normalize_bk_workbook(
                source_book, month_service=self.months
            )
            items_by_target, item_errors = _source_target_items(
                source_book[source_sheet_name],
                normalization_issues=normalization.issues,
            )
            targets: dict[str, PaymentTargetPlan] = {}
            for target_type in ("HP", "NAM"):
                profile = PAYMENT_PROFILES[target_type]
                target_sheet_name = self.months.find_payment_sheet(
                    target_book.sheetnames, month, year, target_type
                )
                sheet_to_create = target_sheet_name is None
                template_sheet = None
                if sheet_to_create:
                    target_sheet_name = self.months.payment_target_name(
                        month, year, target_type
                    )
                    template_sheet = self.months.nearest_previous_payment_template(
                        target_book.sheetnames, month, year, target_type
                    )
                    if template_sheet is None:
                        raise PaymentSyncError(
                            f"Không tìm thấy sheet {target_type} mẫu"
                        )
                    _progress(
                        progress_callback,
                        f"Đang dựng {target_sheet_name} từ mẫu {template_sheet} trong bộ nhớ…",
                    )
                    _copy_profile_sheet(
                        target_book,
                        month_service=self.months,
                        profile=profile,
                        month=month,
                        year=year,
                        target_name=target_sheet_name,
                        template_name=template_sheet,
                    )
                _progress(
                    progress_callback,
                    f"Đang đối chiếu {source_sheet_name} với {target_sheet_name}…",
                )
                target_items = items_by_target[target_type]
                conflicts = _analyze_profile_target(
                    target_book[target_sheet_name],
                    profile,
                    target_items,
                    item_errors,
                )
                targets[target_type] = PaymentTargetPlan(
                    target_type=target_type,
                    sheet_name=target_sheet_name,
                    items=target_items,
                    conflicts=conflicts,
                    sheet_to_create=sheet_to_create,
                    template_sheet=template_sheet,
                )
            plan = PaymentSyncPlan(
                source_path=source,
                target_path=target,
                source_fingerprint=source_fp,
                target_fingerprint=target_fp,
                source_sheet=source_sheet_name,
                targets=targets,
                normalization_required=normalization.changed,
                normalization_sheet_count=len(normalization.changed_sheets),
                source_vba_present=_package_has_vba(source),
                target_vba_present=_package_has_vba(target),
                run_id=run_id,
            )
            self._update_run(
                run_id,
                status=(
                    ExcelRunStatus.WAITING_USER
                    if plan.requires_user_input
                    else ExcelRunStatus.ANALYZING
                ),
                sheet_name=plan.selected_sheet,
                total_items=len(plan.items),
                changed_items=plan.update_count + plan.new_count,
                conflict_count=plan.conflict_count,
            )
            return plan
        except Exception as exc:
            self._finish_failed(run_id, exc)
            raise
        finally:
            if source_book is not None:
                source_book.close()
            if target_book is not None:
                target_book.close()

    def apply(
        self,
        plan: PaymentSyncPlan,
        resolutions: Mapping[str, Any] | None = None,
        *,
        progress_callback: ProgressCallback = None,
    ) -> PaymentSyncResult:
        resolved_values = dict(resolutions or {})
        selected_raw = resolved_values.get("selected_new_rows")
        selected_new = (
            {str(value) for value in selected_raw}
            if selected_raw is not None
            else {item.item_id for item in plan.new_rows}
        )
        skipped_ids: set[str] = set()
        selected_targets: dict[str, int] = {}
        clear_fields: dict[str, set[str]] = defaultdict(set)
        for conflict in plan.conflicts:
            value = resolved_values.get(conflict.conflict_id)
            action = _resolution_action(value, conflict.default_action)
            if action in {ResolutionAction.CANCEL, ResolutionAction.CANCEL_ALL}:
                raise PaymentSyncError("Người dùng đã hủy toàn bộ đồng bộ Thanh toán.")
            if conflict.conflict_type is ConflictType.PAYMENT_CLEAR_VALUE:
                if action is ResolutionAction.OVERWRITE:
                    clear_fields[conflict.item_id].update(
                        str(field) for field in conflict.details.get("clear_fields", ())
                    )
                elif action is ResolutionAction.SKIP:
                    skipped_ids.add(conflict.item_id)
                # KEEP_EXISTING is deliberately a no-op for blank BK cells.
            elif action is ResolutionAction.SELECT_ROW:
                row = _resolution_value(value, "selected_row")
                if row is None:
                    raise PaymentSyncError("Xung đột chưa được chọn dòng đích.")
                selected_targets[conflict.item_id] = int(row)
            else:
                skipped_ids.add(conflict.item_id)

        self.gateway.assert_unchanged(
            plan.source_path, plan.source_fingerprint, label="File BK"
        )
        self.gateway.assert_unchanged(
            plan.target_path, plan.target_fingerprint, label="File Thanh toán"
        )
        self._update_run(plan.run_id, status=ExcelRunStatus.APPLYING)
        source_working = target_working = None
        source_backup = target_backup = None
        source_after = plan.source_fingerprint
        target_after = plan.target_fingerprint
        timestamp = self.clock().replace(tzinfo=None, microsecond=0)
        target_results: dict[str, PaymentTargetResult] = {}
        written: dict[str, dict[str, int]] = {"HP": {}, "NAM": {}}
        try:
            _progress(progress_callback, "Đang tạo bản làm việc an toàn…")
            source_working = self.backups.create_working_copy(
                plan.source_path, run_id=f"{plan.run_id}-bk"
            )
            target_working = self.backups.create_working_copy(
                plan.target_path, run_id=f"{plan.run_id}-payment"
            )
            source_book = self.gateway.load(source_working, read_only=False)
            target_book = self.gateway.load(target_working, read_only=False)
            try:
                normalization = normalize_bk_workbook(
                    source_book, month_service=self.months
                )
                parsed_source = self.months.parse_target_sheet(plan.source_sheet)
                if parsed_source is None:
                    raise PaymentSyncError("Tên sheet BK trong kế hoạch không hợp lệ.")
                month, year = parsed_source
                all_items = {item.item_id: item for item in plan.items}
                for target_type in ("HP", "NAM"):
                    target_plan = plan.targets[target_type]
                    profile = PAYMENT_PROFILES[target_type]
                    created = False
                    if target_plan.sheet_to_create:
                        if target_plan.sheet_name in target_book.sheetnames:
                            raise PaymentSyncError(
                                f"Sheet {target_plan.sheet_name} đã xuất hiện; hãy phân tích lại."
                            )
                        if target_plan.template_sheet is None:
                            raise PaymentSyncError(
                                f"Không tìm thấy sheet {target_type} mẫu"
                            )
                        _copy_profile_sheet(
                            target_book,
                            month_service=self.months,
                            profile=profile,
                            month=month,
                            year=year,
                            target_name=target_plan.sheet_name,
                            template_name=target_plan.template_sheet,
                        )
                        created = True
                    elif target_plan.sheet_name not in target_book.sheetnames:
                        raise PaymentSyncError(
                            f"Không tìm thấy sheet {target_plan.sheet_name}; hãy phân tích lại."
                        )
                    worksheet = target_book[target_plan.sheet_name]
                    profile_resolved = profile.resolve(worksheet)
                    new_items = [
                        item
                        for item in target_plan.new_rows
                        if item.item_id in selected_new and item.item_id not in skipped_ids
                    ]
                    blank_rows = [
                        row
                        for row in range(
                            profile_resolved.data_start_row,
                            profile_resolved.total_row,
                        )
                        if profile.is_blank_row(worksheet, row, profile_resolved)
                    ]
                    if len(blank_rows) < len(new_items):
                        needed = len(new_items) - len(blank_rows)
                        old_total = profile_resolved.total_row
                        _insert_target_rows(worksheet, total_row=old_total, amount=needed)
                        blank_rows.extend(range(old_total, old_total + needed))
                    inserted = updated = effective_unchanged = 0
                    for item in target_plan.update_rows:
                        if item.item_id in skipped_ids or item.target_row is None:
                            continue
                        if _write_profile_item(
                            worksheet,
                            row=item.target_row,
                            item=item,
                            profile=profile,
                            resolved=profile_resolved,
                            timestamp=timestamp,
                            write_identity=False,
                        ):
                            updated += 1
                            written[target_type][item.item_id] = item.target_row
                    clear_conflict_items = {
                        conflict.item_id
                        for conflict in target_plan.conflicts
                        if conflict.conflict_type is ConflictType.PAYMENT_CLEAR_VALUE
                    }
                    for item_id in clear_conflict_items:
                        item = all_items[item_id]
                        if item_id in skipped_ids or item.target_row is None:
                            continue
                        if _write_profile_item(
                            worksheet,
                            row=item.target_row,
                            item=item,
                            profile=profile,
                            resolved=profile_resolved,
                            timestamp=timestamp,
                            write_identity=False,
                            clear_fields=clear_fields.get(item_id, set()),
                        ):
                            updated += 1
                            written[target_type][item_id] = item.target_row
                        else:
                            effective_unchanged += 1
                    for item_id, row in selected_targets.items():
                        item = all_items.get(item_id)
                        if (
                            item is None
                            or item.target_type != target_type
                            or item_id in skipped_ids
                        ):
                            continue
                        if _write_profile_item(
                            worksheet,
                            row=row,
                            item=item,
                            profile=profile,
                            resolved=profile_resolved,
                            timestamp=timestamp,
                            write_identity=True,
                        ):
                            updated += 1
                            written[target_type][item_id] = row
                    for item, row in zip(
                        new_items, blank_rows[: len(new_items)], strict=True
                    ):
                        if _write_profile_item(
                            worksheet,
                            row=row,
                            item=item,
                            profile=profile,
                            resolved=profile_resolved,
                            timestamp=timestamp,
                            write_identity=True,
                        ):
                            inserted += 1
                            written[target_type][item.item_id] = row
                    skipped = len(
                        {
                            item.item_id
                            for item in target_plan.items
                            if item.item_id in skipped_ids
                            or (item.is_new and item.item_id not in selected_new)
                        }
                    )
                    target_results[target_type] = PaymentTargetResult(
                        target_type=target_type,
                        sheet_name=target_plan.sheet_name,
                        sheet_created=created,
                        template_sheet_name=target_plan.template_sheet,
                        inserted_rows=inserted,
                        updated_rows=updated,
                        unchanged_rows=target_plan.unchanged_count + effective_unchanged,
                        skipped_rows=skipped,
                        conflict_count=target_plan.conflict_count,
                    )
                calculation = getattr(target_book, "calculation", None)
                if calculation is not None:
                    calculation.fullCalcOnLoad = True
                    calculation.forceFullCalc = True
                    calculation.calcMode = "auto"
                self.gateway.save(source_book, source_working)
                self.gateway.save(target_book, target_working)
            finally:
                source_book.close()
                target_book.close()

            self._verify_dual_working_files(
                source_working,
                target_working,
                plan,
                written=written,
                clear_fields=clear_fields,
                timestamp=timestamp,
            )
            source_changed = normalization.changed
            target_changed = any(
                result.sheet_created or result.inserted_rows or result.updated_rows
                for result in target_results.values()
            )
            if source_changed or target_changed:
                with ExitStack() as stack:
                    for path in sorted(
                        (plan.source_path, plan.target_path),
                        key=lambda value: str(value).casefold(),
                    ):
                        self.lock_service.ensure_writable(path)
                    self.gateway.assert_unchanged(
                        plan.source_path, plan.source_fingerprint, label="File BK"
                    )
                    self.gateway.assert_unchanged(
                        plan.target_path,
                        plan.target_fingerprint,
                        label="File Thanh toán",
                    )
                    if source_changed:
                        source_backup = self.backups.create_backup(
                            plan.source_path, run_id=f"{plan.run_id}-bk"
                        )
                    if target_changed:
                        target_backup = self.backups.create_backup(
                            plan.target_path, run_id=f"{plan.run_id}-payment"
                        )
                    if source_changed:
                        source_after = self.gateway.atomic_replace(
                            source_working,
                            plan.source_path,
                            expected=plan.source_fingerprint,
                        )
                        source_working = None
                    try:
                        if target_changed:
                            target_after = self.gateway.atomic_replace(
                                target_working,
                                plan.target_path,
                                expected=plan.target_fingerprint,
                            )
                            target_working = None
                    except Exception:
                        if source_changed and source_backup is not None:
                            restore = self.backups.create_working_copy(
                                source_backup, run_id=f"{plan.run_id}-rollback"
                            )
                            self.gateway.atomic_replace(
                                restore, plan.source_path, expected=source_after
                            )
                        raise
            changed = normalization.changed or any(
                result.sheet_created or result.inserted_rows or result.updated_rows
                for result in target_results.values()
            )
            result = PaymentSyncResult(
                status=ExcelRunStatus.SUCCEEDED if changed else ExcelRunStatus.NO_CHANGES,
                source_path=plan.source_path,
                target_path=plan.target_path,
                target_results=target_results,
                source_sheet_name=plan.source_sheet,
                backup_path=target_backup,
                source_backup_path=source_backup,
                fingerprint_before=plan.target_fingerprint,
                fingerprint_after=target_after,
                source_fingerprint_after=source_after,
                vba_preserved=True,
                run_id=plan.run_id,
                message=(
                    "Đã đồng bộ nguyên tử hai sheet HP/NAM."
                    if changed
                    else "Dữ liệu HP/NAM đã đồng bộ."
                ),
            )
            self._finish_result(result)
            return result
        except Exception as exc:
            self._finish_failed(plan.run_id, exc)
            raise
        finally:
            for path in (source_working, target_working):
                if path is not None:
                    Path(path).unlink(missing_ok=True)

    def _verify_dual_working_files(
        self,
        source_path: Path,
        target_path: Path,
        plan: PaymentSyncPlan,
        *,
        written: Mapping[str, Mapping[str, int]],
        clear_fields: Mapping[str, set[str]],
        timestamp: datetime,
    ) -> None:
        if plan.source_vba_present and not _package_has_vba(source_path):
            raise PaymentSyncError("VBA của file BK không còn trong bản làm việc.")
        if plan.target_vba_present and not _package_has_vba(target_path):
            raise PaymentSyncError("VBA của file Thanh toán không còn trong bản làm việc.")
        target_book = self.gateway.load(target_path, read_only=False)
        try:
            all_items = {item.item_id: item for item in plan.items}
            for target_type, target_plan in plan.targets.items():
                worksheet = target_book[target_plan.sheet_name]
                profile = PAYMENT_PROFILES[target_type]
                profile_resolved = profile.resolve(worksheet)
                for item_id, row in written.get(target_type, {}).items():
                    item = all_items[item_id]
                    if _parse_sqt(
                        worksheet.cell(row, profile_resolved.columns["sqt"]).value
                    ) != item.sqt:
                        raise PaymentSyncError(f"SQT dòng {row} không qua kiểm tra.")
                    if _container_key(
                        worksheet.cell(row, profile_resolved.columns["container"]).value
                    ) != item.container:
                        raise PaymentSyncError(f"Container dòng {row} không qua kiểm tra.")
                    for field in profile.managed_fields:
                        actual = worksheet.cell(row, profile_resolved.columns[field]).value
                        incoming = item.values[field]
                        if _has_money(incoming) and not _numeric_equal(actual, incoming):
                            raise PaymentSyncError(
                                f"Giá trị {field} dòng {row} không qua kiểm tra."
                            )
                        if field in clear_fields.get(item_id, set()) and _has_money(actual):
                            raise PaymentSyncError(
                                f"Giá trị {field} dòng {row} chưa được xóa theo lựa chọn."
                            )
                    if worksheet.cell(
                        row, profile_resolved.columns["date"]
                    ).value != timestamp:
                        raise PaymentSyncError(
                            f"Date cập nhật dòng {row} không qua kiểm tra."
                        )
        finally:
            target_book.close()

    def cancel(self, plan: PaymentSyncPlan) -> None:
        self._update_run(plan.run_id, status=ExcelRunStatus.CANCELLED)

    def _create_run(
        self,
        source: Path,
        target: Path,
        source_fp: Any,
        target_fp: Any,
    ) -> int | None:
        if self.run_repository is None:
            return None
        record = self.run_repository.create_run(
            operation=ExcelOperation.PAYMENT_SYNC,
            source_path=source,
            target_path=target,
            source_fingerprint=source_fp,
            target_fingerprint_before=target_fp,
            status=ExcelRunStatus.ANALYZING,
        )
        return int(getattr(record, "id", record))

    def _update_run(self, run_id: int | None, **changes: Any) -> None:
        if self.run_repository is not None and run_id is not None:
            self.run_repository.update_run(run_id, **changes)

    def _finish_result(self, result: PaymentSyncResult) -> None:
        if self.run_repository is None or result.run_id is None:
            return
        self.run_repository.finish_run(
            result.run_id,
            status=result.status,
            sheet_name=result.sheet_name,
            backup_path=result.backup_path,
            target_fingerprint_after=result.fingerprint_after,
            total_items=(
                result.inserted_rows
                + result.updated_rows
                + result.unchanged_rows
                + result.skipped_rows
            ),
            changed_items=result.inserted_rows + result.updated_rows,
            skipped_items=result.skipped_rows,
            conflict_count=result.conflict_count,
        )

    def _finish_failed(self, run_id: int | None, exc: Exception) -> None:
        if self.run_repository is not None and run_id is not None:
            self.run_repository.finish_run(
                run_id,
                status=ExcelRunStatus.FAILED,
                error_message=str(exc),
            )


__all__ = [
    "BK_HEADER_ALIASES",
    "DATE_HEADER",
    "DATE_NUMBER_FORMAT",
    "HP_FIELDS",
    "NAM_FIELDS",
    "HPSheetProfile",
    "NAMSheetProfile",
    "PAYMENT_FIELDS",
    "PaymentSyncError",
    "PaymentSyncService",
    "SUMMARY_HEADERS",
    "find_summary_start",
    "normalize_bk_workbook",
    "refresh_bk_summary_formulas",
]
