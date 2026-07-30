from __future__ import annotations

from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from app.services.excel.daily_sync import (
    SOURCE_HEADER_ALIASES,
    SYNC_FIELDS,
    DailySyncService,
)
from app.services.excel.headers import (
    HeaderResolutionError,
    HeaderResolver,
    normalize_header,
)
from app.services.excel.resolvers import (
    MonthSheetService,
    YearResolutionError,
    YearResolver,
)
from app.services.excel.models import RowCandidate, TargetCellKind
from app.services.excel.posting import ExpensePostingService, classify_target_cell
from app.services.excel.workbook import (
    ExcelBackupService,
    WorkbookChangedError,
    WorkbookError,
    WorkbookGateway,
)


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("  Số \n Container — ", "so container"),
        ("ĐƠN GIÁ", "don gia"),
        ("VS D/O LỆNH", "vs d o lenh"),
        ("ＳＱＴ　ＰＭ", "sqt pm"),
    ],
)
def test_header_normalization_is_unicode_and_punctuation_tolerant(
    raw: str,
    expected: str,
) -> None:
    assert normalize_header(raw) == expected


def test_header_resolver_finds_a_multirow_header_with_merged_anchors() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A4"] = "SQT\nPM"
    sheet.merge_cells("A4:A5")
    sheet["B4"] = "Thông tin lô hàng"
    sheet["B5"] = "Số Container"
    sheet["C4"] = "Chi phí"
    sheet["C5"] = "Đơn giá"

    resolution = HeaderResolver().resolve(
        sheet,
        {
            "sqt": ("SQT PM",),
            "container": ("Thông tin lô hàng Số Container",),
            "road_rate": ("Chi phí ĐƠN GIÁ",),
        },
    )

    assert resolution.columns == {"sqt": 1, "container": 2, "road_rate": 3}
    assert (resolution.row_start, resolution.row_end) == (4, 5)
    assert resolution.column_letter("container") == "B"


def test_header_resolver_reports_missing_required_fields() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["SQT PM", "Số Container"])

    with pytest.raises(HeaderResolutionError) as error:
        HeaderResolver().resolve(
            sheet,
            {
                "sqt": ("SQT PM",),
                "container": ("Số Container",),
                "closing_date": ("Ngày Đóng",),
            },
        )

    assert error.value.missing == ("closing_date",)
    assert "closing_date" in str(error.value)


def test_month_and_year_resolvers_never_guess_from_system_date() -> None:
    months = MonthSheetService()
    years = YearResolver()

    assert months.parse_daily_sheet("  THÁNG   7 ") == 7
    assert months.parse_target_sheet("T07 26") == (7, 2026)
    assert months.target_name(7, 2026) == "T07 26"
    assert (
        months.nearest_previous_template(
            ["T03 26", "T06 26", "T08 26"],
            month=7,
            year=2026,
        )
        == "T06 26"
    )
    assert years.from_filename(Path("Hàng ngày - 2026.xlsx")) == 2026
    assert years.target_year(
        Path("BK Tổng hợp.xlsx"),
        ["T06 26", "T07 26"],
    ) == 2026

    with pytest.raises(YearResolutionError):
        years.from_filename(Path("Hàng ngày.xlsx"))
    with pytest.raises(YearResolutionError):
        years.from_filename(Path("Hàng ngày 2025 sang 2026.xlsx"))
    with pytest.raises(YearResolutionError):
        years.target_year(
            Path("BK 2025 sang 2026.xlsx"),
            ["T07 26"],
        )
    with pytest.raises(YearResolutionError):
        years.from_target_sheets(["T12 25", "T01 26"])


def test_workbook_gateway_detects_changes_after_analysis(tmp_path: Path) -> None:
    path = tmp_path / "BK 2026.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "before"
    workbook.save(path)
    expected = WorkbookGateway().fingerprint(path)

    changed = Workbook()
    changed.active["A1"] = "after"
    changed.save(path)

    with pytest.raises(WorkbookChangedError):
        WorkbookGateway().assert_unchanged(path, expected, label="File BK")


def test_workbook_gateway_rejects_legacy_xls_without_touching_it(
    tmp_path: Path,
) -> None:
    legacy = tmp_path / "legacy.xls"
    legacy.write_bytes(b"not-an-openxml-workbook")

    with pytest.raises(WorkbookError, match=r"\.xls"):
        WorkbookGateway().load(legacy)

    assert legacy.read_bytes() == b"not-an-openxml-workbook"


@pytest.mark.parametrize(
    ("value", "amount", "expected"),
    [
        (None, 100, TargetCellKind.EMPTY),
        ("", 100, TargetCellKind.EMPTY),
        (0, 100, TargetCellKind.ZERO),
        (100, 100, TargetCellKind.SAME_VALUE),
        (100.0, 100, TargetCellKind.SAME_VALUE),
        (50, 100, TargetCellKind.NUMBER),
        ("=SUM(A1:A2)", 100, TargetCellKind.FORMULA),
        ("đã nhập", 100, TargetCellKind.TEXT),
        (True, 100, TargetCellKind.TEXT),
    ],
)
def test_posting_target_cell_classification(
    value: Any,
    amount: int,
    expected: TargetCellKind,
) -> None:
    workbook = Workbook()
    cell = workbook.active["D7"]
    cell.value = value

    state = classify_target_cell(cell, amount)

    assert state.kind is expected
    assert state.coordinate == "D7"
    assert state.value == value


@pytest.mark.parametrize(
    "candidate",
    [
        RowCandidate(
            row=2,
            sqt=None,
            container="DRYU3026167",
            closing_date="2026-07-28",
        ),
        RowCandidate(
            row=2,
            sqt=700,
            container="DRYU3026167",
            closing_date=None,
        ),
        RowCandidate(
            row=2,
            sqt=700,
            container="DRYU3026167",
            closing_date="2026-07-28",
            cargo_type="Ron",
            is_ron=True,
        ),
    ],
)
def test_posting_auto_selects_the_only_matching_row(
    candidate: RowCandidate,
) -> None:
    assert ExpensePostingService._automatic_row([candidate]) is candidate


def test_posting_auto_selects_a_single_complete_non_ron_row() -> None:
    candidate = RowCandidate(
        row=2,
        sqt=700,
        container="DRYU3026167",
        closing_date="2026-07-28",
    )

    assert ExpensePostingService._automatic_row([candidate]) is candidate


def test_working_copy_is_created_next_to_target_for_atomic_replace(
    tmp_path: Path,
) -> None:
    target = tmp_path / "target-volume" / "BK 2026.xlsx"
    backup_dir = tmp_path / "other-volume" / "Backup"
    target.parent.mkdir(parents=True)
    workbook = Workbook()
    workbook.save(target)
    workbook.close()

    service = ExcelBackupService(backup_dir)
    working = service.create_working_copy(target, run_id=7)

    try:
        assert working.parent == target.parent
        assert working.read_bytes() == target.read_bytes()
    finally:
        working.unlink(missing_ok=True)


def test_backup_service_keeps_only_latest_valid_backup(tmp_path: Path) -> None:
    target = tmp_path / "BK 2026.xlsx"
    backup_dir = tmp_path / "Backup"
    workbook = Workbook()
    workbook.active["A1"] = "first"
    workbook.save(target)
    workbook.close()
    service = ExcelBackupService(backup_dir)
    backup_dir.mkdir()
    legacy = backup_dir / "BK 2026_20260101_120000_1.xlsx"
    legacy.write_bytes(target.read_bytes())

    first_hash = target.read_bytes()
    first_backup = service.create_backup(target)

    workbook = load_workbook(target)
    workbook.active["A1"] = "second"
    workbook.save(target)
    workbook.close()
    second_hash = target.read_bytes()
    second_backup = service.create_backup(target)

    assert first_backup == second_backup
    assert second_backup.name == "BK 2026_latest.xlsx"
    assert second_backup.read_bytes() == second_hash
    assert second_backup.read_bytes() != first_hash
    assert list(backup_dir.glob("*.xlsx")) == [second_backup]
    assert not legacy.exists()


def test_excel_artifact_cleanup_is_narrowly_scoped(tmp_path: Path) -> None:
    target = tmp_path / "BK 2026.xlsx"
    source = tmp_path / "Hàng ngày 2026.xlsx"
    temp_dir = tmp_path / "Temp"
    workbook = Workbook()
    workbook.save(target)
    workbook.save(source)
    workbook.close()
    backups = ExcelBackupService(tmp_path / "Backup")
    gateway = WorkbookGateway()
    working = backups.create_working_copy(target, run_id=7)
    snapshot = gateway.source_snapshot(source, temp_dir)
    unrelated = temp_dir / "unrelated.snapshot.xlsx"
    unrelated.write_bytes(source.read_bytes())

    assert backups.cleanup_working_copies(target) == 1
    assert gateway.cleanup_source_snapshots(source, temp_dir) == 1
    assert not working.exists()
    assert not snapshot.exists()
    assert unrelated.exists()


def test_daily_sync_real_data_boundary_ignores_phantom_max_row() -> None:
    workbook = Workbook()
    sheet = workbook.active
    for column, aliases in enumerate(SOURCE_HEADER_ALIASES.values(), 1):
        if column <= 11:
            sheet.cell(1, column).value = aliases[0]
    sheet["P1"] = SOURCE_HEADER_ALIASES["transport"][0]
    sheet["A2"] = 700
    sheet["A65000"].fill = PatternFill("solid", fgColor="FFFFFF")
    resolution = HeaderResolver().resolve(
        sheet,
        SOURCE_HEADER_ALIASES,
        required=SYNC_FIELDS,
    )

    assert sheet.max_row == 65000
    assert DailySyncService._last_data_row(sheet, resolution) == 2
    workbook.close()


def test_lam_lenh_is_not_resolved_without_ghi_chu_boundary() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["SQT", "Ngày Đóng", "Số Container", "LÀM LỆNH"])
    service = object.__new__(ExpensePostingService)
    service.headers = HeaderResolver()

    base = service._resolve_base_headers(sheet)

    assert "LL" not in service._resolve_fee_columns(sheet, base)
