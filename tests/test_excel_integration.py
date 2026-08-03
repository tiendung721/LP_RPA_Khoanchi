from __future__ import annotations

import hashlib
import json
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Iterable

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from app.services.excel.daily_sync import (
    SOURCE_HEADER_ALIASES,
    DailySyncError,
    DailySyncService,
    parse_sqt,
)
from app.services.excel.models import (
    ConflictType,
    ExcelRunStatus,
    PostingItemStatus,
    ResolutionAction,
    TargetCellKind,
)
from app.services.excel.posting import (
    FEE_HEADER_ALIASES,
    ExpensePostingError,
    ExpensePostingService,
)
from app.services.excel.workbook import WorkbookChangedError


SYNC_HEADERS = [
    aliases[0] for aliases in SOURCE_HEADER_ALIASES.values()
]

POSTING_BASE_HEADERS = (
    "SQT PM",
    "Ngày Đóng",
    "Số Container",
    "Loại hàng",
    "Tên tàu",
    "Người nhận",
)
POSTING_FEE_COLUMNS = {
    fee: column
    for column, fee in enumerate(FEE_HEADER_ALIASES, len(POSTING_BASE_HEADERS) + 1)
}


class _ImmediateStabilityChecker:
    def wait(self, _path: str | Path) -> None:
        return None


class _ReadyProvider:
    def __init__(self, path: Path) -> None:
        self.path = path

    def get_latest_ready_json_path(self) -> Path:
        return self.path

    def get_ready_json_path(self, _batch_id: int) -> Path:
        return self.path


class _PostedIndexRepository:
    def __init__(self, indices: set[int]) -> None:
        self.indices = indices

    def successful_source_indices(self, _batch_hash: str) -> set[int]:
        return set(self.indices)


class _RunHistoryRepository:
    def __init__(self) -> None:
        self.created: list[dict[str, Any]] = []
        self.updated: list[dict[str, Any]] = []
        self.finished: list[dict[str, Any]] = []

    def create_run(self, **values: Any) -> Any:
        self.created.append(dict(values))
        return SimpleNamespace(id=len(self.created))

    def update_run(self, run_id: int, **values: Any) -> None:
        self.updated.append({"run_id": run_id, **values})

    def finish_run(self, run_id: int, **values: Any) -> None:
        self.finished.append({"run_id": run_id, **values})

    def get_latest_sync_sheet(self) -> None:
        return None


class _PostingHistoryRepository:
    def __init__(self) -> None:
        self.items: list[dict[str, Any]] = []
        self.metadata: dict[str, Any] = {}

    def successful_source_indices(self, _batch_hash: str) -> set[int]:
        return set()

    def create_items(self, items: Iterable[dict[str, Any]], **values: Any) -> list[Any]:
        self.items.extend(dict(item) for item in items)
        self.metadata = dict(values)
        return []


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _sync_row(
    sqt: Any,
    container: str = "DRYU3026167",
    *,
    closing_date: str = "2026-07-28",
    weight: int = 20,
    cargo_type: str = "Gạo",
    closing_place: str = "Kho A",
    transport: str = "Xe A",
) -> list[Any]:
    return [
        sqt,
        closing_date,
        container,
        weight,
        cargo_type,
        closing_place,
        "Tàu A",
        "2026-07-30",
        "2026-08-02",
        "Công ty B",
        "VTB",
        transport,
    ]


def _save_daily(
    path: Path,
    rows: Iterable[Iterable[Any]],
    *,
    month: int = 7,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Tháng {month}"
    sheet.append(SYNC_HEADERS)
    for row in rows:
        sheet.append(list(row))
    workbook.save(path)
    workbook.close()


def _populate_target_sheet(
    sheet: Any,
    rows: Iterable[Iterable[Any]],
) -> None:
    for column, header in enumerate(SYNC_HEADERS[:11], 1):
        sheet.cell(1, column).value = header
    sheet["L1"] = "Chi phí L"
    sheet["M1"] = "Chi phí M"
    sheet["N1"] = "Chi phí N"
    sheet["O1"] = "Chi phí O"
    sheet["P1"] = SYNC_HEADERS[11]
    sheet["Q1"] = "Hóa đơn"
    for row_number, source_row in enumerate(rows, 2):
        source_values = list(source_row)
        for column, value in enumerate(source_values[:11], 1):
            sheet.cell(row_number, column).value = value
        sheet.cell(row_number, 16).value = source_values[11]


def _save_target(
    path: Path,
    rows: Iterable[Iterable[Any]],
    *,
    month: int,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"T{month:02d} 26"
    _populate_target_sheet(sheet, rows)
    workbook.save(path)
    workbook.close()


def _save_ready(path: Path, rows: Iterable[Iterable[Any]]) -> None:
    path.write_text(
        json.dumps({"v": 1, "d": [list(row) for row in rows]}, ensure_ascii=False),
        encoding="utf-8",
    )


def _add_posting_row(
    sheet: Any,
    row: int,
    container: str,
    *,
    sqt: Any = 700,
    closing_date: Any = "2026-07-28",
    cargo_type: Any = "Gạo",
) -> None:
    sheet.cell(row, 1).value = sqt
    sheet.cell(row, 2).value = closing_date
    sheet.cell(row, 3).value = container
    sheet.cell(row, 4).value = cargo_type
    sheet.cell(row, 5).value = "Tàu A"
    sheet.cell(row, 6).value = "Công ty B"


def _new_posting_sheet(workbook: Workbook, name: str) -> Any:
    sheet = workbook.active
    if sheet.max_row == 1 and sheet.max_column == 1 and sheet["A1"].value is None:
        sheet.title = name
    else:
        sheet = workbook.create_sheet(name)
    for column, title in enumerate(POSTING_BASE_HEADERS, 1):
        sheet.cell(1, column).value = title
    for fee, column in POSTING_FEE_COLUMNS.items():
        sheet.cell(1, column).value = FEE_HEADER_ALIASES[fee][0]
    sheet.cell(1, len(POSTING_BASE_HEADERS) + len(POSTING_FEE_COLUMNS) + 1).value = (
        "GHI CHÚ"
    )
    sheet.cell(1, len(POSTING_BASE_HEADERS) + len(POSTING_FEE_COLUMNS) + 2).value = (
        "Hóa đơn"
    )
    return sheet


def _posting_service(
    ready: Path,
    target: Path,
    runtime_dir: Path,
    *,
    posting_repository: Any | None = None,
    run_repository: Any | None = None,
    clock: Any | None = None,
) -> ExpensePostingService:
    return ExpensePostingService(
        _ReadyProvider(ready),
        bk_path=target,
        temp_dir=runtime_dir / "Temp",
        backup_dir=runtime_dir / "Backup",
        posting_repository=posting_repository,
        run_repository=run_repository,
        clock=clock,
    )


def _sync_service(
    daily: Path,
    target: Path,
    runtime_dir: Path,
    *,
    run_repository: Any | None = None,
) -> DailySyncService:
    return DailySyncService(
        daily_path=daily,
        bk_path=target,
        temp_dir=runtime_dir / "Temp",
        backup_dir=runtime_dir / "Backup",
        stability_checker=_ImmediateStabilityChecker(),
        run_repository=run_repository,
    )


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        (700, 700),
        (700.0, 700),
        (" 7 0 0 ", 700),
        (True, None),
        (0, None),
        (-1, None),
        (700.5, None),
        ("700A", None),
        ("", None),
        (None, None),
    ],
)
def test_parse_sqt_accepts_only_positive_integers(
    raw: Any,
    expected: int | None,
) -> None:
    assert parse_sqt(raw) == expected


def test_daily_sync_compares_full_sheet_preserves_order_and_is_idempotent(
    tmp_path: Path,
) -> None:
    daily = tmp_path / "Hàng ngày 2026.xlsx"
    target = tmp_path / "BK 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    old_row = _sync_row(699, "OLDU0000001")
    duplicate = _sync_row(701, "NEWC0000003")
    _save_daily(
        daily,
        [
            old_row,
            _sync_row(700, "NEWA0000001"),
            _sync_row(700, "NEWB0000002"),
            duplicate,
            duplicate,
            _sync_row("not-an-sqt", "BADS0000000"),
        ],
    )
    _save_target(target, [old_row], month=7)
    workbook = load_workbook(target)
    sheet = workbook["T07 26"]
    sheet["L2"] = 125_000
    sheet["M2"] = "=1+1"
    sheet["N2"] = "do not touch"
    sheet["O2"] = 0
    sheet["Q2"] = "INV-001"
    workbook.save(target)
    workbook.close()
    source_before = _sha256(daily)
    target_before = _sha256(target)

    service = _sync_service(daily, target, runtime_dir)
    plan = service.analyze(source_sheet_name="Tháng 7")

    assert plan.selected_month == 7
    assert [row.sqt for row in plan.rows] == [700, 700, 701, 701]
    assert plan.conflicts == []
    assert plan.insert_count == 4
    assert plan.update_count == 0
    assert plan.unchanged_count == 1
    assert plan.invalid_count == 1
    assert plan.requires_user_input

    result = service.apply(plan, {})

    assert result.status is ExcelRunStatus.SUCCEEDED
    assert result.added_rows == 4
    assert result.inserted_rows == 4
    assert result.updated_rows == 0
    assert result.skipped_rows == 1
    assert result.backup_path is not None
    assert result.backup_path.is_file()
    assert result.backup_path.name == "BK 2026_latest.xlsx"
    assert _sha256(result.backup_path) == target_before
    assert _sha256(daily) == source_before
    workbook = load_workbook(target, data_only=False)
    try:
        sheet = workbook["T07 26"]
        assert [sheet.cell(row, 1).value for row in range(2, 7)] == [
            699,
            700,
            700,
            701,
            701,
        ]
        assert [sheet.cell(row, 3).value for row in range(3, 7)] == [
            "NEWA0000001",
            "NEWB0000002",
            "NEWC0000003",
            "NEWC0000003",
        ]
        assert sheet["F3"].value == "Kho A"
        assert sheet["P3"].value == "Xe A"
        assert [sheet.cell(2, column).value for column in range(12, 16)] == [
            125_000,
            "=1+1",
            "do not touch",
            0,
        ]
        assert sheet["Q2"].value == "INV-001"
    finally:
        workbook.close()

    target_after_first_apply = _sha256(target)
    backup_count = len(list((runtime_dir / "Backup").iterdir()))
    second_plan = service.analyze()
    second_result = service.apply(second_plan, {})

    assert not second_plan.has_changes
    assert second_result.status is ExcelRunStatus.NO_CHANGES
    assert second_result.backup_path is None
    assert second_plan.invalid_count == 1
    assert _sha256(target) == target_after_first_apply
    assert len(list((runtime_dir / "Backup").iterdir())) == backup_count


def test_daily_sync_updates_existing_rows_and_preserves_bk_only_columns(
    tmp_path: Path,
) -> None:
    daily = tmp_path / "Hàng ngày 2026.xlsx"
    target = tmp_path / "BK 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    updated = _sync_row(
        699,
        "NEWU0000001",
        weight=28,
        closing_place="Kho mới",
        transport="Xe mới",
    )
    inserted = _sync_row(700, "NEWA0000002")
    _save_daily(daily, [updated, inserted])
    _save_target(
        target,
        [
            _sync_row(698, "TARGET000001"),
            _sync_row(699, "OLDU0000001"),
        ],
        month=7,
    )
    workbook = load_workbook(target)
    sheet = workbook["T07 26"]
    sheet["L3"] = 125_000
    sheet["M3"] = "=1+1"
    sheet["N3"] = "không được đổi"
    sheet["O3"] = 0
    sheet["Q3"] = "INV-001"
    workbook.save(target)
    workbook.close()

    service = _sync_service(daily, target, runtime_dir)
    plan = service.analyze(source_sheet_name="Tháng 7")

    assert plan.update_count == 1
    assert plan.insert_count == 1
    assert plan.unchanged_count == 0
    assert plan.target_only_count == 1

    result = service.apply(plan, {})

    assert result.updated_rows == 1
    assert result.inserted_rows == 1
    assert result.target_only_rows == 1
    workbook = load_workbook(target, data_only=False)
    try:
        sheet = workbook["T07 26"]
        assert sheet["A2"].value == 698
        assert sheet["C2"].value == "TARGET000001"
        assert sheet["C3"].value == "NEWU0000001"
        assert sheet["D3"].value == 28
        assert sheet["F3"].value == "Kho mới"
        assert sheet["P3"].value == "Xe mới"
        assert [sheet.cell(3, column).value for column in range(12, 16)] == [
            125_000,
            "=1+1",
            "không được đổi",
            0,
        ]
        assert sheet["Q3"].value == "INV-001"
        assert sheet["A4"].value == 700
        assert sheet["C4"].value == "NEWA0000002"
    finally:
        workbook.close()


def test_daily_sync_blocks_mismatched_duplicate_sqt_without_backup(
    tmp_path: Path,
) -> None:
    daily = tmp_path / "Hàng ngày 2026.xlsx"
    target = tmp_path / "BK 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    _save_daily(
        daily,
        [
            _sync_row(700, "NEWA0000001"),
            _sync_row(700, "NEWB0000002"),
        ],
    )
    _save_target(target, [_sync_row(700, "NEWA0000001")], month=7)
    service = _sync_service(daily, target, runtime_dir)

    plan = service.analyze(source_sheet_name="Tháng 7")

    assert plan.conflict_count == 1
    assert plan.conflicts[0].conflict_type is (
        ConflictType.SYNC_GROUP_COUNT_MISMATCH
    )
    with pytest.raises(DailySyncError, match="số dòng"):
        service.apply(plan, {})
    assert not (runtime_dir / "Backup").exists()


def test_daily_sync_aborts_when_source_changes_after_analysis(
    tmp_path: Path,
) -> None:
    daily = tmp_path / "Hàng ngày 2026.xlsx"
    target = tmp_path / "BK 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    _save_daily(daily, [_sync_row(700)])
    _save_target(target, [_sync_row(699)], month=7)
    service = _sync_service(daily, target, runtime_dir)
    plan = service.analyze(source_sheet_name="Tháng 7")

    _save_daily(daily, [_sync_row(700, "CHANGED00001")])

    with pytest.raises(WorkbookChangedError):
        service.apply(plan, {})
    assert not (runtime_dir / "Backup").exists()


def test_daily_sync_accepts_vt_bo_as_target_transport_header(
    tmp_path: Path,
) -> None:
    daily = tmp_path / "Hàng ngày 2026.xlsx"
    target = tmp_path / "BK 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    _save_daily(daily, [_sync_row(700)], month=7)
    _save_target(target, [_sync_row(699)], month=7)

    workbook = load_workbook(target)
    try:
        workbook["T07 26"]["P1"] = "VT bộ"
        workbook.save(target)
    finally:
        workbook.close()

    service = _sync_service(daily, target, runtime_dir)
    plan = service.analyze(source_sheet_name="Tháng 7")

    assert plan.selected_sheet == "T07 26"
    assert [row.sqt for row in plan.rows] == [700]
    service.cancel(plan)


def test_daily_sync_lists_month_sheets_and_analyzes_only_selected_source(
    tmp_path: Path,
) -> None:
    daily = tmp_path / "Hàng ngày.xlsx"
    target = tmp_path / "BK.xlsx"
    runtime_dir = tmp_path / "Excel"

    workbook = Workbook()
    july = workbook.active
    july.title = "Tháng 7"
    july.append(SYNC_HEADERS)
    july.append(_sync_row(700, "JULY0000001"))
    august = workbook.create_sheet("Tháng 8")
    august.append(["Header sai"])
    august.append([800])
    workbook.create_sheet("Ghi chú")
    workbook.save(daily)
    workbook.close()

    workbook = Workbook()
    july_target = workbook.active
    july_target.title = "T07 26"
    _populate_target_sheet(july_target, [_sync_row(699, "OLDU0000001")])
    august_target = workbook.create_sheet("T08 26")
    _populate_target_sheet(august_target, [_sync_row(799, "OLDU0000002")])
    workbook.save(target)
    workbook.close()

    service = _sync_service(daily, target, runtime_dir)

    assert [
        (candidate.month, candidate.sheet_name)
        for candidate in service.source_sheet_candidates()
    ] == [(7, "Tháng 7"), (8, "Tháng 8")]

    progress: list[str] = []
    plan = service.analyze(
        source_sheet_name="Tháng 7",
        progress_callback=progress.append,
    )

    assert plan.selected_month == 7
    assert plan.selected_sheet == "T07 26"
    assert [row.sqt for row in plan.rows] == [700]
    assert [candidate.source_sheet for candidate in plan.month_candidates] == [
        "Tháng 7"
    ]
    assert any("Tháng 7" in message for message in progress)
    assert not any("Tháng 8" in message for message in progress)
    service.cancel(plan)


def test_daily_sync_rejects_source_sheet_not_in_daily_workbook(
    tmp_path: Path,
) -> None:
    daily = tmp_path / "Hàng ngày.xlsx"
    target = tmp_path / "BK.xlsx"
    _save_daily(daily, [_sync_row(700)], month=7)
    _save_target(target, [_sync_row(699)], month=7)
    service = _sync_service(daily, target, tmp_path / "Excel")

    with pytest.raises(DailySyncError, match="Không tìm thấy sheet nguồn"):
        service.analyze(source_sheet_name="Tháng 8")


def test_daily_sync_creates_new_month_from_previous_nonempty_template(
    tmp_path: Path,
) -> None:
    daily = tmp_path / "Hàng ngày 2026.xlsx"
    target = tmp_path / "BK Tổng hợp 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    _save_daily(daily, [_sync_row(700)], month=7)
    _save_target(target, [_sync_row(699, "OLDU0000001")], month=6)
    workbook = load_workbook(target)
    template = workbook["T06 26"]
    template.freeze_panes = "A2"
    template.page_setup.orientation = "landscape"
    template.auto_filter.ref = "A1:Q2"
    template.merge_cells("R1:S1")
    template["R1"] = "Thông tin mẫu"
    template["C2"].fill = PatternFill("solid", fgColor="FFF2CC")
    template["C2"].comment = Comment("old row", "test")
    validation = DataValidation(type="list", formula1='"A,B"')
    template.add_data_validation(validation)
    validation.add("C2:C100")
    template.conditional_formatting.add(
        "A2:A100",
        CellIsRule(operator="greaterThan", formula=["0"]),
    )
    workbook.save(target)
    workbook.close()

    service = _sync_service(daily, target, runtime_dir)
    result = service.apply(service.analyze(), {})

    assert result.status is ExcelRunStatus.SUCCEEDED
    assert result.sheet_name == "T07 26"
    workbook = load_workbook(target, data_only=False)
    try:
        assert workbook.sheetnames == ["T06 26", "T07 26"]
        template = workbook["T06 26"]
        created = workbook["T07 26"]
        assert template["A2"].value == 699
        assert template["C2"].value == "OLDU0000001"
        assert created["A2"].value == 700
        assert created["C2"].value == "DRYU3026167"
        assert created["C2"].comment is None
        assert created["C2"].fill.fgColor.rgb == template["C2"].fill.fgColor.rgb
        assert created.freeze_panes == "A2"
        assert created.page_setup.orientation == "landscape"
        assert created.auto_filter.ref == "A1:Q2"
        assert "R1:S1" in {str(item) for item in created.merged_cells.ranges}
        assert created["R1"].value == "Thông tin mẫu"
        assert len(created.data_validations.dataValidation) == 1
        assert len(created.conditional_formatting) == 1
        assert created["L2"].value is None
        assert created["Q2"].value is None
    finally:
        workbook.close()


def test_daily_sync_aborts_before_backup_when_bk_changed_after_analyze(
    tmp_path: Path,
) -> None:
    daily = tmp_path / "Hàng ngày 2026.xlsx"
    target = tmp_path / "BK 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    _save_daily(daily, [_sync_row(700)])
    _save_target(target, [_sync_row(699)], month=7)
    service = _sync_service(daily, target, runtime_dir)
    plan = service.analyze()

    workbook = load_workbook(target)
    workbook["T07 26"]["L2"] = "external change"
    workbook.save(target)
    workbook.close()
    changed_hash = _sha256(target)

    with pytest.raises(WorkbookChangedError):
        service.apply(plan, {})

    assert _sha256(target) == changed_hash
    assert not (runtime_dir / "Backup").exists()


def test_posting_groups_normalized_container_and_keeps_bl_only_rows_separate(
    tmp_path: Path,
) -> None:
    ready = tmp_path / "ready.json"
    target = tmp_path / "BK 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    _save_ready(
        ready,
        [
            [" dryu-302 6167 ", None, "VTN", "ST", 100],
            ["DRYU3026167", "BL-01", "VTN", "ST", 250],
            [None, "BL-ONLY-01", "CB", "HD", 500],
            [None, "BL-ONLY-02", "CB", "HD", 600],
        ],
    )
    workbook = Workbook()
    sheet = _new_posting_sheet(workbook, "T07 26")
    _add_posting_row(sheet, 2, "DRYU3026167")
    workbook.save(target)
    workbook.close()

    service = _posting_service(ready, target, runtime_dir)
    unselected = service.analyze()
    assert unselected.selected_sheet is None
    assert [item.target_sheet for item in unselected.sheet_candidates] == ["T07 26"]
    plan = service.analyze(sheet_name="T07 26")

    assert plan.selected_sheet == "T07 26"
    assert len(plan.items) == 4
    assert plan.items[0].container == "DRYU3026167"
    assert plan.items[0].source_indices == [0]
    assert plan.items[0].amount == 100
    assert plan.items[1].source_indices == [1]
    assert plan.items[1].amount == 250
    assert plan.items[2].source_indices == [2]
    assert plan.items[3].source_indices == [3]
    assert [
        conflict.conflict_type for conflict in plan.conflicts
    ] == [
        ConflictType.REPEATED_SOURCE_CONTAINER,
        ConflictType.REPEATED_SOURCE_CONTAINER,
        ConflictType.BL_ONLY_NO_CONTAINER,
        ConflictType.BL_ONLY_NO_CONTAINER,
    ]

    repeated = [
        conflict
        for conflict in plan.conflicts
        if conflict.conflict_type is ConflictType.REPEATED_SOURCE_CONTAINER
    ]
    refined = service.refine(
        plan,
        {
            conflict.conflict_id: {
                "action": "SELECT_ROW",
                "selected_row": 2,
            }
            for conflict in repeated
        },
    )

    assert len(refined.items) == 3
    assert refined.items[0].source_indices == [0, 1]
    assert refined.items[0].amount == 350
    assert [
        conflict.conflict_type for conflict in refined.conflicts
    ] == [
        ConflictType.BL_ONLY_NO_CONTAINER,
        ConflictType.BL_ONLY_NO_CONTAINER,
    ]

    result = service.apply(refined, {})

    assert result.status is ExcelRunStatus.SUCCEEDED
    assert result.posted_source_items == 2
    assert result.written_cells == 1
    assert result.skipped_source_items == 2
    assert result.backup_path is None
    assert not (runtime_dir / "Backup").exists()
    workbook = load_workbook(target, data_only=False)
    try:
        assert workbook["T07 26"].cell(
            2, POSTING_FEE_COLUMNS["VTN"]
        ).value == 350
    finally:
        workbook.close()


def test_posting_keeps_repeated_container_separate_for_different_sqt_choices(
    tmp_path: Path,
) -> None:
    ready = tmp_path / "ready.json"
    target = tmp_path / "BK 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    container = "DRYU3026167"
    _save_ready(
        ready,
        [
            [container, "BL-01", "VTN", "ST", 100],
            [container, "BL-02", "VTN", "ST", 250],
        ],
    )
    workbook = Workbook()
    sheet = _new_posting_sheet(workbook, "T07 26")
    _add_posting_row(sheet, 2, container, sqt=701)
    _add_posting_row(sheet, 3, container, sqt=702)
    workbook.save(target)
    workbook.close()

    service = _posting_service(ready, target, runtime_dir)
    plan = service.analyze(sheet_name="T07 26")
    conflicts = [
        conflict
        for conflict in plan.conflicts
        if conflict.conflict_type is ConflictType.REPEATED_SOURCE_CONTAINER
    ]

    assert len(plan.items) == 2
    assert [item.source_indices for item in plan.items] == [[0], [1]]
    assert [conflict.amount for conflict in conflicts] == [100, 250]
    assert len(conflicts) == 2

    refined = service.refine(
        plan,
        {
            conflicts[0].conflict_id: {
                "action": "SELECT_ROW",
                "selected_row": 2,
            },
            conflicts[1].conflict_id: {
                "action": "SELECT_ROW",
                "selected_row": 3,
            },
        },
    )

    assert not refined.conflicts
    assert [(item.target_row, item.amount) for item in refined.items] == [
        (2, 100),
        (3, 250),
    ]

    result = service.apply(refined, {})

    assert result.posted_source_items == 2
    assert result.written_cells == 2
    workbook = load_workbook(target, data_only=False)
    try:
        sheet = workbook["T07 26"]
        column = POSTING_FEE_COLUMNS["VTN"]
        assert sheet.cell(2, column).value == 100
        assert sheet.cell(3, column).value == 250
    finally:
        workbook.close()


def test_posting_sums_repeated_container_only_after_same_sqt_is_selected(
    tmp_path: Path,
) -> None:
    ready = tmp_path / "ready.json"
    target = tmp_path / "BK 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    container = "DRYU3026167"
    _save_ready(
        ready,
        [
            [container, "BL-01", "VTN", "ST", 100],
            [container, "BL-02", "VTN", "ST", 250],
        ],
    )
    workbook = Workbook()
    sheet = _new_posting_sheet(workbook, "T07 26")
    _add_posting_row(sheet, 2, container, sqt=701)
    _add_posting_row(sheet, 3, container, sqt=702)
    workbook.save(target)
    workbook.close()

    service = _posting_service(ready, target, runtime_dir)
    plan = service.analyze(sheet_name="T07 26")
    conflicts = [
        conflict
        for conflict in plan.conflicts
        if conflict.conflict_type is ConflictType.REPEATED_SOURCE_CONTAINER
    ]
    refined = service.refine(
        plan,
        {
            conflict.conflict_id: {
                "action": "SELECT_ROW",
                "selected_row": 2,
            }
            for conflict in conflicts
        },
    )

    assert not refined.conflicts
    assert len(refined.items) == 1
    assert refined.items[0].source_indices == [0, 1]
    assert refined.items[0].amount == 350
    assert refined.items[0].target_row == 2

    result = service.apply(refined, {})

    assert result.posted_source_items == 2
    assert result.written_cells == 1
    workbook = load_workbook(target, data_only=False)
    try:
        sheet = workbook["T07 26"]
        column = POSTING_FEE_COLUMNS["VTN"]
        assert sheet.cell(2, column).value == 350
        assert sheet.cell(3, column).value is None
    finally:
        workbook.close()


def test_posting_requires_confirmation_for_repeated_container_with_one_bk_row(
    tmp_path: Path,
) -> None:
    ready = tmp_path / "ready.json"
    target = tmp_path / "BK 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    container = "DRYU3026167"
    _save_ready(
        ready,
        [
            [container, "BL-01", "VTN", "ST", 100],
            [container, "BL-02", "VTN", "ST", 250],
        ],
    )
    workbook = Workbook()
    sheet = _new_posting_sheet(workbook, "T07 26")
    _add_posting_row(sheet, 2, container, sqt=701)
    workbook.save(target)
    workbook.close()

    service = _posting_service(ready, target, runtime_dir)
    plan = service.analyze(sheet_name="T07 26")
    conflicts = [
        conflict
        for conflict in plan.conflicts
        if conflict.conflict_type is ConflictType.REPEATED_SOURCE_CONTAINER
    ]

    assert len(conflicts) == 2
    assert [conflict.amount for conflict in conflicts] == [100, 250]
    assert [candidate.sqt for candidate in conflicts[0].row_candidates] == [701]
    assert all(item.target_row is None for item in plan.items)

    refined = service.refine(
        plan,
        {
            conflicts[0].conflict_id: {
                "action": "SELECT_ROW",
                "selected_row": 2,
            },
            conflicts[1].conflict_id: {"action": "SKIP"},
        },
    )

    assert not refined.conflicts
    result = service.apply(refined, {})

    assert result.posted_source_items == 1
    assert result.skipped_source_items == 1
    assert result.written_cells == 1
    workbook = load_workbook(target, data_only=False)
    try:
        assert workbook["T07 26"].cell(
            2, POSTING_FEE_COLUMNS["VTN"]
        ).value == 100
    finally:
        workbook.close()


def test_posting_cell_conflicts_honor_add_overwrite_and_skip_resolutions(
    tmp_path: Path,
) -> None:
    ready = tmp_path / "ready.json"
    target = tmp_path / "BK 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    containers = [
        "DRYU3026167",
        "MSCU1234567",
        "OOLU7654321",
        "TGHU2345678",
        "CMAU3456789",
        "TEMU4567890",
    ]
    amounts = [100, 200, 300, 100, 400, 500]
    _save_ready(
        ready,
        [
            [container, None, "CB", "ST", amount]
            for container, amount in zip(containers, amounts, strict=True)
        ],
    )
    workbook = Workbook()
    sheet = _new_posting_sheet(workbook, "T07 26")
    for row, container in enumerate(containers, 2):
        _add_posting_row(sheet, row, container, sqt=698 + row)
    cb_column = POSTING_FEE_COLUMNS["CB"]
    sheet.cell(2, cb_column).value = None
    sheet.cell(3, cb_column).value = 0
    sheet.cell(4, cb_column).value = 300
    sheet.cell(5, cb_column).value = 50
    sheet.cell(6, cb_column).value = "=SUM(A1:A2)"
    sheet.cell(7, cb_column).value = "đã nhập tay"
    workbook.save(target)
    workbook.close()

    service = _posting_service(ready, target, runtime_dir)
    plan = service.analyze(sheet_name="T07 26")
    by_container = {item.container: item for item in plan.items}

    assert by_container[containers[0]].cell_state.kind is TargetCellKind.EMPTY
    assert by_container[containers[1]].cell_state.kind is TargetCellKind.ZERO
    assert by_container[containers[2]].status is PostingItemStatus.ALREADY_EXISTS
    conflicts = {
        conflict.conflict_type: conflict for conflict in plan.conflicts
    }
    assert set(conflicts) == {
        ConflictType.TARGET_CELL_OCCUPIED,
        ConflictType.TARGET_CELL_FORMULA,
        ConflictType.TARGET_CELL_TEXT,
    }
    assert conflicts[
        ConflictType.TARGET_CELL_OCCUPIED
    ].default_action is ResolutionAction.KEEP_EXISTING
    assert ResolutionAction.ADD in conflicts[
        ConflictType.TARGET_CELL_OCCUPIED
    ].allowed_actions
    resolutions = {
        conflicts[ConflictType.TARGET_CELL_OCCUPIED].conflict_id: {
            "action": "ADD"
        },
        conflicts[ConflictType.TARGET_CELL_FORMULA].conflict_id: {
            "action": "OVERWRITE"
        },
        conflicts[ConflictType.TARGET_CELL_TEXT].conflict_id: {
            "action": "SKIP"
        },
    }

    result = service.apply(plan, resolutions)

    assert result.posted_source_items == 4
    assert result.written_cells == 4
    assert result.already_existing_items == 1
    assert result.skipped_source_items == 1
    workbook = load_workbook(target, data_only=False)
    try:
        sheet = workbook["T07 26"]
        assert [sheet.cell(row, cb_column).value for row in range(2, 8)] == [
            100,
            200,
            300,
            150,
            400,
            "đã nhập tay",
        ]
    finally:
        workbook.close()


def test_posting_uses_primary_row_beside_ron_and_reports_unsafe_matches(
    tmp_path: Path,
) -> None:
    ready = tmp_path / "ready.json"
    target = tmp_path / "BK 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    primary_and_ron = "DRYU3026167"
    ambiguous = "MSCU1234567"
    missing = "OOLU7654321"
    _save_ready(
        ready,
        [
            [primary_and_ron, None, "NV", "ST", 100],
            [ambiguous, None, "NV", "ST", 200],
            [missing, None, "NV", "ST", 300],
            [None, "BL-ONLY", "CB", "HD", 400],
        ],
    )
    workbook = Workbook()
    sheet = _new_posting_sheet(workbook, "T07 26")
    _add_posting_row(sheet, 2, primary_and_ron, sqt=700)
    _add_posting_row(
        sheet,
        3,
        primary_and_ron,
        sqt=None,
        closing_date=None,
        cargo_type="Ron",
    )
    _add_posting_row(sheet, 4, ambiguous, sqt=701)
    _add_posting_row(sheet, 5, ambiguous, sqt=702)
    workbook.save(target)
    workbook.close()

    plan = _posting_service(ready, target, runtime_dir).analyze(
        sheet_name="T07 26"
    )
    by_container = {item.container: item for item in plan.items if item.container}

    assert by_container[primary_and_ron].target_row is None
    assert by_container[ambiguous].target_row is None
    assert by_container[missing].target_row is None
    conflict_types = {
        conflict.conflict_type for conflict in plan.conflicts
    }
    multiple = [
        conflict
        for conflict in plan.conflicts
        if conflict.conflict_type is ConflictType.MULTIPLE_CONTAINER_MATCH
    ]
    assert len(multiple) == 2
    assert ConflictType.CONTAINER_NOT_FOUND in conflict_types
    assert ConflictType.BL_ONLY_NO_CONTAINER in conflict_types


def test_posting_unknown_fee_and_invoice_header_are_never_auto_mapped(
    tmp_path: Path,
) -> None:
    ready = tmp_path / "ready.json"
    target = tmp_path / "BK 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    _save_ready(
        ready,
        [
            ["DRYU3026167", None, "CXD", "ST", 100],
            ["MSCU1234567", None, "CB", "ST", 200],
        ],
    )
    workbook = Workbook()
    sheet = _new_posting_sheet(workbook, "T07 26")
    _add_posting_row(sheet, 2, "DRYU3026167")
    _add_posting_row(sheet, 3, "MSCU1234567")
    sheet.cell(1, POSTING_FEE_COLUMNS["CB"]).value = "Hóa đơn"
    workbook.save(target)
    workbook.close()

    plan = _posting_service(ready, target, runtime_dir).analyze(
        sheet_name="T07 26"
    )
    conflicts = {
        conflict.conflict_type: conflict for conflict in plan.conflicts
    }

    assert ConflictType.UNKNOWN_FEE_CODE in conflicts
    assert conflicts[ConflictType.UNKNOWN_FEE_CODE].fee == "CXD"
    assert set(conflicts[ConflictType.UNKNOWN_FEE_CODE].details["fees"]) == set(
        FEE_HEADER_ALIASES
    )
    assert ConflictType.FEE_COLUMN_MISSING in conflicts
    assert conflicts[ConflictType.FEE_COLUMN_MISSING].fee == "CB"
    assert conflicts[ConflictType.FEE_COLUMN_MISSING].target_column is None


def test_posting_rejects_invoice_unit_price_alias_after_notes(
    tmp_path: Path,
) -> None:
    ready = tmp_path / "ready.json"
    target = tmp_path / "BK 2026.xlsx"
    _save_ready(ready, [["DRYU3026167", None, "CBDH", "ST", 200]])
    workbook = Workbook()
    sheet = _new_posting_sheet(workbook, "T07 26")
    _add_posting_row(sheet, 2, "DRYU3026167")
    sheet.cell(1, POSTING_FEE_COLUMNS["CBDH"]).value = None
    notes_column = len(POSTING_BASE_HEADERS) + len(POSTING_FEE_COLUMNS) + 1
    sheet.cell(1, notes_column + 1).value = "ĐƠN GIÁ"
    workbook.save(target)
    workbook.close()

    plan = _posting_service(ready, target, tmp_path / "Excel").analyze(
        sheet_name="T07 26"
    )

    conflict = next(
        item
        for item in plan.conflicts
        if item.conflict_type is ConflictType.FEE_COLUMN_MISSING
    )
    assert conflict.fee == "CBDH"
    assert conflict.target_column is None


def test_posting_partial_replay_sums_only_unposted_source_rows(
    tmp_path: Path,
) -> None:
    ready = tmp_path / "ready.json"
    target = tmp_path / "BK 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    _save_ready(
        ready,
        [
            ["DRYU3026167", None, "VTN", "ST", 100],
            ["DRYU3026167", None, "VTN", "ST", 250],
        ],
    )
    workbook = Workbook()
    sheet = _new_posting_sheet(workbook, "T07 26")
    _add_posting_row(sheet, 2, "DRYU3026167")
    workbook.save(target)
    workbook.close()
    repository = _PostedIndexRepository({0})

    service = _posting_service(
        ready,
        target,
        runtime_dir,
        posting_repository=repository,
    )
    first_plan = service.analyze(sheet_name="T07 26")

    assert first_plan.already_posted_indices == {0}
    assert len(first_plan.previously_posted_items) == 1
    assert first_plan.requires_user_input
    plan = service.analyze(
        sheet_name="T07 26",
        repost_source_indices=[],
    )
    assert len(plan.items) == 1
    assert plan.items[0].source_indices == [1]
    assert plan.items[0].amount == 250

    result = service.apply(plan, {})

    assert result.posted_source_items == 1
    assert result.written_cells == 1
    workbook = load_workbook(target)
    try:
        assert workbook["T07 26"].cell(
            2, POSTING_FEE_COLUMNS["VTN"]
        ).value == 250
    finally:
        workbook.close()


def test_posting_aborts_without_backup_when_target_changes_after_analyze(
    tmp_path: Path,
) -> None:
    ready = tmp_path / "ready.json"
    target = tmp_path / "BK 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    _save_ready(ready, [["DRYU3026167", None, "CB", "ST", 100]])
    workbook = Workbook()
    sheet = _new_posting_sheet(workbook, "T07 26")
    _add_posting_row(sheet, 2, "DRYU3026167")
    workbook.save(target)
    workbook.close()
    service = _posting_service(ready, target, runtime_dir)
    plan = service.analyze(sheet_name="T07 26")

    workbook = load_workbook(target)
    workbook["T07 26"]["A30"] = "external change"
    workbook.save(target)
    workbook.close()
    changed_hash = _sha256(target)

    with pytest.raises(WorkbookChangedError):
        service.apply(plan, {})

    assert _sha256(target) == changed_hash
    assert not (runtime_dir / "Backup").exists()


def test_refine_cxd_surfaces_target_cell_conflict_before_apply(
    tmp_path: Path,
) -> None:
    ready = tmp_path / "ready.json"
    target = tmp_path / "BK 2026.xlsx"
    _save_ready(ready, [["DRYU3026167", None, "CXD", "ST", 100]])
    workbook = Workbook()
    sheet = _new_posting_sheet(workbook, "T07 26")
    _add_posting_row(sheet, 2, "DRYU3026167")
    sheet.cell(2, POSTING_FEE_COLUMNS["VTN"]).value = 50
    workbook.save(target)
    workbook.close()

    service = _posting_service(ready, target, tmp_path / "Excel")
    plan = service.analyze(sheet_name="T07 26")
    conflict = next(
        item
        for item in plan.conflicts
        if item.conflict_type is ConflictType.UNKNOWN_FEE_CODE
    )

    refined = service.refine(
        plan,
        {
            conflict.conflict_id: {
                "action": "SELECT_FEE",
                "selected_fee": "VTN",
            }
        },
    )

    assert refined.items[0].selected_fee == "VTN"
    assert [
        item.conflict_type for item in refined.conflicts
    ] == [ConflictType.TARGET_CELL_OCCUPIED]


def test_refine_manual_row_surfaces_formula_conflict(
    tmp_path: Path,
) -> None:
    ready = tmp_path / "ready.json"
    target = tmp_path / "BK 2026.xlsx"
    _save_ready(ready, [[None, "BL-01", "CB", "HD", 100]])
    workbook = Workbook()
    sheet = _new_posting_sheet(workbook, "T07 26")
    _add_posting_row(sheet, 2, "DRYU3026167")
    sheet.cell(2, POSTING_FEE_COLUMNS["CB"]).value = "=10+20"
    workbook.save(target)
    workbook.close()

    service = _posting_service(ready, target, tmp_path / "Excel")
    plan = service.analyze(sheet_name="T07 26")
    conflict = next(
        item
        for item in plan.conflicts
        if item.conflict_type is ConflictType.BL_ONLY_NO_CONTAINER
    )
    assert [candidate.row for candidate in conflict.row_candidates] == [2]

    refined = service.refine(
        plan,
        {
            conflict.conflict_id: {
                "action": "SELECT_ROW",
                "selected_row": 2,
            }
        },
    )

    assert [
        item.conflict_type for item in refined.conflicts
    ] == [ConflictType.TARGET_CELL_FORMULA]


def test_apply_rejects_direct_fee_selection_until_plan_is_refined(
    tmp_path: Path,
) -> None:
    ready = tmp_path / "ready.json"
    target = tmp_path / "BK 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    _save_ready(ready, [["DRYU3026167", None, "CXD", "ST", 100]])
    workbook = Workbook()
    sheet = _new_posting_sheet(workbook, "T07 26")
    _add_posting_row(sheet, 2, "DRYU3026167")
    sheet.cell(2, POSTING_FEE_COLUMNS["VTN"]).value = 50
    workbook.save(target)
    workbook.close()

    service = _posting_service(ready, target, runtime_dir)
    plan = service.analyze(sheet_name="T07 26")
    conflict = next(
        item
        for item in plan.conflicts
        if item.conflict_type is ConflictType.UNKNOWN_FEE_CODE
    )
    before = _sha256(target)

    with pytest.raises(ExpensePostingError, match="refine"):
        service.apply(
            plan,
            {
                conflict.conflict_id: {
                    "action": "SELECT_FEE",
                    "selected_fee": "VTN",
                }
            },
        )

    assert _sha256(target) == before
    assert not (runtime_dir / "Backup").exists()


@pytest.mark.parametrize(
    ("existing", "action", "conflict_type"),
    [
        (50, ResolutionAction.KEEP_EXISTING, ConflictType.TARGET_CELL_OCCUPIED),
        (
            "=10+20",
            ResolutionAction.KEEP_FORMULA,
            ConflictType.TARGET_CELL_FORMULA,
        ),
    ],
)
def test_posting_history_preserves_keep_action_and_cell_value(
    tmp_path: Path,
    existing: Any,
    action: ResolutionAction,
    conflict_type: ConflictType,
) -> None:
    ready = tmp_path / "ready.json"
    target = tmp_path / "BK 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    _save_ready(ready, [["DRYU3026167", None, "CB", "ST", 100]])
    workbook = Workbook()
    sheet = _new_posting_sheet(workbook, "T07 26")
    _add_posting_row(sheet, 2, "DRYU3026167")
    sheet.cell(2, POSTING_FEE_COLUMNS["CB"]).value = existing
    workbook.save(target)
    workbook.close()
    runs = _RunHistoryRepository()
    postings = _PostingHistoryRepository()
    service = _posting_service(
        ready,
        target,
        runtime_dir,
        posting_repository=postings,
        run_repository=runs,
    )
    plan = service.analyze(batch_id=71, sheet_name="T07 26")
    conflict = next(
        item for item in plan.conflicts if item.conflict_type is conflict_type
    )

    result = service.apply(
        plan,
        {conflict.conflict_id: {"action": action.value}},
    )

    assert result.status is ExcelRunStatus.NO_CHANGES
    assert len(postings.items) == 1
    assert postings.items[0]["action"] is action
    assert postings.items[0]["value_before"] == existing
    assert postings.items[0]["value_after"] == existing
    assert postings.metadata["batch_id"] == 71
    assert not (runtime_dir / "Backup").exists()


def test_daily_sync_ignores_filename_year_and_requires_target_sheet_when_ambiguous(
    tmp_path: Path,
) -> None:
    daily = tmp_path / "bao cao noi bo 2025-2027.xlsx"
    target = tmp_path / "so cai tong hop.xlsx"
    _save_daily(daily, [_sync_row(700)])
    workbook = Workbook()
    first = workbook.active
    first.title = "T07 25"
    _populate_target_sheet(first, [_sync_row(700, "OLDU0000001")])
    second = workbook.create_sheet("T07 26")
    _populate_target_sheet(second, [_sync_row(699, "OLDU0000002")])
    workbook.save(target)
    workbook.close()

    service = _sync_service(daily, target, tmp_path / "Excel")
    plan = service.analyze(source_sheet_name="Tháng 7")

    assert plan.source_year is None
    assert plan.selected_sheet is None
    assert {item.target_sheet for item in plan.month_candidates} == {
        "T07 25",
        "T07 26",
    }
    assert {
        item.target_sheet: item.new_row_count
        for item in plan.month_candidates
    } == {"T07 25": 0, "T07 26": 1}
    conflict = next(
        item
        for item in plan.conflicts
        if item.conflict_type is ConflictType.TARGET_MONTH_AMBIGUOUS
    )
    result = service.apply(
        plan,
        {
            conflict.conflict_id: {
                "action": "SELECT_MONTH",
                "selected_month": 7,
                "selected_sheet": "T07 26",
            }
        },
    )

    assert result.sheet_name == "T07 26"
    workbook = load_workbook(target)
    try:
        assert workbook["T07 25"]["A3"].value is None
        assert workbook["T07 26"]["A3"].value == 700
    finally:
        workbook.close()


def test_daily_sync_uses_the_only_existing_sheet_for_source_month(
    tmp_path: Path,
) -> None:
    daily = tmp_path / "du lieu tuy chon.xlsx"
    target = tmp_path / "bk nhieu nam.xlsx"
    _save_daily(daily, [_sync_row(700)])
    workbook = Workbook()
    july = workbook.active
    july.title = "T07 25"
    _populate_target_sheet(july, [_sync_row(699, "OLDU0000001")])
    june = workbook.create_sheet("T06 26")
    _populate_target_sheet(june, [_sync_row(699, "OLDU0000002")])
    workbook.save(target)
    workbook.close()

    service = _sync_service(daily, target, tmp_path / "Excel")
    plan = service.analyze()

    assert [item.target_sheet for item in plan.month_candidates] == ["T07 25"]
    assert plan.selected_sheet == "T07 25"
    assert not any(
        item.conflict_type is ConflictType.TARGET_MONTH_AMBIGUOUS
        for item in plan.conflicts
    )

    result = service.apply(plan, {})

    assert result.sheet_name == "T07 25"
    workbook = load_workbook(target)
    try:
        assert workbook["T07 25"]["A3"].value == 700
        assert "T07 26" not in workbook.sheetnames
    finally:
        workbook.close()


@pytest.mark.parametrize("existing_header", [None, "Data cập nhật", "Date cập nhật"])
def test_posting_writes_or_reuses_update_date_column(
    tmp_path: Path,
    existing_header: str | None,
) -> None:
    ready = tmp_path / "ready.json"
    target = tmp_path / "BK tuy chon.xlsx"
    fixed = datetime(2026, 7, 29, 15, 4, 5)
    _save_ready(ready, [["DRYU3026167", None, "NV", "ST", 100]])
    workbook = Workbook()
    sheet = _new_posting_sheet(workbook, "T07 26")
    _add_posting_row(sheet, 2, "DRYU3026167")
    last_header = len(POSTING_BASE_HEADERS) + len(POSTING_FEE_COLUMNS) + 2
    sheet.cell(1, last_header).fill = PatternFill("solid", fgColor="FFF2CC")
    update_column = last_header + 1
    if existing_header is not None:
        sheet.cell(1, update_column).value = existing_header
    workbook.save(target)
    workbook.close()

    service = _posting_service(
        ready,
        target,
        tmp_path / "Excel",
        clock=lambda: fixed,
    )
    result = service.apply(
        service.analyze(sheet_name="T07 26"),
        {},
    )

    assert result.written_cells == 1
    workbook = load_workbook(target)
    try:
        sheet = workbook["T07 26"]
        assert sheet.cell(1, update_column).value == (
            existing_header or "Date cập nhật"
        )
        assert sheet.cell(2, update_column).value == fixed
        assert sheet.cell(2, update_column).number_format == "dd/mm/yyyy hh:mm:ss"
        if existing_header is None:
            assert (
                sheet.cell(1, update_column).fill.fgColor.rgb
                == sheet.cell(1, last_header).fill.fgColor.rgb
            )
    finally:
        workbook.close()


def test_posting_same_value_only_updates_date_when_explicitly_reposted(
    tmp_path: Path,
) -> None:
    ready = tmp_path / "ready.json"
    target = tmp_path / "BK.xlsx"
    fixed = datetime(2026, 7, 29, 16, 30, 0)
    _save_ready(ready, [["DRYU3026167", None, "NV", "ST", 100]])
    workbook = Workbook()
    sheet = _new_posting_sheet(workbook, "T07 26")
    _add_posting_row(sheet, 2, "DRYU3026167")
    sheet.cell(2, POSTING_FEE_COLUMNS["NV"]).value = 100
    workbook.save(target)
    workbook.close()

    normal = _posting_service(
        ready,
        target,
        tmp_path / "Normal",
        clock=lambda: fixed,
    )
    normal_result = normal.apply(normal.analyze(sheet_name="T07 26"), {})
    assert normal_result.written_cells == 0
    assert normal_result.backup_path is None
    workbook = load_workbook(target)
    try:
        headers = [
            workbook["T07 26"].cell(1, column).value
            for column in range(1, workbook["T07 26"].max_column + 1)
        ]
        assert "Date cập nhật" not in headers
    finally:
        workbook.close()

    repost = _posting_service(
        ready,
        target,
        tmp_path / "Repost",
        posting_repository=_PostedIndexRepository({0}),
        clock=lambda: fixed,
    )
    repost_plan = repost.analyze(
        sheet_name="T07 26",
        repost_source_indices=[0],
    )
    repost_result = repost.apply(repost_plan, {})

    assert repost_plan.items[0].force_repost
    assert repost_result.written_cells == 1
    workbook = load_workbook(target)
    try:
        sheet = workbook["T07 26"]
        date_column = next(
            column
            for column in range(1, sheet.max_column + 1)
            if sheet.cell(1, column).value == "Date cập nhật"
        )
        assert sheet.cell(2, date_column).value == fixed
    finally:
        workbook.close()


def test_posting_sheet_cancel_is_audited_as_cancelled(
    tmp_path: Path,
) -> None:
    ready = tmp_path / "ready.json"
    target = tmp_path / "BK 2026.xlsx"
    _save_ready(ready, [[None, "BL-ONLY", "CB", "ST", 100]])
    workbook = Workbook()
    _new_posting_sheet(workbook, "T07 26")
    _new_posting_sheet(workbook, "T08 26")
    workbook.save(target)
    workbook.close()
    runs = _RunHistoryRepository()
    service = _posting_service(
        ready,
        target,
        tmp_path / "Excel",
        run_repository=runs,
    )
    plan = service.analyze()
    assert plan.selected_sheet is None
    assert not plan.conflicts
    service.cancel(plan)

    assert runs.finished[-1]["status"] is ExcelRunStatus.CANCELLED


def test_daily_month_cancel_is_audited_as_cancelled(
    tmp_path: Path,
) -> None:
    daily = tmp_path / "Hàng ngày 2026.xlsx"
    target = tmp_path / "BK 2026.xlsx"
    workbook = Workbook()
    for month, container in ((7, "DRYU3026167"), (8, "MSCU1234567")):
        sheet = workbook.active if month == 7 else workbook.create_sheet()
        sheet.title = f"Tháng {month}"
        sheet.append(SYNC_HEADERS)
        sheet.append(_sync_row(700, container))
    workbook.save(daily)
    workbook.close()
    workbook = Workbook()
    for month, container in ((7, "DRYU3026167"), (8, "MSCU1234567")):
        sheet = workbook.active if month == 7 else workbook.create_sheet()
        sheet.title = f"T{month:02d} 26"
        _populate_target_sheet(sheet, [_sync_row(699, container)])
    workbook.save(target)
    workbook.close()
    runs = _RunHistoryRepository()
    service = _sync_service(
        daily,
        target,
        tmp_path / "Excel",
        run_repository=runs,
    )
    plan = service.analyze()
    conflict = next(
        item
        for item in plan.conflicts
        if item.conflict_type is ConflictType.TARGET_MONTH_AMBIGUOUS
    )

    with pytest.raises(DailySyncError, match="Người dùng đã hủy"):
        service.apply(
            plan,
            {conflict.conflict_id: {"action": "CANCEL_ALL"}},
        )

    assert runs.finished[-1]["status"] is ExcelRunStatus.CANCELLED
