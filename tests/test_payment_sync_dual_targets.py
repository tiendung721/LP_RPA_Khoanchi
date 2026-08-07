from __future__ import annotations

import hashlib
import shutil
import zipfile
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import app.services.excel.payment_sync as payment_module
from app.services.excel.models import ConflictType, ResolutionAction
from app.services.excel.payment_sync import (
    DATE_NUMBER_FORMAT,
    SUMMARY_HEADERS,
    PaymentSyncError,
    PaymentSyncService,
)
from app.services.excel.resolvers import MonthSheetService


class _ImmediateStabilityChecker:
    def wait(self, _path: str | Path) -> None:
        return None


SOURCE_COLUMNS = {
    "sea_freight": 3,
    "north_freight": 4,
    "empty_lift": 5,
    "loaded_drop": 6,
    "loaded_lift": 7,
    "empty_drop": 8,
    "south_freight": 9,
    "storage": 10,
    "overweight": 11,
    "vs_do": 12,
    "command_fee": 14,
    "repair": 15,
}


def _save_bk(path: Path, rows: list[dict[str, object]], *, month: int = 6) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"T{month:02d} 26"
    headers = [
        "SQT", "Số Container", "Cước biển", "Cước bộ đóng hàng",
        "Nâng vỏ", "Hạ Hàng", "Nâng Hàng", "Hạ vỏ", "Cước VTN",
        "Lưu cont", "Quá tải", "VS + D/O", "Hóa đơn VS",
        "LÀM LỆNH", "SỬA CHỮA", *SUMMARY_HEADERS,
    ]
    for column, header in enumerate(headers, 1):
        sheet.cell(1, column).value = header
    for row_number, values in enumerate(rows, 2):
        sheet.cell(row_number, 1).value = values.get("sqt", row_number + 500)
        sheet.cell(row_number, 2).value = values.get("container", f"CONT{row_number:07d}")
        for field, column in SOURCE_COLUMNS.items():
            sheet.cell(row_number, column).value = values.get(field)
    workbook.save(path)
    workbook.close()


def _add_payment_sheet(
    workbook: Workbook,
    name: str,
    target_type: str,
    *,
    rows: list[dict[str, object]] | None = None,
    total_row: int = 12,
) -> None:
    sheet = workbook.create_sheet(name)
    headers = (
        ("QT", "SỐ CONT", "HẠ HÀNG", "Số HD", "Ghi chú", "Date cập nhật")
        if target_type == "HP"
        else (
            "QT", "SỐ CONT", "NÂNG VỎ", "NÂNG HÀNG", "HẠ VỎ", "VS + D/O",
            "LÀM LỆNH", "Lưu Cont", "Sửa chữa Cont", "QUÁ TẢI",
            "Date cập nhật", "Số HD", "Ghi chú",
        )
    )
    sheet["A5"] = f"BẢNG KÊ THÁNG {name[1:6]}/26"
    for column, header in enumerate(headers, 1):
        sheet.cell(7, column).value = header
    money_fields = (
        {"loaded_drop": 3}
        if target_type == "HP"
        else {
            "empty_lift": 3, "loaded_lift": 4, "empty_drop": 5, "vs_do": 6,
            "command_fee": 7, "storage": 8, "repair": 9, "overweight": 10,
        }
    )
    for row_number, values in enumerate(rows or [], 8):
        sheet.cell(row_number, 1).value = values.get("sqt")
        sheet.cell(row_number, 2).value = values.get("container")
        for field, column in money_fields.items():
            sheet.cell(row_number, column).value = values.get(field)
        date_column = 6 if target_type == "HP" else 11
        sheet.cell(row_number, date_column).value = values.get("date")
        invoice_column = 4 if target_type == "HP" else 12
        sheet.cell(row_number, invoice_column).value = values.get("invoice")
        sheet.cell(row_number, invoice_column + 1).value = values.get("note")
    sheet.cell(total_row, 1).value = "TỔNG TIỀN"
    for column in money_fields.values():
        letter = sheet.cell(1, column).column_letter
        sheet.cell(total_row, column).value = f"=SUM({letter}8:{letter}{total_row - 1})"
    sheet.cell(total_row + 2, 1).value = "Người duyệt"
    sheet.cell(total_row + 5, 1).value = "Số tiền thanh toán"
    sheet.cell(total_row + 6, 4).value = "Giữ nguyên khu vực thanh toán"
    sheet.freeze_panes = "A8"


def _save_payment(
    path: Path,
    *,
    hp_rows: list[dict[str, object]] | None = None,
    nam_rows: list[dict[str, object]] | None = None,
    month: int = 6,
    total_row: int = 12,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_payment_sheet(
        workbook, f"T{month:02d} 26 HP", "HP", rows=hp_rows, total_row=total_row
    )
    _add_payment_sheet(
        workbook, f"T{month:02d} 26 NAM", "NAM", rows=nam_rows, total_row=total_row
    )
    workbook.save(path)
    workbook.close()


def _save_invoice_bk(path: Path, rows: list[dict[str, object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "T06 26"
    headers = [
        "SQT", "Số Container", "Cước biển", "Cước bộ đóng hàng",
        "Nâng vỏ", "Số HĐ", "Hạ Hàng", "Cột phụ", "Hóa đơn",
        "Cước VTN", "Hóa đơn VTN", "Nâng Hàng", "số hd", "Hạ vỏ",
        "Hoá đơn", "VS + D/O", "HD", "LÀM LỆNH", "Lưu cont",
        "Số hóa đơn", "Quá tải", "HD", "SỬA CHỮA", "Hoá đơn",
        *SUMMARY_HEADERS,
    ]
    for column, header in enumerate(headers, 1):
        sheet.cell(1, column).value = header
    amount_columns = {
        "empty_lift": 5,
        "loaded_drop": 7,
        "loaded_lift": 12,
        "empty_drop": 14,
        "vs_do": 16,
        "command_fee": 18,
        "storage": 19,
        "overweight": 21,
        "repair": 23,
    }
    invoice_columns = {
        "empty_lift": 6,
        "loaded_drop": 9,
        "loaded_lift": 13,
        "empty_drop": 15,
        "vs_do": 17,
        "storage": 20,
        "overweight": 22,
        "repair": 24,
    }
    for row_number, values in enumerate(rows, 2):
        sheet.cell(row_number, 1).value = values["sqt"]
        sheet.cell(row_number, 2).value = values["container"]
        for field, column in amount_columns.items():
            sheet.cell(row_number, column).value = values.get(field)
        for field, column in invoice_columns.items():
            sheet.cell(row_number, column).value = values.get(f"invoice_{field}")
    workbook.save(path)
    workbook.close()


def _save_invoice_payment(
    path: Path,
    *,
    hp_invoice: object | None = None,
    nam_invoice: object | None = None,
) -> None:
    workbook = Workbook()
    hp = workbook.active
    hp.title = "T06 26 HP"
    hp_headers = ("QT", "SỐ CONT", "HẠ HÀNG", "Cột phụ", "Số HD", "Ghi chú", "Date cập nhật")
    for column, header in enumerate(hp_headers, 1):
        hp.cell(7, column).value = header
    hp.append([])
    hp["A8"], hp["B8"], hp["C8"], hp["E8"] = 700, "CONT700", 20, hp_invoice
    hp["C12"] = "=SUM(C8:C11)"

    nam = workbook.create_sheet("T06 26 NAM")
    nam_headers = (
        "QT", "SỐ CONT", "NÂNG VỎ", "Số HD", "NÂNG HÀNG", "Số HD",
        "HẠ VỎ", "Số HD", "VS + D/O", "Số HD", "LÀM LỆNH",
        "Lưu Cont", "Số HD", "Sửa chữa Cont", "Số HD", "QUÁ TẢI",
        "Số HD", "QT", "Tổng", "Date cập nhật",
    )
    for column, header in enumerate(nam_headers, 1):
        nam.cell(7, column).value = header
    nam["A8"], nam["B8"] = 700, "CONT700"
    for coordinate, value in {
        "C8": 10, "E8": 30, "G8": 40, "I8": 50, "K8": 60,
        "L8": 70, "N8": 80, "P8": 90,
    }.items():
        nam[coordinate] = value
    nam["D8"] = nam_invoice
    for column in (3, 5, 7, 9, 11, 12, 14, 16):
        letter = nam.cell(1, column).column_letter
        nam.cell(12, column).value = f"=SUM({letter}8:{letter}11)"
    workbook.save(path)
    workbook.close()


def _service(
    bk: Path,
    payment: Path,
    runtime: Path,
    *,
    clock=lambda: datetime(2026, 8, 3, 14, 15, 16),
) -> PaymentSyncService:
    return PaymentSyncService(
        bk_path=bk,
        payment_path=payment,
        temp_dir=runtime / "Temp",
        backup_dir=runtime / "Backup",
        stability_checker=_ImmediateStabilityChecker(),
        clock=clock,
    )


def _sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_payment_sheet_resolver_is_exact_case_insensitive_and_type_safe() -> None:
    service = MonthSheetService()
    parsed = service.parse_payment_sheet("  t06 26 hp  ")
    assert parsed is not None
    assert (parsed.month, parsed.year, parsed.sheet_type) == (6, 2026, "HP")
    for invalid in ("T06 26", "LUU TRU T06 26 NAM", "T06 26 NAM (2)", "COPY T06 26 HP"):
        assert service.parse_payment_sheet(invalid) is None
    names = ["T04 26", "T04 26 HP", "T05 26 NAM", "T03 26 NAM"]
    assert service.nearest_previous_payment_template(names, 6, 2026, "HP") == "T04 26 HP"
    assert service.nearest_previous_payment_template(names, 6, 2026, "NAM") == "T05 26 NAM"


@pytest.mark.parametrize(
    ("field", "expected_hp", "expected_nam"),
    [
        ("loaded_drop", 1, 0),
        ("empty_lift", 0, 1),
        ("loaded_lift", 0, 1),
        ("empty_drop", 0, 1),
        ("vs_do", 0, 1),
        ("command_fee", 0, 1),
        ("storage", 0, 1),
        ("repair", 0, 1),
        ("overweight", 0, 1),
        ("sea_freight", 0, 0),
        ("north_freight", 0, 0),
        ("south_freight", 0, 0),
    ],
)
def test_fee_classification_routes_only_in_scope_fields(
    tmp_path: Path, field: str, expected_hp: int, expected_nam: int
) -> None:
    bk, payment = tmp_path / "bk.xlsx", tmp_path / "payment.xlsx"
    _save_bk(bk, [{"sqt": 598, "container": " vsgu 2250713 ", field: 123_000}])
    _save_payment(payment)

    plan = _service(bk, payment, tmp_path / "runtime").analyze(
        source_sheet_name="T06 26"
    )

    assert plan.targets["HP"].new_count == expected_hp
    assert plan.targets["NAM"].new_count == expected_nam
    assert all(item.container == "VSGU2250713" for item in plan.items)


def test_one_source_row_can_update_both_targets_without_touching_manual_columns(
    tmp_path: Path,
) -> None:
    bk, payment = tmp_path / "bk.xlsx", tmp_path / "payment.xlsx"
    _save_bk(
        bk,
        [{"sqt": 598, "container": "VSGU2250713", "loaded_drop": 1_252_800, "loaded_lift": 463_300}],
    )
    old_date = datetime(2026, 7, 1, 8, 0, 0)
    manual = {"sqt": 598, "container": "VSGU2250713", "invoice": "HD-88", "note": "Không sửa", "date": old_date}
    _save_payment(payment, hp_rows=[manual | {"loaded_drop": 1}], nam_rows=[manual | {"loaded_lift": 2}])
    service = _service(bk, payment, tmp_path / "runtime")

    result = service.apply(service.analyze(source_sheet_name="T06 26"), {})

    assert result.updated_rows == 2
    workbook = load_workbook(payment)
    try:
        hp, nam = workbook["T06 26 HP"], workbook["T06 26 NAM"]
        assert (hp["D8"].value, hp["E8"].value) == ("HD-88", "Không sửa")
        assert (nam["L8"].value, nam["M8"].value) == ("HD-88", "Không sửa")
        assert hp["F8"].value == datetime(2026, 8, 3, 14, 15, 16)
        assert nam["K8"].value == datetime(2026, 8, 3, 14, 15, 16)
        assert hp["F8"].number_format == DATE_NUMBER_FORMAT
    finally:
        workbook.close()


def test_unchanged_row_keeps_existing_update_date(tmp_path: Path) -> None:
    bk, payment = tmp_path / "bk.xlsx", tmp_path / "payment.xlsx"
    old_date = datetime(2026, 7, 1, 8, 0, 0)
    _save_bk(bk, [{"sqt": 599, "container": "CONT599", "loaded_drop": 123}])
    _save_payment(
        payment,
        hp_rows=[{"sqt": 599, "container": "CONT599", "loaded_drop": 123, "date": old_date}],
    )
    service = _service(bk, payment, tmp_path / "runtime")

    plan = service.analyze(source_sheet_name="T06 26")
    result = service.apply(plan, {})

    assert plan.targets["HP"].unchanged_count == 1
    assert result.target_results["HP"].unchanged_rows == 1
    workbook = load_workbook(payment)
    assert workbook["T06 26 HP"]["F8"].value == old_date
    workbook.close()


def test_blank_bk_money_defaults_to_keep_and_can_be_explicitly_cleared(tmp_path: Path) -> None:
    def arrange(folder: Path) -> tuple[Path, Path, PaymentSyncService, object]:
        folder.mkdir()
        bk, payment = folder / "bk.xlsx", folder / "payment.xlsx"
        _save_bk(bk, [{"sqt": 600, "container": "CONT600", "empty_lift": 100}])
        _save_payment(
            payment,
            nam_rows=[{"sqt": 600, "container": "CONT600", "empty_lift": 1, "loaded_lift": 999}],
        )
        service = _service(bk, payment, folder / "runtime")
        plan = service.analyze(source_sheet_name="T06 26")
        conflict = next(c for c in plan.conflicts if c.conflict_type is ConflictType.PAYMENT_CLEAR_VALUE)
        return bk, payment, service, (plan, conflict)

    _bk, payment, service, payload = arrange(tmp_path / "keep")
    plan, conflict = payload
    assert conflict.default_action is ResolutionAction.KEEP_EXISTING
    service.apply(plan, {})
    workbook = load_workbook(payment)
    assert workbook["T06 26 NAM"]["D8"].value == 999
    workbook.close()

    _bk, payment, service, payload = arrange(tmp_path / "clear")
    plan, conflict = payload
    service.apply(plan, {conflict.conflict_id: {"action": "OVERWRITE"}})
    workbook = load_workbook(payment)
    assert workbook["T06 26 NAM"]["D8"].value is None
    workbook.close()


@pytest.mark.parametrize(
    "target_rows",
    [
        [{"sqt": 700, "container": "OTHER", "loaded_drop": 1}],
        [{"sqt": 701, "container": "CONT700", "loaded_drop": 1}],
        [
            {"sqt": 700, "container": "CONT700", "loaded_drop": 1},
            {"sqt": 700, "container": "CONT700", "loaded_drop": 2},
        ],
    ],
)
def test_partial_or_duplicate_keys_require_row_selection(
    tmp_path: Path, target_rows: list[dict[str, object]]
) -> None:
    bk, payment = tmp_path / "bk.xlsx", tmp_path / "payment.xlsx"
    _save_bk(bk, [{"sqt": 700, "container": "CONT700", "loaded_drop": 123}])
    _save_payment(payment, hp_rows=target_rows)

    plan = _service(bk, payment, tmp_path / "runtime").analyze(source_sheet_name="T06 26")

    conflict = plan.targets["HP"].conflicts[0]
    assert conflict.conflict_type is ConflictType.PARTIAL_KEY_MATCH
    assert conflict.row_candidates


def test_inserting_before_total_extends_local_sum_but_not_external_reference(tmp_path: Path) -> None:
    bk, payment = tmp_path / "bk.xlsx", tmp_path / "payment.xlsx"
    _save_bk(bk, [{"sqt": 900, "container": "NEW900", "loaded_drop": 300}])
    hp_rows = [
        {"sqt": 1, "container": "A", "loaded_drop": 1},
        {"sqt": 2, "container": "B", "loaded_drop": 2},
    ]
    _save_payment(payment, hp_rows=hp_rows, total_row=10)
    workbook = load_workbook(payment)
    workbook["T06 26 HP"]["D12"] = "='T04 26'!S148"
    workbook.save(payment)
    workbook.close()
    service = _service(bk, payment, tmp_path / "runtime")

    service.apply(service.analyze(source_sheet_name="T06 26"), {})

    workbook = load_workbook(payment, data_only=False)
    sheet = workbook["T06 26 HP"]
    assert sheet["C11"].value == "=SUM(C8:C10)"
    assert sheet["D13"].value == "='T04 26'!S148"
    workbook.close()


def test_row_with_blank_identity_but_managed_money_is_not_reused(tmp_path: Path) -> None:
    bk, payment = tmp_path / "bk.xlsx", tmp_path / "payment.xlsx"
    _save_bk(bk, [{"sqt": 901, "container": "NEW901", "loaded_drop": 300}])
    _save_payment(
        payment,
        hp_rows=[{"sqt": None, "container": None, "loaded_drop": 999}],
    )
    service = _service(bk, payment, tmp_path / "runtime")

    service.apply(service.analyze(source_sheet_name="T06 26"), {})

    workbook = load_workbook(payment)
    sheet = workbook["T06 26 HP"]
    assert sheet["C8"].value == 999
    assert (sheet["A9"].value, sheet["B9"].value, sheet["C9"].value) == (
        901, "NEW901", 300,
    )
    workbook.close()


def test_nam_write_failure_never_persists_hp(tmp_path: Path, monkeypatch) -> None:
    bk, payment = tmp_path / "bk.xlsx", tmp_path / "payment.xlsx"
    _save_bk(bk, [{"sqt": 800, "container": "CONT800", "loaded_drop": 1, "empty_lift": 2}])
    _save_payment(payment)
    original = payment.read_bytes()
    service = _service(bk, payment, tmp_path / "runtime")
    plan = service.analyze(source_sheet_name="T06 26")
    original_writer = payment_module._write_profile_item

    def fail_nam(*args, **kwargs):
        if kwargs["profile"].target_type == "NAM":
            raise PaymentSyncError("NAM failed")
        return original_writer(*args, **kwargs)

    monkeypatch.setattr(payment_module, "_write_profile_item", fail_nam)
    with pytest.raises(PaymentSyncError, match="NAM failed"):
        service.apply(plan, {})
    assert payment.read_bytes() == original
    assert not list((tmp_path / "runtime" / "Backup").glob("*"))


def test_analyze_is_read_only_and_backup_directory_keeps_only_latest(tmp_path: Path) -> None:
    bk, payment = tmp_path / "bk.xlsx", tmp_path / "payment.xlsx"
    _save_bk(bk, [{"sqt": 801, "container": "CONT801", "loaded_drop": 1}])
    _save_payment(payment)
    service = _service(bk, payment, tmp_path / "runtime")
    before = (_sha(bk), _sha(payment))
    plan = service.analyze(source_sheet_name="T06 26")
    assert (_sha(bk), _sha(payment)) == before
    service.apply(plan, {})
    backups = list((tmp_path / "runtime" / "Backup").glob("*payment*"))
    assert [path.name for path in backups] == ["payment_latest.xlsx"]
    assert not any(char.isdigit() for char in backups[0].stem.removesuffix("_latest"))


def _add_dummy_vba(path: Path) -> None:
    with zipfile.ZipFile(path, "a") as package:
        package.writestr("xl/vbaProject.bin", b"dummy-vba-project")


def test_xlsm_working_copy_preserves_vba_package(tmp_path: Path) -> None:
    bk, payment = tmp_path / "bk.xlsx", tmp_path / "payment.xlsm"
    _save_bk(bk, [{"sqt": 802, "container": "CONT802", "loaded_drop": 1}])
    _save_payment(payment)
    _add_dummy_vba(payment)
    service = _service(bk, payment, tmp_path / "runtime")
    plan = service.analyze(source_sheet_name="T06 26")
    assert plan.target_vba_present

    result = service.apply(plan, {})

    assert result.vba_preserved
    with zipfile.ZipFile(payment) as package:
        assert "xl/vbaProject.bin" in package.namelist()


def test_real_workbooks_analyze_and_apply_only_on_temporary_copies(tmp_path: Path) -> None:
    project_root = Path(__file__).resolve().parents[1]
    source_fixture = project_root / "Output" / "BK_TONG_HOP_TT_NOI_DIA_2026.xlsx"
    target_fixture = project_root / "Output" / "THANH_TOAN_NANG_HA_VS_DO.xlsm"
    if not source_fixture.is_file() or not target_fixture.is_file():
        pytest.skip("Không có cặp workbook tích hợp thật trong workspace này.")
    original_hashes = (_sha(source_fixture), _sha(target_fixture))
    source = tmp_path / source_fixture.name
    target = tmp_path / target_fixture.name
    shutil.copy2(source_fixture, source)
    shutil.copy2(target_fixture, target)
    service = _service(source, target, tmp_path / "runtime")

    plan = service.analyze(source_sheet_name="T06 26")
    result = service.apply(plan, {})

    assert not plan.targets["HP"].sheet_to_create
    assert not plan.targets["NAM"].sheet_to_create
    assert plan.invoice_change_count == 117
    assert result.invoice_written_cells == 117
    written = load_workbook(target, read_only=False, data_only=False, keep_vba=True)
    try:
        assert written["T06 26 HP"]["D8"].value == "17510 - HA"
        assert written["T06 26 NAM"]["D8"].value == "102"
    finally:
        written.close()
    with zipfile.ZipFile(target) as package:
        assert any(name.lower().endswith("vbaproject.bin") for name in package.namelist())
    assert [path.name for path in (tmp_path / "runtime" / "Backup").glob("*")] == [
        "THANH_TOAN_NANG_HA_VS_DO_latest.xlsm"
    ]
    assert (_sha(source_fixture), _sha(target_fixture)) == original_hashes


def _invoice_source_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {
        "sqt": 700,
        "container": "CONT700",
        "empty_lift": 10,
        "loaded_drop": 20,
        "loaded_lift": 30,
        "empty_drop": 40,
        "vs_do": 50,
        "command_fee": 60,
        "storage": 70,
        "repair": 80,
        "overweight": 90,
        "invoice_empty_lift": "INV-NV",
        "invoice_loaded_drop": "INV-HH",
        "invoice_loaded_lift": "INV-NH",
        "invoice_empty_drop": "INV-HV",
        "invoice_vs_do": "INV-VS",
        "invoice_storage": "INV-LC",
        "invoice_repair": "INV-SC",
        "invoice_overweight": "INV-QT",
    }
    row.update(overrides)
    return row


def test_payment_sync_resolves_shifted_invoice_headers_and_writes_invoice_only(
    tmp_path: Path,
) -> None:
    bk, payment = tmp_path / "bk.xlsx", tmp_path / "payment.xlsx"
    _save_invoice_bk(bk, [_invoice_source_row()])
    _save_invoice_payment(payment)
    service = _service(bk, payment, tmp_path / "runtime")

    plan = service.analyze(source_sheet_name="T06 26")

    assert plan.conflict_count == 0
    assert plan.invoice_change_count == 8
    assert plan.update_count == 2
    result = service.apply(plan, {})
    assert result.invoice_written_cells == 8
    workbook = load_workbook(payment, data_only=False)
    try:
        assert workbook["T06 26 HP"]["E8"].value == "INV-HH"
        assert [
            workbook["T06 26 NAM"][coordinate].value
            for coordinate in ("D8", "F8", "H8", "J8", "M8", "O8", "Q8")
        ] == [
            "INV-NV", "INV-NH", "INV-HV", "INV-VS", "INV-LC", "INV-SC", "INV-QT"
        ]
        assert workbook["T06 26 HP"]["G8"].value is not None
        assert workbook["T06 26 NAM"]["T8"].value is not None
    finally:
        workbook.close()


def test_payment_invoice_conflict_defaults_to_keep_existing(tmp_path: Path) -> None:
    bk, payment = tmp_path / "bk.xlsx", tmp_path / "payment.xlsx"
    row = _invoice_source_row(**{
        key: None
        for key in _invoice_source_row()
        if key.startswith("invoice_")
    })
    row["invoice_loaded_drop"] = "00102"
    _save_invoice_bk(bk, [row])
    _save_invoice_payment(payment, hp_invoice=102)
    service = _service(bk, payment, tmp_path / "runtime")

    plan = service.analyze(source_sheet_name="T06 26")
    conflict = next(
        conflict
        for conflict in plan.conflicts
        if conflict.conflict_type is ConflictType.INVOICE_VALUE_CONFLICT
    )
    assert conflict.default_action is ResolutionAction.KEEP_EXISTING

    result = service.apply(plan, {})
    assert result.invoice_written_cells == 0
    workbook = load_workbook(payment, data_only=False)
    try:
        assert workbook["T06 26 HP"]["E8"].value == 102
    finally:
        workbook.close()


def test_payment_invoice_conflict_can_overwrite_without_skipping_amount(
    tmp_path: Path,
) -> None:
    bk, payment = tmp_path / "bk.xlsx", tmp_path / "payment.xlsx"
    row = _invoice_source_row(loaded_drop=25)
    _save_invoice_bk(bk, [row])
    _save_invoice_payment(payment, hp_invoice="OLD")
    service = _service(bk, payment, tmp_path / "runtime")
    plan = service.analyze(source_sheet_name="T06 26")
    conflict = next(
        conflict
        for conflict in plan.conflicts
        if conflict.conflict_type is ConflictType.INVOICE_VALUE_CONFLICT
        and conflict.details["field"] == "loaded_drop"
    )

    result = service.apply(
        plan,
        {conflict.conflict_id: {"action": "OVERWRITE"}},
    )

    assert result.invoice_written_cells >= 1
    workbook = load_workbook(payment, data_only=False)
    try:
        assert workbook["T06 26 HP"]["C8"].value == 25
        assert workbook["T06 26 HP"]["E8"].value == "INV-HH"
    finally:
        workbook.close()


def test_payment_multiple_source_invoices_are_refined_before_apply(
    tmp_path: Path,
) -> None:
    bk, payment = tmp_path / "bk.xlsx", tmp_path / "payment.xlsx"
    first = _invoice_source_row(invoice_loaded_drop="INV-A")
    second = _invoice_source_row(invoice_loaded_drop="INV-B")
    for field in ("empty_lift", "loaded_lift", "empty_drop", "vs_do", "command_fee", "storage", "repair", "overweight"):
        second[field] = None
        second[f"invoice_{field}"] = None
    _save_invoice_bk(bk, [first, second])
    _save_invoice_payment(payment)
    service = _service(bk, payment, tmp_path / "runtime")
    plan = service.analyze(source_sheet_name="T06 26")
    conflict = next(
        conflict
        for conflict in plan.conflicts
        if conflict.conflict_type is ConflictType.MULTIPLE_SOURCE_INVOICES
    )

    refined = service.refine(
        plan,
        {
            conflict.conflict_id: {
                "action": "SELECT_INVOICE",
                "selected_invoice": "INV-B",
            }
        },
    )

    assert not refined.conflicts
    result = service.apply(refined, {})
    assert result.invoice_written_cells == 8
    workbook = load_workbook(payment, data_only=False)
    try:
        assert workbook["T06 26 HP"]["E8"].value == "INV-B"
    finally:
        workbook.close()


@pytest.mark.parametrize("headers", [(None, "Ghi chú"), ("Số HD", "HD")])
def test_missing_or_ambiguous_payment_invoice_header_skips_only_invoice(
    tmp_path: Path,
    headers: tuple[str | None, str],
) -> None:
    bk, payment = tmp_path / "bk.xlsx", tmp_path / "payment.xlsx"
    row = _invoice_source_row(loaded_drop=25)
    for key in tuple(row):
        if key.startswith("invoice_") and key != "invoice_loaded_drop":
            row[key] = None
    _save_invoice_bk(bk, [row])
    _save_invoice_payment(payment)
    workbook = load_workbook(payment)
    workbook["T06 26 HP"]["E7"], workbook["T06 26 HP"]["F7"] = headers
    workbook.save(payment)
    workbook.close()
    service = _service(bk, payment, tmp_path / "runtime")

    plan = service.analyze(source_sheet_name="T06 26")
    conflict = next(
        conflict
        for conflict in plan.conflicts
        if conflict.conflict_type is ConflictType.INVOICE_COLUMN_MISSING
    )
    assert conflict.default_action is ResolutionAction.SKIP_INVOICE

    result = service.apply(plan, {})
    assert result.invoice_written_cells == 0
    workbook = load_workbook(payment, data_only=False)
    try:
        assert workbook["T06 26 HP"]["C8"].value == 25
    finally:
        workbook.close()


def test_blank_bk_invoice_never_clears_existing_payment_invoice(tmp_path: Path) -> None:
    bk, payment = tmp_path / "bk.xlsx", tmp_path / "payment.xlsx"
    row = _invoice_source_row()
    for key in tuple(row):
        if key.startswith("invoice_"):
            row[key] = None
    _save_invoice_bk(bk, [row])
    _save_invoice_payment(payment, hp_invoice="KEEP-ME", nam_invoice="KEEP-NAM")
    service = _service(bk, payment, tmp_path / "runtime")

    plan = service.analyze(source_sheet_name="T06 26")
    assert not any(
        conflict.conflict_type
        in {ConflictType.INVOICE_VALUE_CONFLICT, ConflictType.INVOICE_COLUMN_MISSING}
        for conflict in plan.conflicts
    )
    result = service.apply(plan, {})

    assert result.invoice_written_cells == 0
    workbook = load_workbook(payment, data_only=False)
    try:
        assert workbook["T06 26 HP"]["E8"].value == "KEEP-ME"
        assert workbook["T06 26 NAM"]["D8"].value == "KEEP-NAM"
    finally:
        workbook.close()
