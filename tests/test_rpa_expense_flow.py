from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from PySide6.QtCore import Qt

from app.config import AppSettings
from app.rpa_expense import (
    RPA_STATUS_IMPORTED,
    RPA_STATUS_NOT_IMPORTED,
    RpaExpenseService,
    RpaExpenseStatusService,
)
from app.rpa_expense.launcher import RpaExpenseBatLauncher
from app.rpa_expense.service import STATUS_HEADER, SUMMARY_HEADERS
from app.ui.rpa_expense_dialog import RpaSqtSelectionDialog


def _build_bk(path: Path, *, with_status: bool = True) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "T07 26"
    sheet["A1"] = "SQT"
    sheet["B1"] = "Ghi chú"
    summary_start = 3
    for offset, header in enumerate(SUMMARY_HEADERS):
        sheet.cell(1, summary_start + offset).value = header
    if with_status:
        sheet.cell(1, summary_start + len(SUMMARY_HEADERS)).value = STATUS_HEADER

    # Hai dòng cùng SQT 101 để kiểm tra tổng hợp nhiều dòng.
    rows = (
        (101, (100, 20, 30, 40, 50, 0, 5, 0, 60), None),
        (101, (200, 10, 0, 5, 25, 0, 15, 0, 40), RPA_STATUS_IMPORTED),
        (102, (300, 30, 10, 0, 20, 0, 0, 0, 50), RPA_STATUS_IMPORTED),
    )
    for row_number, (sqt, values, status) in enumerate(rows, 2):
        sheet.cell(row_number, 1).value = sqt
        sheet.cell(row_number, summary_start).value = f"=A{row_number}"
        for offset, value in enumerate(values, 1):
            sheet.cell(row_number, summary_start + offset).value = value
        if with_status:
            sheet.cell(
                row_number,
                summary_start + len(SUMMARY_HEADERS),
            ).value = status
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()


def _settings(tmp_path: Path, bk: Path) -> AppSettings:
    return AppSettings(
        data_root=tmp_path / "runtime",
        output_dir=tmp_path / "Output",
        bk_workbook_path=str(bk),
    )


def test_analyze_groups_rows_and_imported_items_remain_runnable(
    tmp_path: Path,
) -> None:
    bk = tmp_path / "Output" / "BK.xlsx"
    _build_bk(bk)
    service = RpaExpenseService(_settings(tmp_path, bk))

    assert [item.sheet_name for item in service.sheet_candidates()] == ["T07 26"]
    plan = service.analyze_sheet("T07 26")
    first, second = plan.items

    assert first.sqt == "101"
    assert first.source_rows == (2, 3)
    assert first.status == RPA_STATUS_NOT_IMPORTED
    assert first.amounts.cuoc_bo_dong_hang == 300
    assert first.amounts.nang_ha_dong_hang == 30
    assert first.amounts.luu_cont_qua_tai == 100
    assert first.amounts.sua_chua_cont == 20
    assert first.can_run

    assert second.status == RPA_STATUS_IMPORTED
    assert second.can_run


def test_prepare_json_and_mark_all_source_rows_after_success(
    tmp_path: Path,
) -> None:
    bk = tmp_path / "Output" / "BK.xlsx"
    _build_bk(bk)
    settings = _settings(tmp_path, bk)
    service = RpaExpenseService(settings)
    plan = service.analyze_sheet("T07 26")

    prepared = service.prepare_selection(plan, ["101", "102"])
    payload = json.loads(prepared.selection_path.read_text(encoding="utf-8"))

    assert payload["version"] == 1
    assert payload["operation"] == "NHAP_KHOAN_CHI_BK"
    assert payload["items"][0]["source_rows"] == [2, 3]
    assert payload["items"][0]["status_before"] == RPA_STATUS_NOT_IMPORTED
    assert payload["status_callback"]["when"] == "AFTER_WEB_SAVE_SUCCESS"
    assert payload["status_callback"]["arguments"][-1] == "{sqt}"

    status = RpaExpenseStatusService(
        backup_dir=settings.paths.excel_backup_dir
    )
    first_result = status.mark_imported(prepared.selection_path, "101")
    second_result = status.mark_imported(prepared.selection_path, "102")

    workbook = load_workbook(bk, data_only=False)
    try:
        sheet = workbook["T07 26"]
        status_column = 3 + len(SUMMARY_HEADERS)
        assert sheet.cell(2, status_column).value == RPA_STATUS_IMPORTED
        assert sheet.cell(3, status_column).value == RPA_STATUS_IMPORTED
        assert sheet.cell(4, status_column).value == RPA_STATUS_IMPORTED
    finally:
        workbook.close()
    assert first_result["backup_path"] == second_result["backup_path"]
    assert len(list(settings.paths.excel_backup_dir.glob("*.xlsx"))) == 1


def test_status_column_is_created_when_missing(tmp_path: Path) -> None:
    bk = tmp_path / "Output" / "BK.xlsx"
    _build_bk(bk, with_status=False)
    settings = _settings(tmp_path, bk)
    service = RpaExpenseService(settings)
    prepared = service.prepare_selection(
        service.analyze_sheet("T07 26"),
        ["101"],
    )

    RpaExpenseStatusService(
        backup_dir=settings.paths.excel_backup_dir
    ).mark_imported(prepared.selection_path, "101")

    workbook = load_workbook(bk, data_only=False)
    try:
        sheet = workbook["T07 26"]
        status_column = 3 + len(SUMMARY_HEADERS)
        assert sheet.cell(1, status_column).value == STATUS_HEADER
        assert sheet.cell(2, status_column).value == RPA_STATUS_IMPORTED
        assert sheet.cell(3, status_column).value == RPA_STATUS_IMPORTED
    finally:
        workbook.close()


def test_dialog_keeps_imported_sqt_selectable(qtbot, tmp_path: Path) -> None:
    bk = tmp_path / "Output" / "BK.xlsx"
    _build_bk(bk)
    plan = RpaExpenseService(_settings(tmp_path, bk)).analyze_sheet("T07 26")
    dialog = RpaSqtSelectionDialog(plan)
    qtbot.addWidget(dialog)

    imported_check = dialog.table.item(1, 0)
    assert imported_check.flags() & Qt.ItemFlag.ItemIsUserCheckable
    imported_check.setCheckState(Qt.CheckState.Checked)
    assert dialog.selected_sqt == ["102"]


def test_zero_amount_sqt_can_be_selected_singly_or_in_bulk(
    qtbot,
    tmp_path: Path,
) -> None:
    bk = tmp_path / "Output" / "BK.xlsx"
    _build_bk(bk)
    workbook = load_workbook(bk, data_only=False)
    try:
        sheet = workbook["T07 26"]
        for row_number in range(2, 5):
            for column in range(4, 13):
                sheet.cell(row_number, column).value = 0
        workbook.save(bk)
    finally:
        workbook.close()

    service = RpaExpenseService(_settings(tmp_path, bk))
    plan = service.analyze_sheet("T07 26")
    assert plan.runnable_count == 2
    assert all(item.amounts.total == 0 for item in plan.items)
    assert all(item.can_run for item in plan.items)
    assert all(
        item.validation_message == "Tất cả khoản tiền đều bằng 0."
        for item in plan.items
    )

    dialog = RpaSqtSelectionDialog(plan)
    qtbot.addWidget(dialog)
    first_check = dialog.table.item(0, 0)
    second_check = dialog.table.item(1, 0)
    assert first_check.flags() & Qt.ItemFlag.ItemIsUserCheckable
    assert second_check.flags() & Qt.ItemFlag.ItemIsUserCheckable

    first_check.setCheckState(Qt.CheckState.Checked)
    assert dialog.selected_sqt == ["101"]

    second_check.setCheckState(Qt.CheckState.Checked)
    assert dialog.selected_sqt == ["101", "102"]

    prepared = service.prepare_selection(plan, dialog.selected_sqt)
    assert prepared.item_count == 2
    assert [
        sum(item["amounts"].values()) for item in prepared.payload["items"]
    ] == [0, 0]


def test_launcher_passes_the_fixed_json_path_to_bat(
    monkeypatch,
    tmp_path: Path,
) -> None:
    bk = tmp_path / "Output" / "BK.xlsx"
    _build_bk(bk)
    bat = tmp_path / "run_pad.bat"
    bat.write_text("@echo off\r\n", encoding="utf-8")
    settings = _settings(tmp_path, bk)
    settings.rpa_expense_bat_path = str(bat)
    service = RpaExpenseService(settings)
    prepared = service.prepare_selection(
        service.analyze_sheet("T07 26"),
        ["101"],
    )
    captured: dict[str, object] = {}

    def start_detached(
        command: str,
        arguments: list[str],
        working_directory: str,
    ) -> tuple[bool, int]:
        captured.update(
            command=command,
            arguments=arguments,
            working_directory=working_directory,
        )
        return True, 4321

    monkeypatch.setattr(
        "app.rpa_expense.launcher.QProcess.startDetached",
        start_detached,
    )
    result = RpaExpenseBatLauncher(settings).launch(prepared)

    assert result.process_id == 4321
    assert captured["arguments"][-1] == str(prepared.selection_path)
    assert prepared.selection_path.name == "rpa_input_selection.json"
    assert captured["working_directory"] == str(bat.parent.resolve())
