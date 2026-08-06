from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from app.services.excel.models import ExcelRunStatus
from app.services.excel.payment_sync import (
    SUMMARY_HEADERS,
    PaymentSyncError,
    PaymentSyncService,
)


class _ImmediateStabilityChecker:
    def wait(self, _path: str | Path) -> None:
        return None


def _save_bk(path: Path, *, month: int, include_data: bool = True) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"T{month:02d} 26"
    headers = [
        "SQT", "Số Container", "Cước biển", "Cước bộ đóng hàng",
        "Nâng vỏ", "Hạ Hàng", "Nâng Hàng", "Hạ vỏ", "Cước VTN",
        "Lưu cont", "Quá tải", "VS + D/O", "Hóa đơn VS",
        "LÀM LỆNH", "SỬA CHỮA", *SUMMARY_HEADERS,
    ]
    for column, value in enumerate(headers, 1):
        sheet.cell(1, column).value = value
    if include_data:
        values = [
            601, "CONT0000001", 7_000_000, 3_000_000, 100_000,
            200_000, 300_000, 400_000, 5_000_000, 600_000, 700_000,
            800_000, None, 150_000, 900_000,
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(2, column).value = value
    workbook.save(path)
    workbook.close()


def _save_payment_templates(path: Path, *, month: int = 5) -> None:
    workbook = Workbook()
    hp = workbook.active
    hp.title = f"T{month:02d} 26 HP"
    nam = workbook.create_sheet(f"T{month:02d} 26 NAM")
    profiles = (
        (hp, ("QT", "SỐ CONT", "HẠ HÀNG", "Số HD", "Ghi chú", "Date cập nhật")),
        (
            nam,
            (
                "QT", "SỐ CONT", "NÂNG VỎ", "NÂNG HÀNG", "HẠ VỎ",
                "VS + D/O", "LÀM LỆNH", "Lưu Cont", "Sửa chữa Cont",
                "QUÁ TẢI", "Date cập nhật", "Số HD", "Ghi chú",
            ),
        ),
    )
    for sheet, headers in profiles:
        sheet.merge_cells("A5:F5")
        sheet["A5"] = f"BẢNG KÊ THÁNG {month:02d}/26"
        for column, value in enumerate(headers, 1):
            sheet.cell(7, column).value = value
        sheet["A8"] = 500
        sheet["B8"] = f"OLD-{sheet.title[-3:].strip()}"
        money_columns = (3,) if sheet is hp else tuple(range(3, 11))
        for column in money_columns:
            letter = sheet.cell(1, column).column_letter
            sheet.cell(10, column).value = f"=SUM({letter}8:{letter}9)"
        sheet["A10"] = "TỔNG TIỀN"
        sheet["A12"] = "Người duyệt"
        sheet["A15"] = "Số tiền thanh toán"
        sheet["D16"] = "Khu vực thanh toán giữ nguyên"
        sheet.freeze_panes = "A8"
        sheet.auto_filter.ref = "A7:M9"
    hp["C8"] = 999
    hp["D8"] = "HD-HP"
    hp["E8"] = "Ghi chú HP"
    hp["C8"].fill = PatternFill("solid", fgColor="FFF2CC")
    workbook.save(path)
    workbook.close()


def _service(bk: Path, payment: Path, runtime_dir: Path) -> PaymentSyncService:
    return PaymentSyncService(
        bk_path=bk,
        payment_path=payment,
        temp_dir=runtime_dir / "Temp",
        backup_dir=runtime_dir / "Backup",
        stability_checker=_ImmediateStabilityChecker(),
    )


def test_payment_sync_creates_both_missing_sheets_from_matching_profiles(
    tmp_path: Path,
) -> None:
    bk = tmp_path / "BK 2026.xlsx"
    payment = tmp_path / "Thanh toan 2026.xlsx"
    runtime_dir = tmp_path / "Excel"
    _save_bk(bk, month=6)
    _save_payment_templates(payment, month=5)
    service = _service(bk, payment, runtime_dir)

    plan = service.analyze(source_sheet_name="T06 26")

    assert plan.targets["HP"].sheet_name == "T06 26 HP"
    assert plan.targets["NAM"].sheet_name == "T06 26 NAM"
    assert plan.targets["HP"].template_sheet == "T05 26 HP"
    assert plan.targets["NAM"].template_sheet == "T05 26 NAM"
    assert plan.new_count == 2

    result = service.apply(plan, {})

    assert result.status is ExcelRunStatus.SUCCEEDED
    assert result.inserted_rows == 2
    assert result.backup_path == runtime_dir / "Backup" / "Thanh toan 2026_latest.xlsx"
    workbook = load_workbook(payment, data_only=False)
    try:
        assert "T06 26" not in workbook.sheetnames
        hp = workbook["T06 26 HP"]
        nam = workbook["T06 26 NAM"]
        assert (hp["A8"].value, hp["B8"].value, hp["C8"].value) == (
            601, "CONT0000001", 200_000,
        )
        assert hp["D8"].value is None
        assert hp["E8"].value is None
        assert hp["F8"].value is not None
        assert hp["F8"].number_format == "dd/mm/yyyy hh:mm:ss"
        assert [nam.cell(8, column).value for column in range(1, 11)] == [
            601, "CONT0000001", 100_000, 300_000, 400_000, 800_000,
            150_000, 600_000, 900_000, 700_000,
        ]
        assert hp["D16"].value == "Khu vực thanh toán giữ nguyên"
        assert nam["D16"].value == "Khu vực thanh toán giữ nguyên"
        assert nam["K8"].value is not None
        assert hp.max_column < 100
        assert nam.max_column < 100
    finally:
        workbook.close()

    second_plan = service.analyze(source_sheet_name="T06 26")
    assert second_plan.new_count == 0
    assert second_plan.unchanged_count == 2


def test_missing_previous_profile_template_stops_before_write(tmp_path: Path) -> None:
    bk = tmp_path / "BK 2026.xlsx"
    payment = tmp_path / "Thanh toan 2026.xlsx"
    _save_bk(bk, month=5)
    _save_payment_templates(payment, month=6)
    service = _service(bk, payment, tmp_path / "Excel")
    original = payment.read_bytes()

    with pytest.raises(PaymentSyncError, match="Không tìm thấy sheet HP mẫu"):
        service.analyze(source_sheet_name="T05 26")

    assert payment.read_bytes() == original


def test_empty_month_still_creates_both_sheets_after_confirmation(tmp_path: Path) -> None:
    bk = tmp_path / "BK 2026.xlsx"
    payment = tmp_path / "Thanh toan 2026.xlsx"
    _save_bk(bk, month=6, include_data=False)
    _save_payment_templates(payment, month=5)
    service = _service(bk, payment, tmp_path / "Excel")

    plan = service.analyze(source_sheet_name="T06 26")
    assert plan.requires_user_input
    result = service.apply(plan, {})
    assert result.inserted_rows == 0
    workbook = load_workbook(payment, read_only=True)
    try:
        assert "T06 26 HP" in workbook.sheetnames
        assert "T06 26 NAM" in workbook.sheetnames
    finally:
        workbook.close()


@pytest.mark.parametrize(("existing_type", "missing_type"), (("HP", "NAM"), ("NAM", "HP")))
def test_only_the_missing_target_sheet_is_created(
    tmp_path: Path, existing_type: str, missing_type: str
) -> None:
    bk = tmp_path / "BK 2026.xlsx"
    payment = tmp_path / "Thanh toan 2026.xlsx"
    _save_bk(bk, month=6)
    _save_payment_templates(payment, month=5)
    workbook = load_workbook(payment)
    current = workbook.copy_worksheet(workbook[f"T05 26 {existing_type}"])
    current.title = f"T06 26 {existing_type}"
    workbook.save(payment)
    workbook.close()
    service = _service(bk, payment, tmp_path / "Excel")

    plan = service.analyze(source_sheet_name="T06 26")

    assert not plan.targets[existing_type].sheet_to_create
    assert plan.targets[missing_type].sheet_to_create
    result = service.apply(plan, {})
    assert not result.target_results[existing_type].sheet_created
    assert result.target_results[missing_type].sheet_created
